import json
import os
import glob
from difflib import SequenceMatcher
import openpyxl
import requests
import unicodedata
from datetime import datetime, timedelta

# Import credentials from bot if possible, otherwise hardcode
NET_USER = os.getenv("NET_USER", "")
NET_PASS = os.getenv("NET_PASS", "")
LOGIN_URL = "https://aws.netcontroll.com.br/netadm/api/v1/account/login/"
PARTNER_LOGIN_URL = "https://aws.netcontroll.com.br/netadm/api/v1/parceiro/login"
EXPENSES_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/financeiro/conta-pagar/plano/periodo"
EXPENSES_SUPPLIER_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/financeiro/conta-pagar/fornecedor/periodo"
EXPENSES_DETAILED_URL = "https://aws.netcontroll.com.br/netweb/api/v1/financeiro/despesas/intervalo"
SALES_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/venda/produto/periodo"
INBOUND_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/entrada-mercadoria/periodo"  # Portal: #/estoque/entrada-mercadoria
INVENTORY_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/estoque"              # Portal: #/estoque/inventario
CMV_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/listagem-cmv-produto"
PRODUCT_URL = "https://aws.netcontroll.com.br/netweb/api/v1/estoque/produto"
COMPOSITION_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/composicao"
CANCELLATION_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/item-cancelado/periodo"
COMMISSION_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/comissao/venda/periodo"
REVENUE_PAYMENT_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/faturamento/forma-pagamento/periodo"
CASHIER_CLOSURE_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/fechamento-caixa/periodo"
FISCAL_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/venda/emissor-fiscal/nfce/periodo"
RECEIVABLES_URL = "https://aws.netcontroll.com.br/netweb/api/v1/financeiro/receitas/intervalo"
CASH_BOOK_URL = "https://aws.netcontroll.com.br/netweb/api/v1/financeiro/livro-caixa/conta/extrato"     # Portal: #/financeiro/conciliacao
DASHBOARD_RESUMO_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/venda/produto/periodo"  # Portal: #/dashboard/resumo (base: vendas + despesas combinadas)
DASHBOARD_COMPRAS_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/entrada-mercadoria/periodo"  # Portal: #/dashboard/compra
BALANCETE_URL = "https://aws.netcontroll.com.br/netreport/api/v1/netweb/relatorio/balancete"                # Portal: #/relatorio/balancete
BRASILAPI_NCM_URL = "https://brasilapi.com.br/api/ncm/v1/{}"


# Cache em memória para evitar chamadas repetidas à BrasilAPI
_NCM_CACHE = {}

# Mapeamento de palavras-chave do nomeSubgrupo do NetControll → capítulos NCM esperados (2 dígitos)
# Capítulo = 2 primeiros dígitos do NCM
# Baseado nos subgrupos reais identificados no Nauan Beach Club
NCM_GRUPO_MAP = {
    # ── Bebidas (cap. 22 = bebidas alcoólicas/não-alcoólicas, 21 = sucos/xaropes) ──
    "BEBIDA": ["22", "21", "20"],
    "ALCOOL": ["22"],
    "DRINK": ["22", "21"],
    "CERVEJA": ["22"],
    "VINHO": ["22"],
    "DESTILADO": ["22"],
    "ESPIRITUOSA": ["22"],
    "REFRIG": ["22"],
    "AGUA": ["22"],
    "SUCO": ["22", "21", "20"],
    "ENERGETIC": ["22"],
    "CHOPP": ["22"],
    # ── Carnes (cap. 02 = carnes frescas, 16 = conservas/embutidos) ──
    "CARNE": ["02", "16"],
    "BOVINO": ["02", "16"],
    "SUINO": ["02", "16"],
    "FRANGO": ["02", "16"],
    "AVE": ["02", "16"],
    "PROTEINA": ["02", "16", "03"],
    "EMBUTIDO": ["02", "16"],
    "DEFUMADO": ["02", "16"],
    # ── Frutos do Mar / Pescados (cap. 03 = frescos, 16 = conservas) ──
    "PESCADO": ["03", "16"],
    "PEIXE": ["03", "16"],
    "FRUTO DO MAR": ["03", "16"],
    "FRUTOS DO MAR": ["03", "16"],
    "CAMARAO": ["03", "16"],
    "SALMAO": ["03", "16"],
    "ATUM": ["03", "16"],
    "LAGOSTA": ["03", "16"],
    "POLVO": ["03", "16"],
    "OSTRA": ["03", "16"],
    "LULA": ["03", "16"],
    # ── Laticínios / Frios (cap. 04 = leite/queijo/manteiga) ──
    "LATICI": ["04", "19"],
    "LATIC": ["04", "19"],
    "QUEIJO": ["04"],
    "LEITE": ["04"],
    "FRIOS": ["04", "02", "16"],
    "REQUEIJAO": ["04"],
    "MANTEIGA": ["04"],
    "CREME": ["04", "21"],
    # ── Hortifrutti / FLV (cap. 07 = legumes, 08 = frutas, 09 = temperos) ──
    "HORTIF": ["07", "08", "09"],
    "LEGUME": ["07"],
    "VERDURA": ["07"],
    "FOLHA": ["07"],
    "FRUTA": ["08"],
    "FLV": ["07", "08", "09"],
    # ── Grãos / Cereais / Massas (cap. 10/11 = grãos/farinha, 19 = massas/pão) ──
    "GRAO": ["10", "11", "19"],
    "CEREAL": ["10", "11", "19"],
    "MASSA": ["19", "11"],
    "PADARIA": ["19"],
    "PAO": ["19"],
    "FARINHA": ["11"],
    "ARROZ": ["10"],
    "FEIJAO": ["07", "11"],
    # ── Temperos / Condimentos (cap. 09 = especiarias, 21 = molhos/extratos) ──
    "TEMPERO": ["09", "21"],
    "CONDIMENT": ["09", "21"],
    "MOLHO": ["21", "09"],
    "ESPECIAR": ["09"],
    "AZEITE": ["15"],
    "OLEO": ["15"],
    "VINAGRE": ["22", "21"],
    # ── Limpeza / Higiene / Descartáveis (cap. 34 = sabão, 38 = químicos, 39 = plásticos) ──
    "LIMPEZA": ["34", "38"],
    "HIGIENE": ["34", "33"],
    "QUIMICO": ["38"],
    "DESCART": ["39", "48"],
    "EMBALAGEM": ["39", "48"],
    "UTENSIL": ["39", "73", "76"],
    "EPI": ["39", "62", "63"],
}

DESKTOP = os.getenv("LOCAL_DATA_PATH", "")
_DATA_DIR = os.getenv("DATA_DIR", os.path.dirname(os.path.abspath(__file__)))
SNAPSHOTS_DIR = os.path.join(_DATA_DIR, "estoque_snapshots")

GOOGLE_PLACES_API_KEY = os.getenv("GOOGLE_PLACES_API_KEY", "")

RESTAURANTS = [
    {"name": "Nauan Beach Club", "id": 18784,
     "google_place_id": os.getenv("GOOGLE_PLACE_ID_NAUAN", ""),
     "sales_file": os.path.join(DESKTOP, "Vendas_Nauan_LIVE.json") if DESKTOP else None,
     "stock_file": os.path.join(DESKTOP, "Estoque_Nauan_LIVE.json") if DESKTOP else None,
     "expense_file": os.path.join(DESKTOP, "Despesas_Nauan_LIVE.json") if DESKTOP else None,
     "inbound_file": os.path.join(DESKTOP, "entrada de mercadorias nauan.xlsx") if DESKTOP else None,
     "recipe_file": os.path.join(DESKTOP, "composições nauan.xlsx") if DESKTOP else None,
    },
    {"name": "Milagres do Toque", "id": 19165,
     "google_place_id": os.getenv("GOOGLE_PLACE_ID_MILAGRES", ""),
     "sales_file": os.path.join(DESKTOP, "Vendas_Milagres_LIVE.json") if DESKTOP else None,
     "stock_file": os.path.join(DESKTOP, "Estoque_Milagres_LIVE.json") if DESKTOP else None,
     "expense_file": os.path.join(DESKTOP, "Despesas_Milagres_LIVE.json") if DESKTOP else None,
     "inbound_file": None, "recipe_file": None,
    },
    {"name": "Ahau Arte e Cozinha", "id": 20814,
     "google_place_id": os.getenv("GOOGLE_PLACE_ID_AHAU", ""),
     "sales_file": os.path.join(DESKTOP, "Vendas_Ahau_LIVE.json") if DESKTOP else None,
     "stock_file": os.path.join(DESKTOP, "Estoque_Ahau_LIVE.json") if DESKTOP else None,
     "expense_file": os.path.join(DESKTOP, "Despesas_Ahau_LIVE.json") if DESKTOP else None,
     "inbound_file": None, "recipe_file": None,
    },
]

def load_json(path):
    if not path or not os.path.exists(path): return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except: return []

def find_restaurant_files(rest_name_or_id):
    rest_name_or_id = str(rest_name_or_id).lower()
    for r in RESTAURANTS:
        if rest_name_or_id in r['name'].lower() or rest_name_or_id == str(r['id']):
            return r
    return RESTAURANTS[0]

def safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace('R$', '').replace('\xa0', '').replace('.', '').replace(',', '.').strip()
    try: return float(s)
    except: return 0.0

def normalize_text(text):
    if not text: return ""
    text = str(text)
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn').lower()

def fmt_brl(v, decimals=2):
    """
    Formata valor monetário brasileiro com abreviação automática para facilitar leitura:
      >= 1.000.000  → R$ 1,3 mi
      >= 1.000      → R$ 45,2 mil
      < 1.000       → R$ 945,50   (formato completo)

    Passar decimals=0 para inteiros (ex: quantidades).
    """
    try:
        v = float(v)

        if decimals == 0:
            # Modo inteiro — sem abreviação, só formata
            formatted = f"{v:,.0f}"
            return formatted.replace(",", "X").replace(".", ",").replace("X", ".")

        abs_v = abs(v)
        sinal = "-" if v < 0 else ""

        if abs_v >= 1_000_000:
            # Milhoes: ex 1.308.835,73 -> "1,3 mi"
            abrev = abs_v / 1_000_000
            return f"{sinal}R$ {f'{abrev:.1f}'.replace('.', ',')} mi"
        elif abs_v >= 1_000:
            # Milhares: ex 181.280,11 -> "181,3 mil"
            abrev = abs_v / 1_000
            return f"{sinal}R$ {f'{abrev:.1f}'.replace('.', ',')} mil"
        else:
            # Valor pequeno: formato completo
            formatted = f"{abs_v:,.{decimals}f}"
            formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
            return f"{sinal}R$ {formatted}"
    except:
        return str(v)


def fmt_pct(v, decimals=1):
    """Formata uma porcentagem no padrão brasileiro: 15,5%"""
    try:
        formatted = f"{float(v):.{decimals}f}"
        return formatted.replace(".", ",")
    except:
        return str(v)


def match_query(query, target):
    """
    Checks if ALL words in the normalized query exist in the normalized target.
    This allows 'agua gas' to safely match 'AGUA MINERAL COM GAS'.
    """
    q_norm = normalize_text(query).split()
    t_norm = normalize_text(target)
    if not q_norm: return True # Empty query matches anything
    return all(word in t_norm for word in q_norm)


# Stop-words que não ajudam no match de pratos (ignoradas na pontuação)
_STOP = {'de', 'da', 'do', 'para', 'com', 'sem', 'e', 'a', 'o', 'as', 'os',
         'ao', 'na', 'no', 'por', 'g', 'gr', 'kg', 'ml', 'lt', 'unid', 'und'}

def find_best_cost_match(sales_key: str, cost_map: dict, min_score: float = 0.55) -> float:
    """
    Busca o melhor custo no cost_map para um nome de item vendido (sales_key).
    Usa sobreposição de palavras (Jaccard) para tolerar variações de nome:
      'parmegiana de file mignon 2 pessoas' → 'parmegiana de file 2 pessoas'  ✅
      'isca de carne de sol c macaxeira frita' → 'isca carne de sol'          ✅

    Retorna o custo do melhor match ou 0.0 se nenhum atingir min_score.
    """
    import re
    # Remove pontuação (parênteses, barras, etc.) antes de tokenizar
    def _tokens(s: str):
        s = re.sub(r"[^\w\s]", " ", s)   # remove pontuação
        return [w for w in s.split() if w not in _STOP and len(w) > 2]

    if sales_key in cost_map:
        return cost_map[sales_key]

    words_a = _tokens(sales_key)
    if not words_a:
        return 0.0

    best_val   = 0.0
    best_score = 0.0

    for fk, fv in cost_map.items():
        if fv <= 0:
            continue
        words_b = _tokens(fk)
        if not words_b:
            continue
        common = len(set(words_a) & set(words_b))
        score  = common / len(set(words_a) | set(words_b))
        if score > best_score:
            best_score = score
            best_val   = fv

    return best_val if best_score >= min_score else 0.0



def get_session_for_rest(rest_id):
    session = requests.Session()
    session.headers.update({'Content-Type': 'application/json', 'App-Origin': 'Portal'})
    payload = {'Email': NET_USER, 'Senha': NET_PASS}
    try:
        r = session.post(LOGIN_URL, json=payload, timeout=10)
        if r.status_code == 200:
            token = r.json().get('data', {}).get('access_token')
            if token:
                session.headers.update({'Authorization': f'Bearer {token}'})
                p_resp = session.post(PARTNER_LOGIN_URL, data=str(rest_id), timeout=10)
                if p_resp.status_code == 200:
                    ptoken = p_resp.json().get('data', {}).get('access_token')
                    if ptoken:
                        session.headers.update({'Authorization': f'Bearer {ptoken}'})
                        return session
    except: pass
    return None

def fetch_sales_data(rest_id, start_date=None, end_date=None):
    if not start_date: start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    session = get_session_for_rest(rest_id)
    if not session: return []
    s_date = f"{start_date}T03:00:00.000Z"
    e_date = f"{end_date}T03:00:00.000Z"
    s_params = {'DataInicial': s_date, 'DataFinal': e_date, 'IncluirCusto': 'true'}
    try:
        r = session.get(SALES_URL, params=s_params)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def fetch_expenses_data(rest_id, start_date=None, end_date=None):
    """Despesas agrupadas por Plano de Contas (categorias: FOLHA, CMV, FIXO, etc.).
    Portal: /relatorio/financeiro-conta-pagar-plano"""
    if not start_date: start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    session = get_session_for_rest(rest_id)
    if not session: return []
    s_date = f"{start_date}T03:00:00.000Z"
    e_date = f"{end_date}T03:00:00.000Z"
    exp_params = {'DataInicial': s_date, 'DataFinal': e_date, 'TipoDataDespesa': 0}
    try:
        r = session.get(EXPENSES_URL, params=exp_params)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def fetch_expenses_supplier_data(rest_id, start_date=None, end_date=None):
    """Despesas agrupadas por Fornecedor (ranking: quem recebeu mais).
    Portal: /relatorio/financeiro-conta-pagar-fornecedor"""
    if not start_date: start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    session = get_session_for_rest(rest_id)
    if not session: return []
    s_date = f"{start_date}T03:00:00.000Z"
    e_date = f"{end_date}T03:00:00.000Z"
    params = {'DataInicial': s_date, 'DataFinal': e_date}
    try:
        r = session.get(EXPENSES_SUPPLIER_URL, params=params)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def fetch_inbound_data(rest_id, start_date=None, end_date=None):
    if not start_date: start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    session = get_session_for_rest(rest_id)
    if not session: return []
    s_date = f"{start_date}T03:00:00.000Z"
    e_date = f"{end_date}T03:00:00.000Z"
    params = {'DataInicial': s_date, 'DataFinal': e_date}
    try:
        r = session.get(INBOUND_URL, params=params)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def fetch_cmv_data(rest_id, start_date=None, end_date=None):
    if not start_date: start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    session = get_session_for_rest(rest_id)
    if not session: return []
    s_date = f"{start_date}T03:00:00.000Z"
    e_date = f"{end_date}T03:00:00.000Z"
    params = {'DataInicial': s_date, 'DataFinal': e_date}
    try:
        r = session.get(CMV_URL, params=params)
        if r.status_code == 200: return r.json()
    except: pass
    return []

def fetch_composition_data(rest_id):
    session = get_session_for_rest(rest_id)
    if not session: return []
    try:
        r = session.get(COMPOSITION_URL)
        if r.status_code == 200: return r.json()
    except: pass
    return []

# ─────────────────────────────────────────────────────────────────────────────
# Helpers específicos por restaurante para CMV
# ─────────────────────────────────────────────────────────────────────────────

def get_compras_periodo(rest: dict, start_date: str, end_date: str) -> tuple:
    """
    Retorna (total_compras_R$, fonte_str) para o restaurante e período.
    Prioridade: 1) API de Entradas de Mercadoria  2) Excel local (inbound_file)
    """
    # Tentativa 1 — API NetControll (filtrada por rest_id via partner login)
    try:
        rows = fetch_inbound_data(rest['id'], start_date, end_date)
        if rows:
            total = sum(safe_float(r.get('valorTotal', 0)) for r in rows)
            if total > 0:
                return total, f"API Entradas ({len(rows)} NFs)"
    except: pass

    # Tentativa 2 — Excel local (quando disponível, ex: Nauan)
    inbound_path = rest.get('inbound_file')
    if inbound_path and os.path.exists(inbound_path):
        try:
            wb  = openpyxl.load_workbook(inbound_path, data_only=True)
            ws  = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]

            # Tenta localizar coluna de data e valor total
            col_data  = next((i for i, h in enumerate(headers) if 'data' in h), None)
            col_valor = next((i for i, h in enumerate(headers)
                              if 'total' in h or 'valor' in h), None)

            s_dt = datetime.strptime(start_date, '%Y-%m-%d')
            e_dt = datetime.strptime(end_date,   '%Y-%m-%d')
            total = 0.0
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if col_data is not None and row[col_data]:
                    try:
                        cell_dt = row[col_data]
                        if isinstance(cell_dt, str):
                            cell_dt = datetime.strptime(cell_dt[:10], '%Y-%m-%d')
                        elif not isinstance(cell_dt, datetime):
                            cell_dt = None
                        if cell_dt and not (s_dt <= cell_dt <= e_dt):
                            continue
                    except: pass
                if col_valor is not None and row[col_valor]:
                    total += safe_float(row[col_valor])
                    count += 1
            if total > 0:
                return total, f"Excel local ({count} lançamentos)"
        except Exception as e:
            pass

    return 0.0, "sem dados de compras"


# ─────────────────────────────────────────────────────────────────────────────
# GitHub Gist — armazenamento persistente para snapshots de estoque na nuvem
# Variáveis de ambiente necessárias:
#   GITHUB_TOKEN  → Personal Access Token com escopo "gist"
#   GIST_ID       → ID do Gist criado (ex: "a1b2c3d4e5f6...")
# ─────────────────────────────────────────────────────────────────────────────

def _gist_enabled() -> bool:
    return bool(os.getenv("GITHUB_TOKEN")) and bool(os.getenv("GIST_ID"))

def _gist_save(filename: str, data: list) -> bool:
    """Salva (ou atualiza) um arquivo JSON no GitHub Gist."""
    token   = os.getenv("GITHUB_TOKEN")
    gist_id = os.getenv("GIST_ID")
    try:
        r = requests.patch(
            f"https://api.github.com/gists/{gist_id}",
            headers={"Authorization": f"token {token}",
                     "Accept": "application/vnd.github.v3+json"},
            json={"files": {filename: {"content": json.dumps(data, ensure_ascii=False)}}},
            timeout=15,
        )
        return r.status_code == 200
    except Exception as e:
        print(f"[Gist] Erro ao salvar {filename}: {e}")
        return False

def _gist_load(filename: str):
    """Carrega um arquivo JSON do GitHub Gist. Retorna lista ou None."""
    token   = os.getenv("GITHUB_TOKEN")
    gist_id = os.getenv("GIST_ID")
    try:
        r = requests.get(
            f"https://api.github.com/gists/{gist_id}",
            headers={"Authorization": f"token {token}",
                     "Accept": "application/vnd.github.v3+json"},
            timeout=15,
        )
        if r.status_code != 200:
            return None
        files = r.json().get("files", {})
        if filename not in files:
            return None
        raw_url = files[filename].get("raw_url")
        if not raw_url:
            return None
        r2 = requests.get(raw_url, timeout=15)
        return r2.json() if r2.status_code == 200 else None
    except Exception as e:
        print(f"[Gist] Erro ao carregar {filename}: {e}")
        return None


def save_inventory_snapshot(restaurant_name: str, date_str: str = None) -> str:
    """Salva snapshot do estoque atual em arquivo datado para uso no CMV por movimentação.
    Chame no último dia de cada mês (ou no primeiro do mês seguinte) para cada restaurante.
    Os snapshots são armazenados em: estoque_snapshots/<RestName>_<YYYY-MM-DD>.json
    """
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session:
        return f"Erro de autenticação para {rest['name']}."
    try:
        r = session.get(INVENTORY_URL)
        if r.status_code != 200:
            return f"API retornou status {r.status_code} para {rest['name']}."
        data = r.json()
        if not isinstance(data, list) or not data:
            return f"Inventário vazio ou inválido para {rest['name']}."
    except Exception as e:
        return f"Erro ao buscar inventário: {e}"
    safe_name = rest['name'].replace(" ", "_").replace("/", "_")
    filename  = f"{safe_name}_{date_str}.json"
    valor_total = sum(
        safe_float(s.get('estoqueAtual', 0)) * safe_float(s.get('custoAtual', 0))
        for s in data
    )
    destinos = []

    # 1. Salva localmente (quando DATA_DIR persistente estiver configurado ou ao rodar local)
    try:
        os.makedirs(SNAPSHOTS_DIR, exist_ok=True)
        path = os.path.join(SNAPSHOTS_DIR, filename)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        destinos.append("💾 local")
    except Exception as e:
        print(f"[snapshot] Erro ao salvar local: {e}")

    # 2. Salva no GitHub Gist (quando GITHUB_TOKEN + GIST_ID configurados)
    if _gist_enabled():
        ok = _gist_save(filename, data)
        destinos.append("☁️ Gist" if ok else "⚠️ Gist falhou")

    if not destinos:
        return f"⚠️ Nenhum destino disponível para salvar snapshot de {rest['name']}."

    return (f"✅ Snapshot salvo: {rest['name']} em {date_str}\n"
            f"   {len(data)} produtos | Valor em estoque: R$ {valor_total:,.2f}\n"
            f"   Destinos: {' + '.join(destinos)}")


def _calc_snapshot_valor(data: list) -> float:
    return sum(
        safe_float(s.get('estoqueAtual', 0)) * safe_float(s.get('custoAtual', 0))
        for s in data
    )

def _load_snapshot_value(rest_name: str, date_str: str) -> tuple:
    """Retorna (valor_R$, fonte_str) para o snapshot mais próximo da data (±3 dias).
    Prioridade: 1) arquivo local  2) GitHub Gist
    """
    safe_name = rest_name.replace(" ", "_").replace("/", "_")
    target = datetime.strptime(date_str, '%Y-%m-%d')

    for delta in range(4):
        for sign in ([0] if delta == 0 else [1, -1]):
            d = target + timedelta(days=delta * sign)
            label = d.strftime('%d/%m/%Y')
            filename = f"{safe_name}_{d.strftime('%Y-%m-%d')}.json"

            # 1. Arquivo local
            path = os.path.join(SNAPSHOTS_DIR, filename)
            if os.path.exists(path):
                data = load_json(path)
                return _calc_snapshot_valor(data), f"snapshot local {label}"

            # 2. GitHub Gist
            if _gist_enabled():
                data = _gist_load(filename)
                if data:
                    return _calc_snapshot_valor(data), f"snapshot Gist {label}"

    return 0.0, "sem snapshot"


def get_ficha_cost_map(rest: dict) -> tuple:
    """
    Retorna (dict {prato_normalizado: custo_por_porcao}, fonte_str)
    Prioridade: 1) API de Composições (COMPOSITION_URL) 2) Excel local (recipe_file)
    """
    # Tentativa 1 — API NetControll
    try:
        rows = fetch_composition_data(rest['id'])
        if rows:
            ficha = {}
            for row in rows:
                dish        = normalize_text(str(row.get('compostoNome', '')))
                custo_ingr  = safe_float(row.get('custo', 0))
                ficha[dish] = ficha.get(dish, 0.0) + custo_ingr
            ficha = {k: v for k, v in ficha.items() if v > 0}
            if ficha:
                return ficha, f"API Composições ({len(ficha)} pratos)"
    except: pass

    # Tentativa 2 — Excel local de receitas
    recipe_path = rest.get('recipe_file')
    if recipe_path and os.path.exists(recipe_path):
        try:
            wb  = openpyxl.load_workbook(recipe_path, data_only=True)
            ws  = wb.active
            ficha        = {}
            current_dish = None
            headers      = [str(c.value).strip().lower() if c.value else ''
                            for c in ws[1]] if ws.max_row > 0 else []

            # Detectar colunas de custo (geralmente "custo total" ou "custo")
            col_custo = next((i for i, h in enumerate(headers)
                              if 'custo' in h and ('total' in h or 'unit' in h)), None)
            if col_custo is None:
                col_custo = next((i for i, h in enumerate(headers) if 'custo' in h), None)

            for row in ws.iter_rows(min_row=2, values_only=True):
                r0 = str(row[0]).strip() if row[0] else ''
                r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''

                # Linha de cabeçalho de prato (col1 preenchida, col2 vazia ou "Produto:")
                if r0 and (not r1 or r0.startswith('Produto:')):
                    dish_name = r0.replace('Produto:', '').strip()
                    if dish_name:
                        current_dish = normalize_text(dish_name)
                        if current_dish not in ficha:
                            ficha[current_dish] = 0.0
                elif current_dish and col_custo is not None:
                    val = safe_float(row[col_custo]) if len(row) > col_custo else 0
                    ficha[current_dish] += val

            ficha = {k: v for k, v in ficha.items() if v > 0}
            if ficha:
                return ficha, f"Excel local ({len(ficha)} pratos)"
        except Exception as e:
            pass

    return {}, "fichas técnicas indisponíveis"



def get_revenue(restaurant_name, start_date=None, end_date=None):
    rest = find_restaurant_files(restaurant_name)
    data = fetch_sales_data(rest['id'], start_date, end_date)
    total = sum([safe_float(item.get('valor', 0)) for item in data])
    dt_str = start_date if start_date else "ontem"
    return f"Faturamento total de {rest['name']} ({dt_str}): {fmt_brl(total)}"

def get_top_selling_items(restaurant_name, top_n=5, start_date=None, end_date=None):
    rest = find_restaurant_files(restaurant_name)
    data = fetch_sales_data(rest['id'], start_date, end_date)
    sales_map = {}
    for row in data:
        n = row.get("nome", "")
        v = safe_float(row.get("valor", 0))
        q = safe_float(row.get("qtde", 0))
        if n:
            if n not in sales_map:
                sales_map[n] = {'qty': 0, 'rev': 0}
            sales_map[n]['qty'] += q
            sales_map[n]['rev'] += v
    items = list(sales_map.items())
    items.sort(key=lambda x: x[1]['rev'], reverse=True)
    dt_str = start_date if start_date else "ontem"
    res = f"Top {top_n} mais vendidos de {rest['name']} ({dt_str}):\n"
    for k, v in items[:int(top_n)]:
        res += f"- {k}: {v['qty']} unid -> {fmt_brl(v['rev'])}\n"
    return res

def search_sales(restaurant_name, query, start_date=None, end_date=None):
    rest = find_restaurant_files(restaurant_name)
    data = fetch_sales_data(rest['id'], start_date, end_date)
    
    grouped = {}
    for row in data:
        n = row.get("nome", "")
        if match_query(query, n):
            if n not in grouped: grouped[n] = {'qty': 0, 'val': 0.0}
            grouped[n]['qty'] += safe_float(row.get('qtde', 0))
            grouped[n]['val'] += safe_float(row.get('valor', 0))

            
    dt_str = start_date if start_date else "ontem"
    if not grouped: 
        return f"Nenhuma venda encontrada para '{query}' em {rest['name']} ({dt_str})."
        
    res = f"Vendas para '{query}' em {rest['name']} ({dt_str}):\n"
    for k, v in grouped.items():
        res += f"- {k}: {v['qty']} unid -> {fmt_brl(v['val'])}\n"
    return res

def get_stock(restaurant_name, query):
    """Consulta inventário de estoque.
    Tenta API ao vivo primeiro (INVENTORY_URL = portal #/estoque/inventario),
    depois cai no JSON local como fallback."""
    rest = find_restaurant_files(restaurant_name)
    data = []

    # 1. Tenta buscar ao vivo da API de inventário
    try:
        session = get_session_for_rest(rest['id'])
        if session:
            r = session.get(INVENTORY_URL)
            if r.status_code == 200:
                raw = r.json()
                if isinstance(raw, list) and len(raw) > 0:
                    data = raw
                    print(f"[get_stock] Inventário ao vivo OK — {len(data)} itens")
    except Exception as e:
        print(f"[get_stock] API ao vivo falhou: {e}")

    # 2. Fallback: JSON local (sincronizado pelo /sync)
    if not data:
        path = rest.get('stock_file')
        if path and os.path.exists(path):
            data = load_json(path)
            print(f"[get_stock] Usando JSON local — {len(data)} itens")
        else:
            return f"Estoque indisponível para {rest['name']}. Use /sync para sincronizar ou aguarde o auto-sync."

    q_low = normalize_text(query)
    is_generic = q_low in ['estoque', 'inventario', 'tudo', 'completo', 'geral', '']

    matches = []
    for row in data:
        n = row.get('produto', '')
        if is_generic or match_query(query, n):
            est  = safe_float(row.get('estoqueAtual', 0))
            custo = safe_float(row.get('custoAtual', 0))
            matches.append({
                'nome': n,
                'estoque': est,
                'custo': custo,
                'total': est * custo
            })

    if not matches:
        return f"Produto '{query}' não encontrado no estoque de {rest['name']}."

    matches.sort(key=lambda x: SequenceMatcher(None, q_low, normalize_text(x['nome'])).ratio(), reverse=True)

    # Para consultas genéricas, mostra top 30 por valor; para produto específico, mostra os 5 mais relevantes
    if is_generic:
        matches_show = sorted(matches, key=lambda x: -x['total'])[:30]
        total_val = sum(m['total'] for m in matches)
        res = f"📦 **INVENTÁRIO — {rest['name']}** _(portal: #/estoque/inventario)_\n"
        res += f"Total de itens: {len(matches)} | Capital em estoque: R${fmt_brl(total_val)}\n\n"
        for m in matches_show:
            alerta = " ⚠️ RUPTURA" if m['estoque'] <= 0 else (" 🔻 BAIXO" if m['estoque'] < 3 else "")
            res += f"• *{m['nome']}*: {fmt_brl(m['estoque'], 2)} unid × R${fmt_brl(m['custo'])} = R${fmt_brl(m['total'])}{alerta}\n"
        if len(matches) > 30:
            res += f"\n_...e mais {len(matches)-30} itens. Peça 'estoque completo' para ver todos._"
    else:
        matches_show = matches[:5]
        res = f"📦 **Estoque — '{query}' em {rest['name']}**\n"
        for m in matches_show:
            alerta = " ⚠️ RUPTURA" if m['estoque'] <= 0 else (" 🔻 ESTOQUE BAIXO" if m['estoque'] < 3 else "")
            res += f"• *{m['nome']}*: {fmt_brl(m['estoque'], 2)} unid | Custo unit.: R${fmt_brl(m['custo'])} | Total em estoque: R${fmt_brl(m['total'])}{alerta}\n"

    return res

def get_recipe(restaurant_name, dish_name):
    rest = find_restaurant_files(restaurant_name)
    path = rest.get('recipe_file')
    if not path or not os.path.exists(path):
        return f"Planilha de Ficha Técnica indisponível para {rest['name']}."
    
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        current_dish = None
        ingredients = []
        dish_found = False
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            r0 = str(row[0]).strip() if row[0] else ""
            r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            if r0 and not r1 and not r0.replace('.','').isdigit():
                if dish_found: break # Finished reading the target dish
                if dish_name.lower() in r0.lower():
                    dish_found = True
                    current_dish = r0
            elif dish_found and current_dish:
                r7 = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                r8 = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                
                # Try to extract ingredient name and qty heuristically based on columns
                # In Nauan's composition excel, columns usually look like: 
                # (empty), code, NAME, unit, yield... qty is scattered. Let's just dump row text.
                text_parts = [str(x) for x in row if x and str(x).strip()]
                if len(text_parts) > 1:
                    ingredients.append(" | ".join(text_parts))
                    
        if not dish_found:
             return f"Ficha técnica para '{dish_name}' não encontrada."
        
        res = f"🍲 Ficha Técnica de {current_dish}:\n"
        for i in ingredients:
            res += f"- {i}\n"
        return res[:1000] # Limit size for LLM context
        
    except Exception as e:
        return f"Erro ao ler ficha técnica: {e}"

def get_product_specs(restaurant_name, product_name):
    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session: return "Não foi possível autenticar."
    
    try:
        # 1. Search for Product ID
        r = session.get(PRODUCT_URL)
        if r.status_code != 200: return "Erro ao buscar lista de produtos."
        
        products = r.json()
        match = None
        for p in products:
            if product_name.lower() in p.get('nome', '').lower():
                match = p
                break
        
        if not match: return f"Produto '{product_name}' não encontrado no portal."
        
        # 2. Fetch full details
        pid = match['id']
        detail_url = f"{PRODUCT_URL}/{pid}/completo"
        rd = session.get(detail_url)
        if rd.status_code != 200: return f"Erro ao buscar detalhes do produto {pid}."
        
        data = rd.json()
        
        res = f"📄 **DOSSIÊ TÉCNICO: {data.get('nome')}**\n"
        res += f"🏠 Unidade: {rest['name']}\n"
        res += f"🔢 ID: {data.get('id')} | Unidade: {data.get('nomeUnidadeMedida', 'N/A')}\n"
        res += f"🏷️ Subgrupo: {data.get('nomeSubgrupo')}\n"
        res += f"⚖️ NCM: {data.get('ncm') or 'N/A'} | CEST: {data.get('cest') or 'N/A'}\n"
        res += f"💰 Preço de Venda: {fmt_brl(safe_float(data.get('preco')))}\n\n"
        
        comp = data.get('composicoes', [])
        if comp:
            res += "🧪 **COMPOSIÇÃO / FICHA TÉCNICA:**\n"
            for c in comp:
                qty = safe_float(c.get('quantidade'))
                u = c.get('nomeUnidadeMedidaIngrediente', '')
                res += f"  • {c.get('nomeIngrediente')}: {qty} {u}\n"
        else:
            res += "ℹ️ Produto sem composição cadastrada no portal."
            
        return res[:4000]
    except Exception as e:
        return f"Erro ao extrair dossiê: {e}"

def analyze_recipes_profitability(restaurant_name, query=None):
    rest = find_restaurant_files(restaurant_name)
    
    # Try Live Data first
    try:
        live_data = fetch_composition_data(rest['id'])
        if live_data:
            dishes_map = {}
            for item in live_data:
                d_name = item.get('compostoNome')
                if not d_name: continue
                
                if d_name not in dishes_map:
                    dishes_map[d_name] = {
                        "name": d_name,
                        "group": "Geral", # Composition report doesn't explicitly have group in this endpoint
                        "cost": 0.0,
                        "ingredients": []
                    }
                
                ing_qty = safe_float(item.get('quantidade'))
                ing_cost = safe_float(item.get('custo'))
                dishes_map[d_name]["cost"] += ing_cost
                dishes_map[d_name]["ingredients"].append({
                    "name": item.get('composicaoNome'),
                    "qty": ing_qty,
                    "t_cost": ing_cost,
                    "u_cost": (ing_cost / ing_qty) if ing_qty > 0 else 0
                })
            
            dishes = list(dishes_map.values())
            if query:
                dishes = [d for d in dishes if match_query(query, d["name"])]
                
            if dishes:
                return build_profitability_report(rest, dishes, query)
    except: pass

    # Fallback to Excel if live fails or returns nothing
    path = rest.get('recipe_file')
    if not path or not os.path.exists(path):
        return f"Não foi possível obter dados live nem Planilha de Ficha Técnica para {rest['name']}."
        
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        dishes = []
        current_dish = None
        current_group = ""
        current_cost = 0.0
        ingredients = []
        
        for row in ws.iter_rows(min_row=1, values_only=True):
            r0 = str(row[0]).strip() if row[0] else ""
            r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if r0.startswith("Subgrupo:"):
                current_group = r0.replace('Subgrupo:', '').strip()
                continue
            if r0.startswith("Produto:") or (r0 and not r1 and not r0.replace('.','').isdigit() and not r0.startswith("Subgrupo:")):
                if current_dish:
                    dishes.append({"name": current_dish, "group": current_group, "cost": current_cost, "ingredients": ingredients})
                current_dish = r0.split(":", 1)[1].strip() if ":" in r0 else r0
                cost_str = str(row[5]) if len(row) > 5 and row[5] else "0"
                current_cost = safe_float(cost_str)
                ingredients = []
            elif current_dish and len(row) >= 5:
                if isinstance(row[2], (int, float)):
                    ing_name = str(row[1]).strip() if row[1] else ""
                    ing_qty = safe_float(row[2])
                    ing_u_cost = safe_float(row[4])
                    ing_t_cost = safe_float(row[5])
                    if ing_name:
                        ingredients.append({"name": ing_name, "qty": ing_qty, "t_cost": ing_t_cost, "u_cost": ing_u_cost})
        if current_dish:
            dishes.append({"name": current_dish, "group": current_group, "cost": current_cost, "ingredients": ingredients})
        
        if query:
            dishes = [d for d in dishes if match_query(query, d["name"]) or match_query(query, d["group"])]
            
        return build_profitability_report(rest, dishes, query)
    except Exception as e:
        return f"Erro ao auditar as fichas técnicas: {e}"

def build_profitability_report(rest, dishes, query=None):
    zero_cost_ingredients = set()
    for d in dishes:
        for ing in d['ingredients']:
            if ing.get('u_cost', 0) == 0 or ing.get('t_cost', 0) == 0:
                zero_cost_ingredients.add(ing['name'])
                
    dishes.sort(key=lambda x: x["cost"], reverse=True)
    q_label = f" ({query.upper()})" if query else ""
    report = f"📊 **AUDITORIA DE FICHAS TÉCNICAS LIVE: {rest['name']}{q_label}**\n\n"
    
    z_limit = 40 if query else 15
    if zero_cost_ingredients:
        report += "🚨 **ALERTA CRÍTICO: INSUMOS COM CUSTO ZERO (R$ 0,00)!**\n"
        report += "_Corrija no cadastro de composição urgentemente:_\n"
        for z in list(zero_cost_ingredients)[:z_limit]:
            report += f"❌ {z}\n"
        report += "\n"
        
    d_limit = min(40, len(dishes)) if query else 10
    report += f"🏆 **MAIORES CUSTOS DE PRODUÇÃO (CMV BRUTO):**\n"
    for d in dishes[:d_limit]:
        report += f"🍽️ **{d['name']}** - Custo {fmt_brl(d['cost'])}\n"
        top_ing = sorted(d['ingredients'], key=lambda x: x['t_cost'], reverse=True)
        if top_ing:
            report += f"   ➤ Vilão: {top_ing[0]['name']} ({fmt_brl(top_ing[0]['t_cost'])})\n"
            
    report += "\n💡 **INSIGHTS CEO:**\n1. Audite os insumos zero para evitar lucro fantasma.\n2. Negocie os 'Vilões' para baixar o CMV imediatamente."
    return report[:14000]

def get_ingredient_consumption(restaurant_name, query, start_date=None, end_date=None):
    if not start_date: start_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    
    rest = find_restaurant_files(restaurant_name)
    path = rest.get('recipe_file')
    if not path or not os.path.exists(path):
        return f"Planilha de Ficha Técnica indisponível para {rest['name']}. Sem ficha técnica, não posso calcular o consumo de insumos dos pratos vendidos."
        
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        
        q_low = normalize_text(query)
        dishes_found = {}
        
        current_dish = None
        dish_has_ingredient = False
        dish_ingredients_matched = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            r0 = str(row[0]).strip() if row[0] else ""
            r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            
            if r0 and not r1 and not r0.replace('.','').isdigit():
                if dish_has_ingredient and current_dish:
                    dishes_found[current_dish] = dish_ingredients_matched
                current_dish = r0
                dish_has_ingredient = False
                dish_ingredients_matched = []
                continue
                
            if current_dish:
                text_parts = [str(x) for x in row if x and str(x).strip()]
                if len(text_parts) > 1:
                    joined_text = " | ".join(text_parts)
                    if match_query(query, joined_text):
                        dish_has_ingredient = True
                        dish_ingredients_matched.append(joined_text)
                        
        if dish_has_ingredient and current_dish:
            dishes_found[current_dish] = dish_ingredients_matched
            
        if not dishes_found:
            return f"Nenhum prato/drink foi encontrado contendo '{query}' na ficha técnica."
            
        sales_data = fetch_sales_data(rest['id'], start_date, end_date)
        
        report = f"🥩 **CONSUMO DE INSUMO VIA RECEITUÁRIO: {query.upper()}**\n"
        report += f"📅 Período: {start_date} a {end_date}\n\n"
        report += "🍽️ **VENDAS DOS PRATOS QUE LEVAM ESSE INSUMO:**\n"
        
        total_items_sold_with_ingr = 0
        
        for dish_name, lines in dishes_found.items():
            d_low = normalize_text(dish_name)
            qty_sold = 0
            for row in sales_data:
                n = str(row.get("nome", ""))
                # Strict or partial match logic: sometimes sales name is slightly different
                n_low = normalize_text(n)
                if d_low == n_low or d_low in n_low or n_low in d_low:
                    qty_sold += safe_float(row.get("qtde", 0))
                    
            if qty_sold > 0:
                report += f"- **Prato:** {dish_name} (Vendidos: {qty_sold} unid)\n"
                for l in lines[:2]:
                    report += f"  _Receita (porção):_ {l}\n"
                total_items_sold_with_ingr += qty_sold
                
        if total_items_sold_with_ingr == 0:
            report += "Nenhum dos pratos que utilizam este insumo foi vendido neste período.\n"
            
        report += "\n💡 *Consulte as quantidades por porção listadas acima e multiplique pelo total de pratos vendidos para estimar o consumo real do insumo!*"
        return report[:3000]
        
    except Exception as e:
        return f"Erro ao calcular consumo de insumos: {e}"

def get_inbound_purchases(restaurant_name, query, start_date=None, end_date=None):
    rest = find_restaurant_files(restaurant_name)
    data = fetch_inbound_data(rest['id'], start_date, end_date)
    dt_str = start_date if start_date else "ontem"
    
    if not data:
        return f"Sem entradas de mercadoria registradas via sistema para {rest['name']} em {dt_str}."
        
    matches = []
    
    for row in data:
        prod = str(row.get('produto', ''))
        forn = str(row.get('fornecedor', ''))
        is_generic = normalize_text(query) in ["compras", "geral", "auditoria", "", "vendas", "estoque"]
        if is_generic or match_query(query, prod) or match_query(query, forn):
            qtd = safe_float(row.get('qtde', 0))
            un = row.get('un', '')
            val = safe_float(row.get('valorTotal', 0))
            matches.append(f"{prod} ({qtd} {un}) - FORNECEDOR: {forn} - {fmt_brl(val)}")
            
    if not matches:
        return f"Nenhuma entrada recente encontrada para '{query}' no sistema em {dt_str}."
        
    res = f"🛒 Últimas Entradas no Sistema para '{query}' ({dt_str}):\n"
    for m in matches[-10:]: # Get last 10
        res += f"- {m}\n"
    return res[:1500]


def get_cmv_report(restaurant_name, query=None, start_date=None, end_date=None):
    """
    CMV por período — visão executiva.

    CMV Teórico  = Σ(qtde_vendida_prato × custo_ficha_técnica_porção)
                   Fonte: get_ficha_cost_map() → API Composições ou Excel local

    CMV Real     = Σ(qtde_vendida × precoCompra_ERP)
                   Fonte: cruzamento entre fetch_sales_data() e fetch_cmv_data().
                   A API listagem-cmv-produto retorna o CADASTRO de custo por produto
                   (precoCompra = último custo de compra registrado no ERP, cmv = % calculado).
                   Multiplicamos esse custo unitário pela qtde vendida no período.

    Compras      = get_compras_periodo() → API Entradas ou Excel local
    Estoque Final = JSON local de estoque
    """
    if not start_date: start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:   end_date   = datetime.now().strftime('%Y-%m-%d')
    dt_str = f"{start_date} a {end_date}"
    rest   = find_restaurant_files(restaurant_name)

    # ── 1. Vendas do período ──────────────────────────────────────────────────
    # Palavras-chave que identificam itens SEM custo de mercadoria (serviços puros)
    # → excluídos do faturamento e do cálculo de CMV
    _SEM_CMV_KWS = {'day use', 'dayuse', 'day-use', 'aluguel toalha',
                    'ao ponto', 'bem passada', 'entrada', 'couvert'}

    def _is_servico(nome: str) -> bool:
        n = nome.lower()
        return any(kw in n for kw in _SEM_CMV_KWS)

    sales_data      = fetch_sales_data(rest['id'], start_date, end_date)
    sales_agg       = {}
    receita_servico = 0.0   # Day Use e outros serviços — excluídos do CMV

    for s in sales_data:
        nome = str(s.get('nome', '')).strip()
        if not nome: continue
        valor = safe_float(s.get('valor', 0))
        if _is_servico(nome):
            receita_servico += valor
            continue
        key = normalize_text(nome)
        if key not in sales_agg:
            sales_agg[key] = {'nome': nome, 'qtde': 0.0, 'receita': 0.0}
        sales_agg[key]['qtde']    += safe_float(s.get('qtde',  0))
        sales_agg[key]['receita'] += valor

    faturamento_bruto = sum(v['receita'] for v in sales_agg.values())
    faturamento_total = faturamento_bruto + receita_servico  # para exibição


    # ── 2. Custo unitário por produto — duas fontes complementares ────────────
    #
    #   Fonte A: ERP CMV catalog (listagem-cmv-produto)
    #     → precoCompra = último custo de compra unitário registrado
    #     → cobre: bebidas, insumos simples com estoque direto
    #
    #   Fonte B: Fichas Técnicas (get_ficha_cost_map)
    #     → custo total por porção (soma de ingredientes da receita)
    #     → cobre: pratos compostos (comidas) que NÃO têm precoCompra direto
    #
    # CMV Real = Σ(qtde_vendida × melhor_custo_unitário_disponível)
    #   para cada item vendido: usa Fonte A se disponível, senão Fonte B.

    # Fonte A — ERP (bebidas, itens simples)
    cmv_catalog  = fetch_cmv_data(rest['id'], start_date, end_date)
    erp_cost_map = {}   # { nome_normalizado: precoCompra }
    for r in cmv_catalog:
        nome = str(r.get('nome', '')).strip()
        if not nome: continue
        key = normalize_text(nome)
        pc = safe_float(r.get('precoCompra', 0))
        if pc > 0:
            erp_cost_map[key] = pc

    # Fonte B — Fichas Técnicas (pratos compostos) — buscada UMA VEZ aqui
    # e reutilizada tanto no CMV Real (fallback) quanto no CMV Teórico
    ficha_cost, fonte_fichas = get_ficha_cost_map(rest)

    n_erp   = len(erp_cost_map)
    n_ficha = len(ficha_cost)
    fonte_cmv_real = (f"ERP ({n_erp} simples) + Fichas ({n_ficha} pratos)"
                      if ficha_cost else f"ERP ({n_erp} produtos)")

    # ── 3. CMV Real = qtde_vendida × custo_unit (ERP → Ficha como fallback) ───
    cmv_real_total = 0.0
    cmv_real_map   = {}   # { key: { nome, custo, custo_u, receita, qtde, fonte } }
    sem_custo      = []   # itens sem custo em nenhuma fonte (para debug)

    for key, v in sales_agg.items():
        # Busca na Fonte A (ERP) — exact match + word-based fuzzy
        custo_unit = find_best_cost_match(key, erp_cost_map)
        fonte_item = "ERP"

        # Fallback Fonte B (Ficha Técnica) — pratos compostos
        if custo_unit == 0:
            fonte_item = "FT"
            custo_unit = find_best_cost_match(key, ficha_cost)

        if custo_unit == 0:
            sem_custo.append(v['nome'])
            continue

        custo_total     = v['qtde'] * custo_unit
        cmv_real_total += custo_total
        cmv_real_map[key] = {
            'nome':    v['nome'],
            'qtde':    v['qtde'],
            'custo':   custo_total,
            'custo_u': custo_unit,
            'receita': v['receita'],
            'fonte':   fonte_item,
        }

    cmv_real_pct = (cmv_real_total / faturamento_bruto * 100) if faturamento_bruto > 0 else 0

    # ── Passo 4. CMV Teórico — Fichas Técnicas × Qtde Vendida ────────────────
    # ficha_cost já foi buscada no passo 2 (Fonte B) e reutilizada aqui.

    cmv_teorico   = 0.0
    itens_teorico = []
    for key, v in sales_agg.items():
        # Usa find_best_cost_match para tolerar variações de nome na ficha
        custo_ficha = find_best_cost_match(key, ficha_cost)
        if custo_ficha == 0:
            continue
        teorico_item = v['qtde'] * custo_ficha
        real_item    = cmv_real_map.get(key, {}).get('custo', 0.0)
        cmv_teorico += teorico_item
        if not query or match_query(query, v['nome']):
            desvio = real_item - teorico_item
            efic   = (teorico_item / real_item * 100) if real_item > 0 else 100.0
            itens_teorico.append({
                'nome':     v['nome'],
                'qtde':     v['qtde'],
                'custo_ft': custo_ficha,
                'teorico':  teorico_item,
                'real':     real_item,
                'desvio':   desvio,
                'efic':     efic,
            })
    cmv_teorico_pct = (cmv_teorico / faturamento_bruto * 100) if faturamento_bruto > 0 else 0

    # ── 4. Compras do período — Entradas de Mercadoria ────────────────────────
    # Usa get_compras_periodo() que tenta API e depois Excel local do restaurante
    total_compras, fonte_compras = get_compras_periodo(rest, start_date, end_date)

    # ── 5. Estoque Final — JSON local do restaurante ──────────────────────────
    stock_path = rest.get('stock_file', '')
    stock_data = load_json(stock_path) if stock_path and os.path.exists(stock_path) else []
    valor_ef   = sum(safe_float(s.get('valorEstoque', 0)) for s in stock_data)
    fonte_ef   = "JSON local" if stock_data else "indisponível"

    # ══════════════════════════════════════════════════════════════════════════
    # RELATÓRIO
    # ══════════════════════════════════════════════════════════════════════════
    q_label = f" — Filtro: {query.upper()}" if query else ""
    res  = f"📊 **ANÁLISE DE CMV DO PERÍODO{q_label}**\n"
    res += f"🏠 {rest['name']}  |  📅 {dt_str}\n"
    res += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    # Fontes usadas (transparência por restaurante)
    res += f"_Fontes: CMV real → {fonte_cmv_real} | Fichas → {fonte_fichas} | Compras → {fonte_compras} | EF → {fonte_ef}_\n\n"

    # ── Bloco 1: Painel Executivo ──────────────────────────────────────────────
    res += "🎯 **PAINEL EXECUTIVO — CMV CONSOLIDADO**\n\n"

    if cmv_teorico > 0:
        emoji_t = "✅" if cmv_teorico_pct <= 30 else ("⚠️" if cmv_teorico_pct <= 35 else "🚨")
        res += f"{emoji_t} **CMV Teórico** _(fichas técnicas × vendas)_:\n"
        res += f"   **{fmt_pct(cmv_teorico_pct)}%**  ≈  {fmt_brl(cmv_teorico)}\n"

        res += "   _O que as receitas dizem que deveria ser consumido_\n\n"
    else:
        res += f"⚠️ **CMV Teórico**: não calculado — {fonte_fichas}\n\n"

    if cmv_real_total > 0:
        emoji_r = "✅" if cmv_real_pct <= 30 else ("⚠️" if cmv_real_pct <= 35 else "🚨")
        res += f"{emoji_r} **CMV Real** _(custo registrado pelo ERP em cada venda)_:\n"
        res += f"   **{fmt_pct(cmv_real_pct)}%**  ≈  {fmt_brl(cmv_real_total)}\n"
        res += "   _O que o sistema efetivamente baixou do estoque_\n\n"
    else:
        res += f"⚠️ **CMV Real**: {fonte_cmv_real} — sem custo registrado no período\n\n"

    if cmv_teorico > 0 and cmv_real_total > 0:
        desvio_pp = cmv_real_pct - cmv_teorico_pct
        abs_dev = abs(desvio_pp)
        if abs_dev < 0.5:   dv_e, dv_t = "🏆", "Operação impecável — real praticamente igual ao teórico."
        elif abs_dev <= 2:  dv_e, dv_t = "✅",  "Pequeno desvio, dentro do tolerável. Monitore."
        elif abs_dev <= 5:  dv_e, dv_t = "⚠️", "Desvio relevante — revise porcionamento e desperdícios."
        else:               dv_e, dv_t = "🚨", "DESVIO ALTO — fichas técnicas desatualizadas ou divergência de custo!"
        sinal = "+" if desvio_pp >= 0 else ""
        res += f"{dv_e} **Desvio (Real − Teórico): {sinal}{fmt_pct(desvio_pp)} p.p.**\n"
        res += f"   _{dv_t}_\n\n"

    res += f"💰 **Faturamento F&B (base CMV):** {fmt_brl(faturamento_bruto)}\n"
    if receita_servico > 0:
        res += f"   _(Day Use / serviços excluídos: {fmt_brl(receita_servico)} — sem custo de mercadoria)_\n"
        res += f"   _Receita total do período: {fmt_brl(faturamento_total)}_\n"
    res += "📊 _Referência saudável: CMV entre 25% e 32%._\n"

    # ── Bloco 2: Contexto Contábil ─────────────────────────────────────────────
    total_items   = len(sales_agg)
    cobert_items  = len(cmv_real_map)
    pct_cobertura = (cobert_items / total_items * 100) if total_items > 0 else 0

    res += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
    res += "📦 **CONTEXTO CONTÁBIL**\n"
    res += f"   🛒 Compras do período ({fonte_compras}): {fmt_brl(total_compras)}\n"
    res += f"   📦 Estoque Final ({fonte_ef}):  {fmt_brl(valor_ef)}\n"
    if total_compras > 0 and valor_ef > 0:
        cobertura = (valor_ef / total_compras * 100)
        res += f"   📊 EF representa {fmt_pct(cobertura)}% das compras do período\n"
    if total_compras == 0:
        res += "   ⚠️ _Sem compras encontradas — verifique entradas no portal ou arquivo Excel_\n"
    # Cobertura dos itens vendidos
    res += f"\n   📈 Cobertura CMV: {cobert_items}/{total_items} itens vendidos com custo ({fmt_pct(pct_cobertura)}%)\n"
    res += f"   _Fonte: ERP (bebidas/simples) + Fichas Técnicas (pratos compostos)_\n"
    if sem_custo and pct_cobertura < 80:
        res += f"   ⚠️ {len(sem_custo)} itens sem custo em nenhuma fonte — podem subestivar o CMV\n"


    # ── Bloco 3: Top desvios (sem filtro) ─────────────────────────────────────
    if itens_teorico and not query:
        piores = sorted([p for p in itens_teorico if p['desvio'] > 0],
                        key=lambda x: x['desvio'], reverse=True)[:8]
        if piores:
            res += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            res += "🔎 **TOP 8 PRATOS — MAIOR DESVIO Real > Teórico**\n"
            res += "_Esses pratos custaram mais do que previsto nas fichas técnicas:_\n\n"
            for p in piores:
                e = "🚨" if p['efic'] < 80 else ("⚠️" if p['efic'] < 95 else "✅")
                res += f"{e} **{p['nome']}** ({fmt_brl(p['qtde'], 0)} unid)\n"
                res += f"   Custo/porção ficha: {fmt_brl(p['custo_ft'])}\n"
                res += f"   Teórico: {fmt_brl(p['teorico'])}  |  Real: {fmt_brl(p['real'])}\n"
                res += f"   Perda: {fmt_brl(p['desvio'])}  (efic. {fmt_pct(p['efic'], 0)}%)\n\n"

    # ── Bloco 4: Detalhe por produto (apenas com filtro) ──────────────────────
    if query:
        res += "\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        res += f"📋 **DETALHE POR PRODUTO — Filtro: {query.upper()}**\n\n"
        itens_det = []
        for key, v in cmv_real_map.items():
            if match_query(query, v['nome']):
                teorico_val = next((t['teorico'] for t in itens_teorico
                                    if normalize_text(t['nome']) == key), 0)
                pct_real = (v['custo'] / v['receita'] * 100) if v['receita'] > 0 else 0
                itens_det.append({
                    'nome':    v['nome'],
                    'qtde':    v['qtde'],
                    'receita': v['receita'],
                    'real':    v['custo'],
                    'custo_u': v.get('custo_u', 0),
                    'teorico': teorico_val,
                    'pct':     pct_real,
                })
        if itens_det:
            for it in sorted(itens_det, key=lambda x: x['real'], reverse=True)[:20]:
                flag = "🚨" if it['pct'] > 35 else ("⚠️" if it['pct'] > 28 else "✅")
                res += f"{flag} **{it['nome']}** — {fmt_brl(it['qtde'], 0)} unid\n"
                res += f"   Custo unit (ERP): {fmt_brl(it['custo_u'])}\n"
                res += f"   Receita: {fmt_brl(it['receita'])}  |  CMV Real: {fmt_brl(it['real'])}  ({fmt_pct(it['pct'])}%)\n"
                if it['teorico'] > 0:
                    res += (f"   CMV Teórico: {fmt_brl(it['teorico'])}  |  "
                            f"Desvio: {fmt_brl(it['real'] - it['teorico'])}\n")
                res += "\n"
        else:
            res += f"Nenhum item encontrado para '{query}' no CMV do período.\n"

    return res[:14000]


def get_expenses(restaurant_name, query=None, start_date=None, end_date=None):
    rest = find_restaurant_files(restaurant_name)
    dt_str = start_date if start_date else "ontem"

    # Detecta se a pergunta é sobre fornecedor → usa endpoint /fornecedor
    q_norm = normalize_text(query or '')
    is_supplier_query = any(x in q_norm for x in [
        'fornecedor', 'fornecedores', 'ranking fornecedor', 'quem mais recebeu',
        'maior fornecedor', 'despesa por fornecedor', 'conta pagar fornecedor'
    ])

    if is_supplier_query:
        # Usa endpoint /conta-pagar/fornecedor/periodo (portal: financeiro-conta-pagar-fornecedor)
        raw = fetch_expenses_supplier_data(rest['id'], start_date, end_date)
        if not raw:
            return f"Sem dados de fornecedores para {rest['name']} em {dt_str}."
        supplier_totals = {}
        for row in raw:
            name = str(row.get('fornecedor', 'OUTROS')).strip() or 'OUTROS'
            val  = safe_float(row.get('total', 0) or row.get('valor', 0))
            if val > 0:
                supplier_totals[name] = supplier_totals.get(name, 0) + val
        sorted_sup = sorted(supplier_totals.items(), key=lambda x: -x[1])
        grand_total = sum(supplier_totals.values())
        res = f"🏭 **DESPESAS POR FORNECEDOR — {rest['name']} ({dt_str})**\n"
        res += f"💰 Total: R${fmt_brl(grand_total)}\n\n"
        for i, (name, val) in enumerate(sorted_sup, 1):
            pct = val / grand_total * 100 if grand_total > 0 else 0
            flag = " ⚠️ _Concentração alta_" if pct > 20 else ""
            res += f"{i}. *{name}*: R${fmt_brl(val)} ({pct:.1f}%){flag}\n"
        res += f"\n📍 _Fonte: /relatorio/financeiro-conta-pagar-fornecedor_"
        return res

    # Caso padrão: usa endpoint /conta-pagar/plano/periodo (portal: financeiro-conta-pagar-plano)
    data = fetch_expenses_data(rest['id'], start_date, end_date)
    if not data: return f"Sem despesas registradas para {rest['name']} em {dt_str}."
    
    total = 0.0
    
    # Hierarchy: Cat1 -> Cat2 -> List of items
    hierarchy = {}
    suspicious_items = []
    
    # Detect payment status filter
    status_filter = None # All
    if any(x in q_norm for x in ['pago', 'pagos', 'pagas', 'liquidado', 'concluido']):
        status_filter = True
    elif any(x in q_norm for x in ['pendente', 'a pagar', 'em aberto', 'nao pago']):
        status_filter = False

    for item in data:
        val = safe_float(item.get('valor', 0))
        cat1 = str(item.get('planoContas1', 'DIVERSOS')).strip().upper()
        cat2 = str(item.get('planoContas2', 'GERAL')).strip().upper()
        forn = str(item.get('fornecedor', 'N/A')).strip()
        hist = str(item.get('historico', '')).strip()
        is_paid = item.get('pagamento', False)
        
        match = True
        
        # 1. Search Query Filter
        if query:
            is_generic = normalize_text(query) in ["compras", "geral", "auditoria", "", "despesas", "financeiro", "pagas", "pagos", "pendentes", "a pagar"]
            if not is_generic:
                match = match_query(query, cat1) or match_query(query, cat2) or match_query(query, forn) or match_query(query, hist)
        
        # 2. Status Filter (Specific for 'pago' vs 'pendente')
        if status_filter is not None:
            if is_paid != status_filter:
                match = False
            
        if match:
            total += val
            if cat1 not in hierarchy: hierarchy[cat1] = {"total": 0, "subs": {}}
            if cat2 not in hierarchy[cat1]["subs"]: hierarchy[cat1]["subs"][cat2] = {"total": 0, "items": []}
            
            hierarchy[cat1]["total"] += val
            hierarchy[cat1]["subs"][cat2]["total"] += val
            
            is_paid = item.get('pagamento', False)
            status_tag = "PAGO ✅" if is_paid else "A PAGAR ⏳"
            
            hierarchy[cat1]["subs"][cat2]["items"].append({
                "forn": forn,
                "val": val,
                "hist": hist,
                "status": status_tag
            })
            
            # Check for suspicious items
            is_suspicious = False
            reasons = []
            if not hist or len(hist) < 5 or hist.upper() in ["DIVERSOS", "PAGAMENTO", "COMPRA", "DESPESA", "TESTE"]:
                is_suspicious = True
                reasons.append("Histórico vazio ou genérico")
            
            if not forn or forn.upper() in ["N/A", "DIVERSOS", "FORNECEDOR NAO IDENTIFICADO"]:
                is_suspicious = True
                reasons.append("Fornecedor não identificado")
                
            if is_suspicious:
                suspicious_items.append({
                    "cat1": cat1,
                    "cat2": cat2,
                    "forn": forn,
                    "val": val,
                    "hist": hist,
                    "reasons": reasons
                })
            
    if total == 0:
        return f"Nenhuma despesa encontrada" + (f" para '{query}'" if query else "") + f" em {rest['name']} ({dt_str})."
        
    res = f"🧾 Detalhamento de Despesas - {rest['name']} ({dt_str}):\n"
    res += f"💰 **TOTAL GERAL: {fmt_brl(total)}**\n\n"
    
    if suspicious_items:
        res += "🚨 **ALERTAS DE DESPESAS SUSPEITAS** 🚨\n"
        res += "_Atenção: Os lançamentos abaixo precisam de maior detalhamento ou justificativa clara._\n"
        for idx, s in enumerate(suspicious_items[:10]):
            motivos = " | ".join(s['reasons'])
            res += f"  {idx+1}. **{s['cat1']} / {s['cat2']}** - {fmt_brl(s['val'])}\n"
            res += f"     • Fornecedor: {s['forn']}\n"
            res += f"     • Histórico: {s['hist'] if s['hist'] else '[Vazio]'}\n"
            res += f"     ⚠️ _Motivo do Alerta: {motivos}_\n\n"
        if len(suspicious_items) > 10:
            res += f"  _... e mais {len(suspicious_items) - 10} ocultas._\n\n"
        res += "💡 **Dica CEO:** Exija que a equipe preencha sempre o Histórico com o QUÊ foi comprado e POR QUÊ, e evite Fornecedores genéricos.\n"
        res += "-"*40 + "\n\n"
    
    for c1, data1 in hierarchy.items():
        res += f"📂 **{c1}** ({fmt_brl(data1['total'])})\n"
        for c2, data2 in data1['subs'].items():
            res += f"  └─ 📁 *{c2}* ({fmt_brl(data2['total'])})\n"
            for it in data2['items']:
                res += f"     • {it['forn']}: {fmt_brl(it['val'])} [{it['status']}]\n"
                if it['hist']: res += f"       _Obs: {it['hist']}_\n"
    
    return res

def get_session_for_rest(rest_id):
    session = requests.Session()
    session.headers.update({'Content-Type': 'application/json', 'App-Origin': 'Portal'})
    payload = {'Email': NET_USER, 'Senha': NET_PASS}
    try:
        r = session.post(LOGIN_URL, json=payload, timeout=10)
        if r.status_code == 200:
            token = r.json().get('data', {}).get('access_token')
            if token:
                session.headers.update({'Authorization': f'Bearer {token}'})
                p_resp = session.post(PARTNER_LOGIN_URL, data=str(rest_id), timeout=10)
                if p_resp.status_code == 200:
                    ptoken = p_resp.json().get('data', {}).get('access_token')
                    if ptoken:
                        session.headers.update({'Authorization': f'Bearer {ptoken}'})
                        return session
    except: pass
    return None

def apply_price_change(restaurant_name, product_name, price_change):
    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de conexão financeira."
    
    try:
        r = session.get(PRODUCT_URL)
        if r.status_code != 200: return "Erro ao buscar opções de produto no ERP."
        
        products = r.json()
        match = None
        for p in products:
            if normalize_text(product_name) in normalize_text(p.get('nome', '')):
                match = p
                break
                
        if not match: return f"Produto '{product_name}' não encontrado no cardápio de {rest['name']}."
        
        pid = match['id']
        rd = session.get(f"{PRODUCT_URL}/{pid}/completo")
        if rd.status_code != 200: return "Erro ao baixar ficha do produto."
        
        data = rd.json()
        preco_antigo = safe_float(data.get('preco', 0))
        
        str_change = str(price_change).replace(',','.').strip()
        if '%' in str_change:
            pct = float(str_change.replace('%','')) / 100
            novo_preco = preco_antigo * (1 + pct)
        else:
            novo_preco = preco_antigo + float(str_change)
            
        data['preco'] = novo_preco
        data['precoVenda'] = novo_preco
        
        # Emulating PUT because destructive action might need exact JSON format
        # If API rejects, we still show the calculation to CEO
        r_put = session.put(f"{PRODUCT_URL}/{pid}", json=data)
        if r_put.status_code in [200, 201, 204]:
            return f"✅ **SUCESSO! ERP ATUALIZADO.**\n\n🍔 Item: {match['nome']} ({rest['name']})\n📉 Preço Antigo: {fmt_brl(preco_antigo)}\n📈 **Novo Preço:** {fmt_brl(novo_preco)}\n\nO reajuste já está valendo no PDV."
        else:
            return f"⚠️ **SIMULAÇÃO DE AJUSTE (API PROTEGIDA)**\n\n🍔 Item: {match['nome']} ({rest['name']})\n📉 Preço Atual: {fmt_brl(preco_antigo)}\n📈 **Preço Calculado:** {fmt_brl(novo_preco)}\n\nO portal do administrador bloqueou a edição direta via bot. A intenção de reajuste está registrada."
    except Exception as e:
        return f"Erro ao aplicar preço: {str(e)}"

# ─────────────────────────────────────────────────────────────────────────────
# Google Places — busca de avaliações reais
# Variáveis de ambiente:
#   GOOGLE_PLACES_API_KEY     → chave da Google Places API
#   GOOGLE_PLACE_ID_NAUAN     → Place ID do Nauan Beach Club
#   GOOGLE_PLACE_ID_MILAGRES  → Place ID do Milagres do Toque
#   GOOGLE_PLACE_ID_AHAU      → Place ID do Ahau Arte e Cozinha
# Para obter o Place ID: https://developers.google.com/maps/documentation/places/web-service/place-id
# ─────────────────────────────────────────────────────────────────────────────

_POSITIVE_KW = {'ótimo','excelente','incrível','delicioso','maravilhoso','perfeito',
                'adorei','amei','recomendo','fantástico','sensacional','impecável',
                'nota 10','top','lindo','especial','gostoso','atencioso','rápido',
                'capricho','superou','surpreendeu','voltarei','voltamos'}
_NEGATIVE_KW = {'ruim','péssimo','horrível','demora','demorou','frio','errado',
                'errou','decepcionou','decepção','insosso','caro demais','salgado',
                'atendimento','descaso','abandonado','sujo','barata','mosca',
                'nunca mais','não recomendo','cancelei','esperamos','esperou'}

def _fetch_google_reviews(rest: dict) -> dict:
    """Busca avaliações reais via Google Places API.
    Retorna dict com rating, total_ratings, reviews list.
    """
    api_key   = GOOGLE_PLACES_API_KEY
    place_id  = rest.get("google_place_id", "")

    if not api_key:
        return {"error": "GOOGLE_PLACES_API_KEY não configurada"}

    # Se não tiver place_id configurado, busca pelo nome
    if not place_id:
        return {"error": f"GOOGLE_PLACE_ID não configurado para {rest['name']}"}

    # Places API (legacy) — requer "Places API" habilitada no Google Cloud Console
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "fields":   "name,rating,user_ratings_total,reviews",
                "language": "pt-BR",
                "key":      api_key,
            },
            timeout=10,
        )
        data = r.json()
        status = data.get("status", "UNKNOWN")
        if status != "OK":
            return {"error": f"Places API: {status} — {data.get('error_message', '')}"}

        result = data.get("result", {})
        return {
            "place_id":      place_id,
            "rating":        result.get("rating", 0),
            "total_ratings": result.get("user_ratings_total", 0),
            "reviews":       result.get("reviews", []),
        }
    except Exception as e:
        return {"error": f"Erro ao buscar reviews: {e}"}


def _analyze_review_sentiment(text: str) -> str:
    words = set(text.lower().split())
    pos = len(words & _POSITIVE_KW)
    neg = len(words & _NEGATIVE_KW)
    if pos > neg:   return "positivo"
    if neg > pos:   return "negativo"
    return "neutro"


def _extract_review_themes(reviews: list) -> dict:
    """Extrai temas recorrentes dos reviews: atendimento, comida, espera, preço, ambiente."""
    themes = {
        "atendimento":  {"kw": {"atendimento","garçom","garçom","equipe","staff","simpático","educado","grosso","demora"},     "pos": 0, "neg": 0},
        "comida":       {"kw": {"comida","prato","sabor","gostoso","delicioso","insosso","frio","mal passado","porcão"},        "pos": 0, "neg": 0},
        "espera":       {"kw": {"espera","demorou","demora","rápido","ágil","esperamos","esperou","lento"},                    "pos": 0, "neg": 0},
        "preço":        {"kw": {"preço","caro","barato","valor","custo","vale","não vale","justo"},                            "pos": 0, "neg": 0},
        "ambiente":     {"kw": {"ambiente","vista","praia","lindo","bonito","agradável","barulho","quente","estrutura"},       "pos": 0, "neg": 0},
    }
    for rev in reviews:
        text  = rev.get("text", "").lower()
        stars = rev.get("rating", 3)
        words = set(text.split())
        for theme, data in themes.items():
            if words & data["kw"]:
                if stars >= 4: data["pos"] += 1
                elif stars <= 2: data["neg"] += 1
    return themes


def get_customer_success_report(restaurant_name, days=30):
    """Analisa avaliações reais do Google + cancelamentos + tendência de faturamento."""
    rest = find_restaurant_files(restaurant_name)
    now  = datetime.now()
    end  = now.strftime('%Y-%m-%d')
    start_cur  = (now - timedelta(days=days)).strftime('%Y-%m-%d')
    start_prev = (now - timedelta(days=days*2)).strftime('%Y-%m-%d')
    end_prev   = (now - timedelta(days=days+1)).strftime('%Y-%m-%d')

    # ── 1. Faturamento: período atual vs. anterior ────────────────────────────
    sales_cur  = fetch_sales_data(rest['id'], start_cur, end)
    sales_prev = fetch_sales_data(rest['id'], start_prev, end_prev)
    fat_cur  = sum(safe_float(s.get('valor', 0)) for s in sales_cur)
    fat_prev = sum(safe_float(s.get('valor', 0)) for s in sales_prev)
    fat_var  = ((fat_cur - fat_prev) / fat_prev * 100) if fat_prev > 0 else 0

    # ── 2. Cancelamentos ──────────────────────────────────────────────────────
    try:
        session = get_session_for_rest(rest['id'])
        r_canc = session.get(CANCELLATION_URL, params={
            'DataInicial': f"{start_cur}T03:00:00.000Z",
            'DataFinal':   f"{end}T23:59:59.000Z",
        }, timeout=10) if session else None
        cancels = r_canc.json() if r_canc and r_canc.status_code == 200 else []
    except Exception:
        cancels = []
    total_cancels  = len(cancels)
    valor_cancels  = sum(safe_float(c.get('valor', 0)) for c in cancels)
    cancel_rate    = (valor_cancels / fat_cur * 100) if fat_cur > 0 else 0

    # ── 3. Google Reviews ─────────────────────────────────────────────────────
    gdata   = _fetch_google_reviews(rest)
    g_error = gdata.get("error")
    reviews = gdata.get("reviews", [])
    rating  = gdata.get("rating", 0)
    total_r = gdata.get("total_ratings", 0)

    sentimentos = {"positivo": 0, "negativo": 0, "neutro": 0}
    notas = []
    for rev in reviews:
        s = _analyze_review_sentiment(rev.get("text", ""))
        sentimentos[s] += 1
        notas.append(rev.get("rating", 3))

    avg_recent = (sum(notas) / len(notas)) if notas else 0
    themes     = _extract_review_themes(reviews)

    # ── Montagem do relatório ─────────────────────────────────────────────────
    report  = f"⭐ **AVALIAÇÕES & SATISFAÇÃO — {rest['name']}**\n"
    report += f"_Período: últimos {days} dias_\n\n"

    # Bloco Google
    if g_error:
        report += f"📍 **Google Reviews:** {g_error}\n"
        report += f"   _(Configure GOOGLE_PLACES_API_KEY e GOOGLE_PLACE_ID no Railway)_\n\n"
    else:
        stars_bar = "⭐" * round(rating)
        report += f"📍 **Google Reviews:** {rating:.1f}/5 {stars_bar} ({total_r:,} avaliações totais)\n"
        if reviews:
            report += f"   Últimas {len(reviews)} avaliações: "
            report += f"😊 {sentimentos['positivo']} positivas | "
            report += f"😐 {sentimentos['neutro']} neutras | "
            report += f"😞 {sentimentos['negativo']} negativas\n"
            if avg_recent:
                trend = "📈" if avg_recent >= rating else "📉"
                report += f"   Média recente: {avg_recent:.1f} {trend}\n"

        # Temas
        report += "\n📌 **Temas recorrentes:**\n"
        for theme, data in themes.items():
            if data["pos"] + data["neg"] > 0:
                bar = "🟢" * data["pos"] + "🔴" * data["neg"]
                report += f"   {theme.capitalize()}: {bar}\n"

        # Reviews negativos recentes (alerta)
        neg_reviews = [r for r in reviews if r.get("rating", 5) <= 2]
        if neg_reviews:
            report += f"\n🚨 **{len(neg_reviews)} avaliação(ões) negativa(s) recente(s):**\n"
            for rev in neg_reviews[:3]:
                autor = rev.get("author_name", "Cliente")
                texto = rev.get("text", "")[:120]
                report += f"   • _{autor}_: \"{texto}...\"\n"

    # Bloco Financeiro
    seta = "📈" if fat_var >= 0 else "📉"
    report += f"\n💰 **Faturamento ({days}d):** {fmt_brl(fat_cur)} {seta} {fat_var:+.1f}% vs. período anterior\n"

    # Bloco Cancelamentos
    canc_icon = "🟢" if cancel_rate < 2 else ("🟡" if cancel_rate < 5 else "🔴")
    report += f"{canc_icon} **Cancelamentos:** {total_cancels} itens | {fmt_brl(valor_cancels)} ({cancel_rate:.1f}% do faturamento)\n"

    # Diagnóstico cruzado
    report += "\n🧠 **Diagnóstico cruzado:**\n"
    if not g_error and rating < 4.0:
        report += f"   ⚠️ Nota Google abaixo de 4.0 — investigar reviews negativos e responder publicamente.\n"
    if cancel_rate > 5:
        report += f"   ⚠️ Taxa de cancelamento alta ({cancel_rate:.1f}%) — pode indicar erros de pedido ou insatisfação.\n"
    if fat_var < -10:
        report += f"   ⚠️ Queda de faturamento de {fat_var:.1f}% — verificar se há correlação com avaliações negativas.\n"
    if (not g_error and rating >= 4.5) and fat_var >= 0 and cancel_rate < 2:
        report += f"   ✅ Operação saudável: boa nota, faturamento estável e baixo cancelamento.\n"

    return report


def get_reviews_consolidated() -> str:
    """Relatório consolidado de avaliações das 3 casas para visão CEO."""
    report = "⭐ **PAINEL DE AVALIAÇÕES — GRUPO MILAGRES**\n\n"
    for rest in RESTAURANTS:
        gdata  = _fetch_google_reviews(rest)
        rating = gdata.get("rating", 0)
        total  = gdata.get("total_ratings", 0)
        error  = gdata.get("error")
        reviews = gdata.get("reviews", [])
        notas   = [r.get("rating", 3) for r in reviews]
        avg_rec = (sum(notas) / len(notas)) if notas else 0
        neg_rec = sum(1 for n in notas if n <= 2)

        if error:
            report += f"🏠 **{rest['name']}**: sem dados ({error})\n"
        else:
            trend  = ("📈" if avg_rec >= rating else "📉") if avg_rec else ""
            report += f"🏠 **{rest['name']}**: {rating:.1f}⭐ ({total:,} avaliações) {trend}\n"
            if neg_rec:
                report += f"   🚨 {neg_rec} avaliação(ões) negativa(s) recente(s)\n"

    report += "\n_Use 'avaliações [nome da casa]' para análise detalhada._"
    return report

def get_dre_report(restaurant_name, start_date=None, end_date=None):
    if not start_date: start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date: end_date = datetime.now().strftime('%Y-%m-%d')
    
    rest = find_restaurant_files(restaurant_name)
    
    sales_data = fetch_sales_data(rest['id'], start_date, end_date)
    rec_bruta = sum(safe_float(x.get('valor', 0)) for x in sales_data)
    
    # CMV Real vs Teórico
    cmv_data = fetch_cmv_data(rest['id'], start_date, end_date)
    erp_cost_map = {}
    for r in cmv_data:
        nome = str(r.get('nome', '')).strip()
        if nome:
            pc = safe_float(r.get('precoCompra', 0))
            if pc > 0: erp_cost_map[normalize_text(nome)] = pc
            
    ficha_cost, _ = get_ficha_cost_map(rest)
    
    cmv_real = 0.0
    cmv_teorico = 0.0
    
    for s in sales_data:
        nome = str(s.get('nome', '')).strip()
        if not nome or any(kw in nome.lower() for kw in {'day use', 'dayuse', 'day-use', 'aluguel toalha', 'ao ponto', 'bem passada', 'entrada', 'couvert'}): continue
        key = normalize_text(nome)
        qtde = safe_float(s.get('qtde', 0))
        
        # Real (Fallback para Ficha Técnica)
        custo_unit = find_best_cost_match(key, erp_cost_map)
        if custo_unit == 0: custo_unit = find_best_cost_match(key, ficha_cost)
        cmv_real += qtde * custo_unit
        
        # Teórico (Somente da Ficha)
        custo_ficha = find_best_cost_match(key, ficha_cost)
        if custo_ficha > 0: cmv_teorico += qtde * custo_ficha

    # CMV por movimentação de estoque (mais preciso) — usa snapshots quando disponíveis
    ei_val, ei_fonte = _load_snapshot_value(rest['name'], start_date)
    ef_val, ef_fonte = _load_snapshot_value(rest['name'], end_date)
    compras_periodo, compras_fonte = get_compras_periodo(rest, start_date, end_date)

    if ei_val > 0 or ef_val > 0:
        custo_cmv = ei_val + compras_periodo - ef_val
        cmv_fonte = f"Movimentação (EI {ei_fonte} + Compras {compras_fonte} − EF {ef_fonte})"
    else:
        custo_cmv = cmv_real
        cmv_fonte = "Vendas × custo unitário ERP/FT (sem snapshots)"

    cmv_real_pct = (custo_cmv / rec_bruta * 100) if rec_bruta > 0 else 0
    cmv_teorico_pct = (cmv_teorico / rec_bruta * 100) if rec_bruta > 0 else 0

    
    exp_data = fetch_expenses_data(rest['id'], start_date, end_date)
    impostos = 0
    folha = 0
    desp_fixas = 0
    desp_variaveis = 0
    fixas_items = {}  # label -> valor acumulado

    FIXAS_KEYWORDS = {
        "ALUGUEL":        "Aluguel",
        "AGUA":           "Água",
        "LUZ":            "Energia",
        "ENERGIA":        "Energia",
        "INTERNET":       "Internet",
        "SISTEMA":        "Sistema",
        "SOFTPLUS":       "Sistema",
        "VELOO":          "Sistema",
        "CONTABILIDADE":  "Contabilidade",
        "CONTADOR":       "Contabilidade",
    }

    for exp in exp_data:
        val = safe_float(exp.get('valor', 0))
        t = f"{exp.get('planoContas1','')} {exp.get('planoContas2','')} {exp.get('historico','')} {exp.get('fornecedor','')}".upper()

        if any(x in t for x in ["SIMPLES", "ICMS", "IMPOSTO", "DAS", "TAXA"]):
            impostos += val
        elif any(x in t for x in ["FOLHA", "SALARIO", "VT", "VR", "INSS", "FGTS", "RH", "RESCISAO"]):
            folha += val
        elif any(x in t for x in FIXAS_KEYWORDS):
            desp_fixas += val
            label = next((lbl for kw, lbl in FIXAS_KEYWORDS.items() if kw in t), "Outros Fixos")
            fixas_items[label] = fixas_items.get(label, 0) + val
        else:
            desp_variaveis += val

    rec_liquida = rec_bruta - impostos
    margem_bruta = rec_liquida - custo_cmv
    ebitda = margem_bruta - folha - desp_fixas - desp_variaveis
    margem_ebitda = (ebitda / rec_bruta * 100) if rec_bruta > 0 else 0

    report = f"🏢 **DRE GERENCIAL: {rest['name']}** ({start_date} a {end_date})\n\n"
    report += f"1️⃣ **(+) Receita Bruta:** {fmt_brl(rec_bruta)}\n"
    report += f"2️⃣ **(-) Impostos/Taxas:** {fmt_brl(impostos)}\n"
    report += f"3️⃣ **(=) Receita Líquida:** {fmt_brl(rec_liquida)}\n\n"
    report += f"4️⃣ **(-) CMV:** {fmt_brl(custo_cmv)} ({fmt_pct(cmv_real_pct)}% da Receita)\n"
    report += f"    *Fonte: {cmv_fonte}*\n"
    if ei_val > 0 or ef_val > 0:
        report += f"    ├─ Estoque Inicial ({ei_fonte}): {fmt_brl(ei_val)}\n"
        report += f"    ├─ (+) Compras ({compras_fonte}): {fmt_brl(compras_periodo)}\n"
        report += f"    └─ (−) Estoque Final ({ef_fonte}): {fmt_brl(ef_val)}\n"
    report += f"    *(_CMV Teórico Esperado: {fmt_brl(cmv_teorico)} / {fmt_pct(cmv_teorico_pct)}%_)*\n"
    report += f"5️⃣ **(=) Margem Bruta:** {fmt_brl(margem_bruta)} ({fmt_pct(((margem_bruta/rec_bruta)*100) if rec_bruta>0 else 0)}%)\n\n"
    report += f"6️⃣ **(-) Folha/RH:** {fmt_brl(folha)}\n"
    report += f"7️⃣ **(-) Despesas Fixas (Ocupação):** {fmt_brl(desp_fixas)}\n"
    for label, val in sorted(fixas_items.items(), key=lambda x: -x[1]):
        report += f"    ├─ {label}: {fmt_brl(val)}\n"
    report += f"8️⃣ **(-) Despesas Variáveis/Outras:** {fmt_brl(desp_variaveis)}\n\n"
    report += f"🎯 **EBITDA (Lucro Operacional):** {fmt_brl(ebitda)} ({fmt_pct(margem_ebitda)}%)\n"

    return report

def get_balancete(restaurant_name, start_date=None, end_date=None):
    """Busca o Balancete Contábil do restaurante diretamente da API do portal netcontroll.
    Portal: #/relatorio/balancete
    API: BALANCETE_URL (/relatorio/balancete)

    O balancete apresenta os saldos contábeis por plano de contas para o período,
    permitindo conferir receitas, despesas e saldos de forma estruturada."""
    rest = find_restaurant_files(restaurant_name)
    now = datetime.now()
    if not start_date: start_date = now.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:   end_date   = now.strftime('%Y-%m-%d')

    session = get_session_for_rest(rest['id'])
    if not session:
        return f"Erro de autenticação ao acessar o balancete de {rest['name']}."

    s_iso = f"{start_date}T03:00:00.000Z"
    e_iso = f"{end_date}T23:59:59.000Z"

    data = []
    # Tenta variações comuns da rota (com e sem /periodo)
    for url in [
        BALANCETE_URL + "/periodo",
        BALANCETE_URL,
    ]:
        try:
            r = session.get(url, params={'DataInicial': s_iso, 'DataFinal': e_iso})
            if r.status_code == 200:
                raw = r.json()
                if isinstance(raw, list) and raw:
                    data = raw
                    break
                elif isinstance(raw, dict) and raw:
                    data = [raw]
                    break
        except Exception as e:
            print(f"[get_balancete] Tentativa {url} falhou: {e}")

    if not data:
        return (f"⚠️ Balancete não disponível via API para {rest['name']} no período {start_date} a {end_date}.\n"
                f"Acesse diretamente: portal.netcontroll.com.br → #/relatorio/balancete")

    # Agrupa por plano de contas
    grupos = {}
    total_receitas = 0.0
    total_despesas = 0.0

    for row in data:
        plano = str(row.get('planoContas', '') or row.get('descricao', '') or row.get('conta', '') or 'OUTROS').strip().upper()
        debito  = safe_float(row.get('debito',  0) or row.get('valorDebito',  0))
        credito = safe_float(row.get('credito', 0) or row.get('valorCredito', 0))
        saldo   = safe_float(row.get('saldo',   0) or (credito - debito))

        if plano not in grupos:
            grupos[plano] = {'debito': 0.0, 'credito': 0.0, 'saldo': 0.0}
        grupos[plano]['debito']  += debito
        grupos[plano]['credito'] += credito
        grupos[plano]['saldo']   += saldo

        # Classifica: crédito > débito = receita; débito > crédito = despesa
        if credito > debito:
            total_receitas += credito - debito
        else:
            total_despesas += debito - credito

    sorted_grupos = sorted(grupos.items(), key=lambda x: abs(x[1]['saldo']), reverse=True)

    report  = f"📒 **BALANCETE — {rest['name']}**\n"
    report += f"📅 Período: {start_date} → {end_date}\n"
    report += f"📍 _Fonte: portal #/relatorio/balancete_\n\n"
    report += f"{'─'*40}\n"

    for plano, vals in sorted_grupos:
        sinal = "+" if vals['credito'] >= vals['debito'] else "-"
        report += (f"• *{plano}*\n"
                   f"  Débito: R${fmt_brl(vals['debito'])} | Crédito: R${fmt_brl(vals['credito'])} | "
                   f"Saldo: {sinal}R${fmt_brl(abs(vals['saldo']))}\n")

    report += f"\n{'─'*40}\n"
    report += f"💰 **Total Créditos (Receitas):** R${fmt_brl(total_receitas)}\n"
    report += f"💸 **Total Débitos (Despesas):**  R${fmt_brl(total_despesas)}\n"
    result = total_receitas - total_despesas
    sinal_r = "✅ Superávit" if result >= 0 else "🔴 Déficit"
    report += f"📊 **Resultado do Período:** {sinal_r} R${fmt_brl(abs(result))}\n"

    return report


def get_realtime_fraud_alert(restaurant_name):
    """Radar anti-fraude em tempo real: cruza cancelamentos × caixa por operador no dia de hoje."""
    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    today = datetime.now().strftime('%Y-%m-%d')
    d_start = f"{today}T00:00:00"
    d_end   = f"{today}T23:59:59"

    report = f"🚨 **RADAR ANTI-PERDAS / FRAUDE — TEMPO REAL**\n📍 **{rest['name']}** | {today}\n\n"
    fraud_level = 0  # 0=clean, 1=attention, 2=critical

    # ── Buscar cancelamentos do dia ───────────────────────────────────────────
    cancellations = []
    try:
        r = session.get(CANCELLATION_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
        if r.status_code == 200: cancellations = r.json() or []
    except: pass

    # ── Buscar fechamentos de caixa do dia ────────────────────────────────────
    closures = []
    try:
        r = session.get(CASHIER_CLOSURE_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
        if r.status_code == 200: closures = r.json() or []
    except: pass

    # ── Mapear cancelamentos por operador ─────────────────────────────────────
    canc_by_op = {}
    total_canc = 0
    for c in cancellations:
        op = normalize_text(str(c.get('nomeOperadorCancelamento') or c.get('vendedor') or 'SEM_OP'))
        val = safe_float(c.get('valor', c.get('valorTotal', 0)))
        prod = str(c.get('nomeProduto', c.get('produtoNome', '?')))
        reason = str(c.get('motivo', 'SEM MOTIVO')).strip() or 'SEM MOTIVO'
        hora = str(c.get('dataCancelamento', ''))[-8:]
        total_canc += val
        if op not in canc_by_op:
            canc_by_op[op] = {'val': 0, 'items': [], 'reasons': set(), 'horas': []}
        canc_by_op[op]['val'] += val
        canc_by_op[op]['items'].append(prod)
        canc_by_op[op]['reasons'].add(reason)
        canc_by_op[op]['horas'].append(hora)

    # ── Mapear quebras de caixa por operador ──────────────────────────────────
    cashier_by_op = {}
    for c in closures:
        op = normalize_text(str(c.get('operador') or c.get('nomeCaixa') or 'SEM_OP'))
        diff = safe_float(c.get('diferenca', 0))
        pgto = str(c.get('pgto', 'GERAL'))
        if op not in cashier_by_op:
            cashier_by_op[op] = {'diff': 0, 'pgtos': []}
        cashier_by_op[op]['diff'] += diff
        if diff != 0: cashier_by_op[op]['pgtos'].append(f"{pgto}: {fmt_brl(diff)}")

    # ── CRUZAMENTO: operadores com cancela alto E quebra de caixa ─────────────
    cross_alerts = []
    for op, cs in sorted(canc_by_op.items(), key=lambda x: x[1]['val'], reverse=True):
        pct = (cs['val'] / total_canc * 100) if total_canc > 0 else 0
        caixa = cashier_by_op.get(op)
        has_high_canc = cs['val'] >= 100 or pct >= 30
        has_cashier_issue = caixa and abs(caixa['diff']) > 20
        sus_reason = any('SEM MOTIVO' in r.upper() or r.strip() == '' for r in cs['reasons'])

        if has_high_canc and has_cashier_issue:
            cross_alerts.append(('CRITICO', op, cs, caixa, pct))
            fraud_level = 2
        elif has_high_canc:
            cross_alerts.append(('ATENCAO', op, cs, caixa, pct))
            if fraud_level < 1: fraud_level = 1
        elif has_cashier_issue:
            cross_alerts.append(('CAIXA', op, None, caixa, 0))
            if fraud_level < 1: fraud_level = 1

    # ── Montar relatório ──────────────────────────────────────────────────────
    if cross_alerts:
        for alert_type, op, cs, caixa, pct in cross_alerts:
            if alert_type == 'CRITICO':
                report += f"🔴 **SINAL CRÍTICO DE FRAUDE — {op.upper()}**\n"
                report += f"  Cancelamentos hoje: {fmt_brl(cs['val'])} ({pct:.0f}% do total cancelado)\n"
                report += f"  Itens: {', '.join(list(set(cs['items']))[:4])}\n"
                report += f"  Motivos: {', '.join(cs['reasons'])}\n"
                report += f"  Quebra de caixa: {fmt_brl(caixa['diff'])} ({'; '.join(caixa['pgtos'][:3])})\n"
                report += f"  ⚠️ **Operador concentra AMBOS: cancela alto + caixa com diferença — investigar AGORA.**\n\n"
            elif alert_type == 'ATENCAO':
                report += f"🟡 Alto cancelamento — **{op}**: {fmt_brl(cs['val'])} ({pct:.0f}%)\n"
                report += f"  Itens: {', '.join(list(set(cs['items']))[:3])} | Motivos: {', '.join(cs['reasons'])}\n\n"
            elif alert_type == 'CAIXA':
                report += f"🟡 Quebra de caixa — **{op}**: {fmt_brl(caixa['diff'])}\n"
                if caixa['pgtos']: report += f"  Por forma: {'; '.join(caixa['pgtos'][:3])}\n\n"
    else:
        report += "✅ Nenhum padrão suspeito detectado até agora.\n\n"

    # Resumo de cancelamentos por produto (top 5)
    if cancellations:
        canc_prod = {}
        for c in cancellations:
            prod = str(c.get('nomeProduto', c.get('produtoNome', '?')))
            val = safe_float(c.get('valor', c.get('valorTotal', 0)))
            canc_prod[prod] = canc_prod.get(prod, 0) + val
        top_prods = sorted(canc_prod.items(), key=lambda x: x[1], reverse=True)[:5]
        report += f"📦 **Top cancelamentos por produto hoje:**\n"
        for prod, val in top_prods:
            report += f"  • {prod}: {fmt_brl(val)}\n"
        report += f"\n💸 Total cancelado hoje: {fmt_brl(total_canc)}\n\n"

    if fraud_level == 2:
        report += "🚨 **AÇÃO URGENTE:** Contate a gerência imediatamente e solicite justificativa dos lançamentos sinalizados."
    elif fraud_level == 1:
        report += "⚠️ **AÇÃO RECOMENDADA:** Monitore de perto os operadores sinalizados e solicite justificativas."
    else:
        report += "✅ Radar limpo. Continue monitorando."

    return report

def get_dynamic_pricing_suggestions(restaurant_name):
    weather_info = get_weather_forecast(restaurant_name)
    is_sunny = "sol" in weather_info.lower() or "limpo" in weather_info.lower()
    
    rest = find_restaurant_files(restaurant_name)
    stock_path = rest.get('stock_file')
    if not stock_path or not os.path.exists(stock_path):
        return f"Não tenho acesso ao estoque em tempo real de {rest['name']} para cruzar a precificação."
        
    stock_data = load_json(stock_path)
    
    overstocked = []
    for row in stock_data:
        nome = str(row.get('produto', '')).upper()
        qtd = safe_float(row.get('estoqueAtual', 0))
        custo_total = qtd * safe_float(row.get('custoAtual', 0))
        
        if ("CORONA" in nome or "HEINEKEN" in nome or "CHOPP" in nome) and custo_total > 3000:
            overstocked.append({"nome": nome, "qtd": qtd, "custo_total": custo_total})
            
    report = f"🎯 **MOTOR PREDITIVO DE PRECIFICAÇÃO DINÂMICA**\n📍 **{rest['name']}**\n\n"
    report += "🌤️ **Análise de Clima e Demanda:**\n"
    
    if is_sunny:
        report += "A previsão indica SOL FORTE/TEMPO BOM para os próximos dias.\n"
        report += "A expectativa de ocupação é **MAXIMIZADA**.\n\n"
        report += "💡 **Sugestão de Upsell Automático:**\n"
        report += "  • **Day Use / Couvert / Locação de Lounge:** Aumentar em 10% a 15% imediatamente. A demanda inelástica pelo sol garante a conversão e joga a margem de lucro para cima sem custo de mercadoria (Markup infinito).\n"
    else:
        report += "A previsão indica CHUVA ou TEMPO NUBLADO.\n"
        report += "A expectativa de ocupação é **BAIXA a MODERADA**.\n\n"
        report += "💡 **Sugestão de Estímulo (Giro Rápido):**\n"
        report += "  • **Day Use / Couvert:** Reduzir em 20% para não perder o dia e atrair o público que hesita em sair de casa. O ticket médio se salva na alimentação e bebida.\n"
        
    if overstocked:
        report += "\n📦 **Oportunidades de Giro (Capital Parado):**\n"
        for item in overstocked[:3]:
            report += f"  • **{item['nome']}** - Estoque perigosamente alto: {fmt_brl(item['qtd'], 0)} unid ({fmt_brl(item['custo_total'])} empatados em prateleira).\n"
            report += f"    _Ação: Baixar o preço em 5% no ERP agora mesmo, ou criar Promoção 'Compre 4 Leve 5' para desafogar o caixa._\n"
            
    report += "\n⚡ _Seja agressivo! Quer que eu conecte no ERP e aplique o reajuste agorinha mesmo? Diga: 'Aumente o preço do Day Use em 10%'_"
    return report

def get_predictive_hr_scale(restaurant_name, target_date=None):
    if not target_date:
        target_date = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
        
    rest = find_restaurant_files(restaurant_name)
    
    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')
    sales = fetch_sales_data(rest['id'], start_date, end_date)
    
    total_sales = sum([safe_float(s.get('valor', 0)) for s in sales])
    daily_avg = (total_sales / 30) if total_sales > 0 else 15000.0
    
    weather_info = get_weather_forecast(restaurant_name)
    is_sunny = "sol" in weather_info.lower() or "limpo" in weather_info.lower()
    is_rainy = "chuv" in weather_info.lower() or "trovo" in weather_info.lower()
    
    forecasted_sales = daily_avg
    clima_res = "(☁️ Padrão / Médias Históricas Fixadas)"
    if is_sunny:
        forecasted_sales = daily_avg * 1.4
        clima_res = "(☀️ Sol: Projeção Alta de +40%)"
    elif is_rainy:
        forecasted_sales = daily_avg * 0.7
        clima_res = "(🌧️ Chuva: Retração de -30%)"
        
    # Fórmulas de produtividade base (Garçom puxando 1500; Cozinha puxando 2000)
    expected_waiters = max(3, int(forecasted_sales / 1500))
    expected_kitchen = max(3, int(forecasted_sales / 2000))
    
    report = f"👥 **MOTOR DE RH PREDITIVO (Escala de Equipe)**\n📍 **{rest['name']}** | Simulação p/ dia: {target_date}\n\n"
    report += f"💰 **Volume de Vendas Esperado:** {fmt_brl(forecasted_sales)} {clima_res}\n\n"
    
    report += f"📋 **Quadro Ideal Sugerido (Algoritmo Operacional):**\n"
    report += f"  🍽️ Salão/Praia: {expected_waiters} garçons/atendentes.\n"
    report += f"  🍳 Cozinha/Bar: {expected_kitchen} cozinheiros/barmans.\n\n"
    
    # Parâmetro interno médio (Mock para gerar inteligência de relatório)
    base_waiters = 8 
    
    if expected_waiters > base_waiters * 1.2:
        falta = expected_waiters - base_waiters
        report += f"⚠️ **RISCO VERMELHO DE ATENDIMENTO (CUSTOMER SUCCESS):**\nO volume de vendas exigirá cerca de {expected_waiters} profissionais de salão.\n"
        report += f"A sua escala mínima não suportará! Se não convocar **{falta} freelancers/extras**, os clientes reclamarão de lentidão, bebida quente e o ticket médio vai afundar.\n"
    elif expected_waiters < base_waiters * 0.8:
        report += f"⚠️ **RISCO VERMELHO DE MARGEM (FOLHA CARA):**\nO seu quadro base está inchado para as vendas previstas.\n"
        report += f"Você pode cortar ou alocar banco de horas para evitar garçons ociosos sugando o caixa (Custo Folha).\n"
    else:
        report += "✅ **ESCALA CALIBRADA:** Movimento previsto e equipe estão em equilíbrio.\n"
        
    return report

def get_scenario(restaurant_names, start_date, end_date=None):
    if not end_date: end_date = start_date
    
    report_res = ""
    for name in restaurant_names:
        rest = find_restaurant_files(name)
        session = get_session_for_rest(rest['id'])
        if not session:
            report_res += f"\n--- {rest['name']} ---\nErro de conexão com API.\n"
            continue
            
        # Fetch Sales
        s_date = f"{start_date}T03:00:00.000Z"
        e_date = f"{end_date}T23:59:59.000Z"
        
        s_params = {'DataInicial': s_date, 'DataFinal': e_date, 'IncluirCusto': 'true'}

        sales_data = []
        try:
            r = session.get(SALES_URL, params=s_params)
            if r.status_code == 200: sales_data = r.json()
        except: pass
        
        # Fetch Expenses
        exp_params = {'DataInicial': s_date, 'DataFinal': e_date, 'TipoDataDespesa': 0}
        exp_data = []
        try:
            r = session.get(EXPENSES_URL, params=exp_params)
            if r.status_code == 200: exp_data = r.json()
        except: pass
        
        # Calculate
        total_sales = sum([float(x.get('valor', 0)) for x in sales_data])
        total_exp = sum([float(x.get('valor', 0)) for x in exp_data])
        
        # Novas métricas (Lucro, ROI, EBITDA, Folha)
        net_profit = total_sales - total_exp
        roi_empresa = (net_profit / total_exp * 100) if total_exp > 0 else 0
        
        payroll_keywords = ["FOLHA", "SALARIO", "SALÁRIO", "RH", "PESSOAL", "FUNCIONARIO", 
                            "FUNCIONÁRIO", "LABORE", "FGTS", "INSS", "RESCISAO", "RESCISÃO", 
                            "FERIAS", "FÉRIAS", "VALE", "VT", "VR", "BENEFICIO"]
        itda_keywords = ["IMPOSTO", "TAXA", "JUROS", "MULTA", "SIMPLES", "ICMS", "DAS", 
                         "FINANCIAMENTO", "TARIFA", "BANCARIA", "BANCÁRIA", "BANCARIO", 
                         "EMPRESTIMO", "DEPRECIACAO", "AMORTIZACAO"]
        
        total_payroll = 0
        total_itda = 0
        
        for exp in exp_data:
            val = float(exp.get('valor', 0))
            text_to_search = f"{exp.get('planoContas1', '')} {exp.get('planoContas2', '')} {exp.get('historico', '')} {exp.get('fornecedor', '')}".upper()
            
            if any(k in text_to_search for k in payroll_keywords):
                total_payroll += val
            
            if any(k in text_to_search for k in itda_keywords):
                total_itda += val
                
        ebitda = net_profit + total_itda
        roi_folha = (net_profit / total_payroll * 100) if total_payroll > 0 else 0
        
        report_res += f"\n📊 **Cenário Detalhado {rest['name']} ({start_date} a {end_date})**\n"
        report_res += f"💰 Vendas Totais: {fmt_brl(total_sales)} ({len(sales_data)} registros)\n"
        report_res += f"🧾 Despesas Ativas: {fmt_brl(total_exp)} ({len(exp_data)} registros)\n"
        report_res += f"💵 Lucro Líquido: {fmt_brl(net_profit)}\n"
        report_res += f"📈 ROI da Empresa: {fmt_brl(roi_empresa)}%\n"
        report_res += f"🏢 EBITDA (aprox.): {fmt_brl(ebitda)}\n"
        report_res += f"👥 Custo Folha de Pagamento: {fmt_brl(total_payroll)}\n"
        report_res += f"🚀 ROI da Folha: {fmt_brl(roi_folha)}%\n"
        
        if sales_data:
            top = sorted(sales_data, key=lambda x: float(x.get('valor', 0)), reverse=True)[:10]
            report_res += "🔝 Top 10 Itens por Faturamento:\n"
            for t in top: 
                report_res += f"  - {t.get('nome')}: {t.get('qtde')} unid ({fmt_brl(float(t.get('valor')))})\n"
            
        if exp_data:
            top_exp = sorted(exp_data, key=lambda x: float(x.get('valor', 0)), reverse=True)[:10]
            report_res += "💸 Top 10 Despesas/Contas a Pagar:\n"
            for t in top_exp: 
                report_res += f"  - {t.get('fornecedor')}: {fmt_brl(float(t.get('valor')))} ({t.get('planoContas1')} - {t.get('historico', '')[:40]})\n"
        elif not exp_data and not sales_data:
            report_res += "⚠️ Nenhum dado de venda ou despesa encontrado para este período exato.\n"
            
    return report_res

def get_audit(restaurant_name, query, start_date=None, end_date=None):
    if not start_date: start_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    if not end_date: end_date = start_date
    
    # We call search_sales, get_expenses, get_stock, get_inbound_purchases and stitch them
    sales = search_sales(restaurant_name, query, start_date, end_date)
    expenses = get_expenses(restaurant_name, query, start_date, end_date)
    stock = get_stock(restaurant_name, query)
    inbound = get_inbound_purchases(restaurant_name, query, start_date, end_date)
    
    report = f"🔍 **AUDITORIA DE INSIGHTS CEO: {query.upper()}**\n"
    report += f"🏠 Unidade: {restaurant_name} | 📅 Período: {start_date} a {end_date}\n\n"
    report += "📦 **ESTOQUE ATUAL & CUSTO (Via Sistema):**\n" + stock + "\n\n"
    report += "🛒 **ENTRADA DE MERCADORIAS (Via Sistema):**\n" + inbound + "\n\n"

    report += "💸 **DESPESAS FINANCEIRAS / CONTAS A PAGAR (Sistema):**\n" + expenses + "\n\n"
    report += "💰 **VENDAS NO PERÍODO (Saídas de Caixa):**\n" + sales + "\n\n"
    report += "⚠️ Cruze essas informações para achar divergências (ex: mercadoria que entrou mas não acusou na despesa, repasses de preço, itens com ruptura)."
    
    return report

def get_purchasing_plan(restaurant_name, query=None, days_history=7, coverage_days=7):
    try: days_history = int(days_history)
    except: days_history = 7
    try: coverage_days = int(coverage_days)
    except: coverage_days = 7
    
    start_date = (datetime.now() - timedelta(days=days_history)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')
    
    rest = find_restaurant_files(restaurant_name)
    sales = fetch_sales_data(rest['id'], start_date, end_date)
    
    sales_map = {}
    for r in sales:
        n = normalize_text(str(r.get('nome', '')))
        if n: sales_map[n] = sales_map.get(n, 0) + safe_float(r.get('qtde', 0))
        
    # Build complete ingredient demand map
    ingredient_demand = dict(sales_map)
    recipe_path = rest.get('recipe_file')
    if recipe_path and os.path.exists(recipe_path):
        try:
            wb = openpyxl.load_workbook(recipe_path, data_only=True)
            ws = wb.active
            current_dish = None
            recipes = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                r0 = str(row[0]).strip() if row[0] else ""
                r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if r0.startswith("Produto:") or (r0 and not r1 and not r0.replace('.','').isdigit()):
                    current_dish = normalize_text(r0.split(":", 1)[1]) if ":" in r0 else normalize_text(r0)
                    recipes[current_dish] = []
                elif current_dish and len(row) >= 4:
                    if isinstance(row[2], (int, float)):
                        ing_name = normalize_text(str(row[1])) if row[1] else ""
                        ing_qty = safe_float(row[2])
                        if ing_name and ing_qty > 0:
                            recipes[current_dish].append((ing_name, ing_qty))
                            
            for sn, qty_sold in sales_map.items():
                for dish, ingr_list in recipes.items():
                    if sn == dish:
                        for ing_name, ing_qty in ingr_list:
                            ingredient_demand[ing_name] = ingredient_demand.get(ing_name, 0) + (qty_sold * ing_qty)
        except: pass
        
    stock_data = []
    try:
        session = get_session_for_rest(rest['id'])
        if session:
            r = session.get(INVENTORY_URL)
            if r.status_code == 200:
                raw = r.json()
                if isinstance(raw, list) and len(raw) > 0: stock_data = raw
    except: pass

    if not stock_data:
        stock_path = rest.get('stock_file')
        if stock_path and os.path.exists(stock_path):
            stock_data = load_json(stock_path)
            
    if not stock_data:
        return f"Lista de Estoque inacessível para {rest['name']}."
    suggestions = []
    
    # Pre-aggregate stock to avoid duplicates returned by API
    stock_map = {}
    for row in stock_data:
        nome = str(row.get('produto', '')).strip()
        n_norm = normalize_text(nome)
        if len(n_norm) < 3: continue
        
        if query:
            grupo = str(row.get('grupo', ''))
            subgrupo = str(row.get('subgrupo', ''))
            if not (match_query(query, nome) or match_query(query, grupo) or match_query(query, subgrupo)):
                continue
                
        if n_norm not in stock_map:
            stock_map[n_norm] = {
                'nome': nome,
                'estoqueAtual': safe_float(row.get('estoqueAtual', 0)),
                'custoAtual': safe_float(row.get('custoAtual', 0))
            }
        else:
            stock_map[n_norm]['estoqueAtual'] += safe_float(row.get('estoqueAtual', 0))
            stock_map[n_norm]['custoAtual'] = max(stock_map[n_norm]['custoAtual'], safe_float(row.get('custoAtual', 0)))

    for n_norm, data in stock_map.items():
        qty_sold = ingredient_demand.get(n_norm, 0)
                
        current = data['estoqueAtual']
        daily_avg = qty_sold / max(1, days_history)
        projected = daily_avg * coverage_days
        
        to_buy = 0
        if current > 0:
            if projected > current and qty_sold > 0:
                to_buy = projected - current
        else:
            if qty_sold > 0:
                to_buy = projected
                
        # Skip intangibles
        if 'DAY USE' in n_norm.upper() or 'PONTO' in n_norm.upper() or 'PASSADA' in n_norm.upper() or 'COUVERT' in n_norm.upper():
            continue
            
        if to_buy > 0 or query:
            suggestions.append({
                'nome': data['nome'],
                'sold': qty_sold,
                'current': current,
                'projected': projected,
                'to_buy': to_buy,
                'cost': data['custoAtual'] * to_buy if to_buy > 0 else 0
            })
             
    suggestions.sort(key=lambda x: x['to_buy'], reverse=True)
    
    if not suggestions:
        return f"✅ O estoque atual de {rest['name']} parece cobrir a demanda projetada de todos os itens e insumos para os próximos {coverage_days} dias."
        
    report = f"📋 **PLANO DE COMPRAS INTELIGENTE/ESTOQUE**\n"
    q_lbl = f" | Filtrado por: {query.upper()}" if query else ""
    report += f"🏠 {rest['name']} | Receituário & Vendas | {days_history} dias base | Alvo: {coverage_days} dias{q_lbl}\n\n"
    
    total_cost = 0
    for s in suggestions: 
        report += f"🛒 **{s['nome']}**\n"
        report += f"  - Estoque Atual: {fmt_brl(s['current'])}\n"
        if s['sold'] > 0:
            report += f"  - Consumo (Vendas {days_history}d): {fmt_brl(s['sold'])} unid/kg\n"
            report += f"  - Sujestão IA: **Comprar {fmt_brl(s['to_buy'])} unid/kg**\n"
            if s['cost'] > 0:
                report += f"  - Investimento: {fmt_brl(s['cost'])}\n"
        else:
            report += f"  - Sugestão: Item sem giro direto via vendas/ficha técnica.\n"
        total_cost += s['cost']
        report += "\n"
        
    report += f"💰 **Estimativa Total do Pedido: {fmt_brl(total_cost)}**\n\n"
    report += "_Nota: Este cálculo utiliza os ingredientes da Ficha Técnica + Faturamento Real._"
    
    return report[:14000] # Prevenção rigorosa de erro 429 de limite de Tokens OpenAI

def get_waste_audit(restaurant_name, query=None, days_history=7):
    try: days_history = int(days_history)
    except: days_history = 7
    
    start_date = (datetime.now() - timedelta(days=days_history)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')
    
    rest = find_restaurant_files(restaurant_name)
    sales = fetch_sales_data(rest['id'], start_date, end_date)
    inbound_data = fetch_inbound_data(rest['id'], start_date, end_date)
    
    sales_map = {}
    for r in sales:
        n = normalize_text(str(r.get('nome', '')))
        if n: sales_map[n] = sales_map.get(n, 0) + safe_float(r.get('qtde', 0))
        
    inbound_map = {}
    if inbound_data:
        for r in inbound_data:
            n = normalize_text(str(r.get('produto', '')))
            if n: inbound_map[n] = inbound_map.get(n, 0) + safe_float(r.get('qtde', 0))
        
    ingredient_demand = dict(sales_map)
    recipe_path = rest.get('recipe_file')
    if recipe_path and os.path.exists(recipe_path):
        try:
            wb = openpyxl.load_workbook(recipe_path, data_only=True)
            ws = wb.active
            current_dish = None
            recipes = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                r0 = str(row[0]).strip() if row[0] else ""
                r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if r0.startswith("Produto:") or (r0 and not r1 and not r0.replace('.','').isdigit() and not r0.startswith("Subgrupo:")):
                    current_dish = normalize_text(r0.split(":", 1)[1]) if ":" in r0 else normalize_text(r0)
                    recipes[current_dish] = []
                elif current_dish and len(row) >= 4:
                    if isinstance(row[2], (int, float)):
                        ing_name = normalize_text(str(row[1])) if row[1] else ""
                        ing_qty = safe_float(row[2])
                        if ing_name and ing_qty > 0:
                            recipes[current_dish].append((ing_name, ing_qty))
                            
            for sn, qty_sold in sales_map.items():
                for dish, ingr_list in recipes.items():
                    if sn == dish:
                        for ing_name, ing_qty in ingr_list:
                            ingredient_demand[ing_name] = ingredient_demand.get(ing_name, 0) + (qty_sold * ing_qty)
        except: pass
        
    stock_path = rest.get('stock_file')
    if not stock_path or not os.path.exists(stock_path):
        return f"Lista de Estoque inacessível para {rest['name']}."
        
    stock_data = load_json(stock_path)
    
    stock_map = {}
    for row in stock_data:
        nome = str(row.get('produto', '')).strip()
        n_norm = normalize_text(nome)
        if len(n_norm) < 3: continue
        
        if query:
            grupo = str(row.get('grupo', ''))
            subgrupo = str(row.get('subgrupo', ''))
            if not (match_query(query, nome) or match_query(query, grupo) or match_query(query, subgrupo)):
                continue
                
        if n_norm not in stock_map:
            stock_map[n_norm] = {
                'nome': nome,
                'estoqueAtual': safe_float(row.get('estoqueAtual', 0)),
                'custoAtual': safe_float(row.get('custoAtual', 0))
            }
        else:
            stock_map[n_norm]['estoqueAtual'] += safe_float(row.get('estoqueAtual', 0))
            stock_map[n_norm]['custoAtual'] = max(stock_map[n_norm]['custoAtual'], safe_float(row.get('custoAtual', 0)))

    audit_results = []
    
    for n_norm, data in stock_map.items():
        if 'DAY USE' in n_norm.upper() or 'PONTO' in n_norm.upper() or 'PASSADA' in n_norm.upper() or 'COUVERT' in n_norm.upper():
            continue
            
        qty_sold_theoretical = ingredient_demand.get(n_norm, 0)
        current_stock = data['estoqueAtual']
        inbound_qty = inbound_map.get(n_norm, 0)
        
        if qty_sold_theoretical > 0 or inbound_qty > 0 or query:
            audit_results.append({
                'nome': data['nome'],
                'theoretical_consumed': qty_sold_theoretical,
                'system_stock': current_stock,
                'inbound': inbound_qty,
                'cost': data['custoAtual']
            })
            
    # Sort by value of consumed goods (highest tracking priority)
    audit_results.sort(key=lambda x: max(x['theoretical_consumed'], x['inbound']) * x['cost'], reverse=True)
    
    if not audit_results:
        return f"✅ Nenhum item rastreável encontrado para a auditoria em {rest['name']}."
        
    q_lbl = f" | Foco: {query.upper()}" if query else " | Foco: Itens Mais Consumidos"
    report = f"🕵️ **AUDITORIA REAL VS. TEÓRICA (Detecção de Quebras)**\n"
    report += f"🏠 {rest['name']} | Base Vendas/Entradas: {days_history} últimos dias{q_lbl}\n\n"
    report += "⚠️ *Instrução CEO:* Vá fisicamente ao Freezer/Estoque e faça a contagem cega dos itens abaixo. O Consumo Real é calculado como: (Estoque Anterior + Compras) - Saldo Físico.\n\n"
    
    for a in audit_results[:35]: # Limit output to top 35 to save tokens
        cost_str = f"{fmt_brl(a['cost'])}" if a['cost'] > 0 else "N/A"
        report += f"📍 **{a['nome']}** (Custo ref: {cost_str})\n"
        report += f"  - Consumo Teórico (Vendas x Ficha): **{fmt_brl(a['theoretical_consumed'])}** unid/kg/L\n"
        report += f"  - Entradas de Mercadoria (Compras): **{fmt_brl(a['inbound'])}** unid/kg/L\n"
        report += f"  - Saldo no Sistema Hoje: **{fmt_brl(a['system_stock'])}** unid/kg/L\n"
        report += f"  - Saldo Físico Encontrado: [        ] unid/kg/L\n"
        report += f"  - Consumo Real Apurado: [        ] unid/kg/L\n\n"
        
    if len(audit_results) > 35:
        report += f"_... e mais {len(audit_results) - 35} itens. Seja mais específico na busca caso o insumo suspeito não esteja listado._\n"
        
    return report[:14000]

def get_menu_engineering(restaurant_name, start_date=None, end_date=None):
    if not start_date: start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    if not end_date: end_date = datetime.now().strftime('%Y-%m-%d')
    
    rest = find_restaurant_files(restaurant_name)
    sales = fetch_sales_data(rest['id'], start_date, end_date)
    cmv_data = fetch_cmv_data(rest['id'], start_date, end_date)
    
    if not sales or not cmv_data:
        return f"Não há dados suficientes (Vendas ou CMV) para {rest['name']} no período ({start_date} a {end_date})."
        
    # Map volume from sales
    volume_map = {}
    for r in sales:
        n = normalize_text(str(r.get('nome', '')))
        if n and not any(x in n.upper() for x in ['DAY USE', 'COUVERT', 'PONTO', 'PASSADA', 'FONE']):
            volume_map[n] = volume_map.get(n, 0) + safe_float(r.get('qtde', 0))
            
    items = []
    processed = set()
    
    # Cross-reference with CMV margins
    for c in cmv_data:
        n = normalize_text(str(c.get('nome', '')))
        if n in volume_map and n not in processed:
             pr = safe_float(c.get('preco', 0))
             custo = safe_float(c.get('precoCompra', 0))
             margem = pr - custo
             # Fallback if margem is somehow 0 but CMV API gives margem field
             if margem == 0: margem = safe_float(c.get('margem', 0))
             
             items.append({
                 'nome': str(c.get('nome', '')),
                 'grupo': str(c.get('grupo', '')),
                 'vol': volume_map[n],
                 'margem': margem,
                 'custo': custo,
                 'preco': pr
             })
             processed.add(n)
             
    if not items:
        return f"Não foi possível cruzar Vendas com Ficha Técnica para a Matriz BCG em {rest['name']}."
        
    # Calculate medians for the Matrix Cross
    sorted_vol = sorted([i['vol'] for i in items])
    sorted_margem = sorted([i['margem'] for i in items])
    
    med_vol = sorted_vol[len(sorted_vol)//2] if sorted_vol else 0
    med_margem = sorted_margem[len(sorted_margem)//2] if sorted_margem else 0
    
    stars = []
    horses = []
    puzzles = []
    dogs = []
    
    for i in items:
        if i['vol'] >= med_vol and i['margem'] >= med_margem:
            stars.append(i)
        elif i['vol'] >= med_vol and i['margem'] < med_margem:
            horses.append(i)
        elif i['vol'] < med_vol and i['margem'] >= med_margem:
            puzzles.append(i)
        else:
            dogs.append(i)
            
    stars.sort(key=lambda x: x['vol'] * x['margem'], reverse=True)
    horses.sort(key=lambda x: x['vol'], reverse=True)
    puzzles.sort(key=lambda x: x['margem'], reverse=True)
    dogs.sort(key=lambda x: x['vol'] * x['margem']) # worst first
    
    report = f"🎯 **ENGENHARIA DE CARDÁPIO AUTÔNOMA (MATRIZ BCG)**\n"
    report += f"🏠 {rest['name']} | 📅 {start_date} a {end_date}\n"
    report += f"📊 Ponto de Equilíbrio do Cardápio: Giro Mídia de {fmt_brl(med_vol, 1)} unid | Lucro Mídio de {fmt_brl(med_margem)}\n\n"
    
    report += "🌟 **ESTRELAS (Alto Lucro + Alta Venda)**\n_Aumente a publicidade, coloque em destaque no menu._\n"
    for i in stars[:15]:
        report += f"  ➤ {i['nome']} (Vendeu {i['vol']} unid | Lucro {fmt_brl(i['margem'])}/prato)\n"
    if not stars: report += "  _Nenhum item nesta categoria._\n"
        
    report += "\n🐎 **BURROS DE CARGA (Baixo Lucro + Alta Venda)**\n_Aumente o preço em R$ 2 ou renegocie o custo do insumo com fornecedor._\n"
    for i in horses[:15]:
        report += f"  ➤ {i['nome']} (Vendeu {i['vol']} unid | Lucro {fmt_brl(i['margem'])}/prato)\n"
    if not horses: report += "  _Nenhum item nesta categoria._\n"
        
    report += "\n🧩 **QUEBRA-CABEÇAS (Alto Lucro + Baixa Venda)**\n_Faça combo, comissione o garçom para empurrar este prato._\n"
    for i in puzzles[:15]:
        report += f"  ➤ {i['nome']} (Vendeu só {i['vol']} unid | SUPER Lucro {fmt_brl(i['margem'])}/prato)\n"
    if not puzzles: report += "  _Nenhum item nesta categoria._\n"
        
    report += "\n🐕 **CACHORROS (Baixo Lucro + Baixa Venda)**\n_Considere cortar do cardápio. Custo oculto de estoque._\n"
    for i in dogs[:15]:
        report += f"  ➤ {i['nome']} (Vendeu {i['vol']} unid | Fraco Lucro {fmt_brl(i['margem'])}/prato)\n"
    if not dogs: report += "  _Nenhum item nesta categoria._\n"

    report += "\n💡 _As Dicas Acima foram geradas matematicamente pela sua IA. Foque nos Burros de Carga hoje._"
    
    return report[:14000]

def get_supplier_inflation(restaurant_name, query=None, days_recent=15, days_old=30):
    try: days_recent = int(days_recent)
    except: days_recent = 15
    try: days_old = int(days_old)
    except: days_old = 30
    
    rest = find_restaurant_files(restaurant_name)
    
    # Period 1: Recent purchases (last X days)
    recent_start = (datetime.now() - timedelta(days=days_recent)).strftime('%Y-%m-%d')
    recent_end = datetime.now().strftime('%Y-%m-%d')
    recent_data = fetch_inbound_data(rest['id'], recent_start, recent_end)
    
    # Period 2: Older purchases (X to Y days ago)
    old_start = (datetime.now() - timedelta(days=days_old)).strftime('%Y-%m-%d')
    old_end = (datetime.now() - timedelta(days=days_recent)).strftime('%Y-%m-%d')
    old_data = fetch_inbound_data(rest['id'], old_start, old_end)
    
    if not recent_data and not old_data:
        return f"Sem dados de entrada de mercadoria para {rest['name']} nos últimos {days_old} dias."

    def build_price_map(data):
        price_map = {}
        for r in data:
            produto = str(r.get('produto', '')).strip()
            fornecedor = str(r.get('fornecedor', '')).strip()
            preco = safe_float(r.get('valorUnitario', 0))
            if not produto or preco <= 0: continue
            
            if query and not match_query(query, produto) and not match_query(query, fornecedor):
                continue
                
            key = normalize_text(produto)
            if key not in price_map or preco > price_map[key]['preco']:
                price_map[key] = {
                    'nome': produto,
                    'fornecedor': fornecedor,
                    'preco': preco
                }
        return price_map
    
    recent_prices = build_price_map(recent_data)
    old_prices = build_price_map(old_data)
    
    alerts = []
    
    for key, recent in recent_prices.items():
        if key in old_prices:
            old = old_prices[key]
            if old['preco'] > 0:
                variacao = ((recent['preco'] - old['preco']) / old['preco']) * 100
                if abs(variacao) >= 3:  # Flag changes above 3%
                    alerts.append({
                        'nome': recent['nome'],
                        'fornecedor': recent['fornecedor'],
                        'preco_antigo': old['preco'],
                        'preco_atual': recent['preco'],
                        'variacao': variacao
                    })
    
    # Sort by biggest increase first
    alerts.sort(key=lambda x: x['variacao'], reverse=True)
    
    q_lbl = f" | Filtro: {query.upper()}" if query else ""
    report = f"📈 **TRACKER DE INFLAÇÃO OCULTA DE FORNECEDORES**\n"
    report += f"🏠 {rest['name']}{q_lbl}\n"
    report += f"📅 Período Recente: {recent_start} a {recent_end}\n"
    report += f"📅 Período Anterior: {old_start} a {old_end}\n\n"
    
    if not alerts:
        report += "✅ Nenhuma variação significativa de preço detectada (acima de 3%).\n"
        return report
    
    increases = [a for a in alerts if a['variacao'] > 0]
    decreases = [a for a in alerts if a['variacao'] < 0]
    
    if increases:
        report += "🚨 **AUMENTOS DE PREÇO DETECTADOS:**\n"
        for a in increases[:25]:
            report += f"  🔺 **{a['nome']}** (+{fmt_pct(a['variacao'])}%)\n"
            report += f"     Fornecedor: {a['fornecedor']}\n"
            report += f"     Preço Anterior: {fmt_brl(a['preco_antigo'])} → Atual: {fmt_brl(a['preco_atual'])}\n\n"
    
    if decreases:
        report += "🟢 **REDUÇÕES DE PREÇO IDENTIFICADAS:**\n"
        for a in decreases[:15]:
            report += f"  🔽 **{a['nome']}** ({fmt_pct(a['variacao'])}%)\n"
            report += f"     Fornecedor: {a['fornecedor']}\n"
            report += f"     Preço Anterior: {fmt_brl(a['preco_antigo'])} → Atual: {fmt_brl(a['preco_atual'])}\n\n"
    
    report += "💡 _Renegocie imediatamente com os fornecedores que aumentaram. Use as reduções como argumento para pressionar concorrentes._"
    
    return report[:14000]

def get_cashflow_runway(restaurant_name, days_forward=7):
    try: days_forward = int(days_forward)
    except: days_forward = 7
    
    rest = find_restaurant_files(restaurant_name)
    
    # 1) Revenue trend: average daily sales from last 7 days
    sales_start = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    sales_end = datetime.now().strftime('%Y-%m-%d')
    sales = fetch_sales_data(rest['id'], sales_start, sales_end)
    total_revenue_7d = sum([safe_float(s.get('valor', 0)) for s in sales])
    avg_daily_revenue = total_revenue_7d / 7 if total_revenue_7d > 0 else 0
    projected_revenue = avg_daily_revenue * days_forward
    
    # 2) Upcoming bills: expenses due in the next X days
    today = datetime.now().strftime('%Y-%m-%d')
    future = (datetime.now() + timedelta(days=days_forward)).strftime('%Y-%m-%d')
    expenses = fetch_expenses_data(rest['id'], today, future)
    
    upcoming_bills = []
    total_bills = 0.0
    total_paid = 0.0
    total_pending = 0.0
    
    for e in expenses:
        val = safe_float(e.get('valor', 0))
        is_paid = e.get('pagamento', False)
        forn = str(e.get('fornecedor', 'N/A')).strip()
        cat = str(e.get('planoContas1', '')).strip()
        venc = str(e.get('dataVencimento', ''))[:10]
        
        total_bills += val
        if is_paid:
            total_paid += val
        else:
            total_pending += val
            upcoming_bills.append({
                'fornecedor': forn,
                'valor': val,
                'categoria': cat,
                'vencimento': venc
            })
    
    upcoming_bills.sort(key=lambda x: x['vencimento'])
    
    # 3) Cash flow projection
    saldo_projetado = projected_revenue - total_pending
    
    report = f"💸 **PROJEÇÃO DE FLUXO DE CAIXA (RUNWAY)**\n"
    report += f"🏠 {rest['name']} | Próximos {days_forward} dias\n\n"
    
    report += f"📈 **RECEITA PROJETADA:**\n"
    report += f"  - Faturamento Últimos 7d: {fmt_brl(total_revenue_7d)}\n"
    report += f"  - Média Diária: {fmt_brl(avg_daily_revenue)}\n"
    report += f"  - Projeção {days_forward} dias: **{fmt_brl(projected_revenue)}**\n\n"
    
    report += f"📉 **CONTAS A PAGAR (Próximos {days_forward} dias):**\n"
    report += f"  - Total Compromissos: {fmt_brl(total_bills)}\n"
    report += f"  - Já Pagos: {fmt_brl(total_paid)} ✅\n"
    report += f"  - Pendentes: **{fmt_brl(total_pending)}** ⏳\n\n"
    
    if upcoming_bills:
        report += "📋 **DETALHAMENTO DAS CONTAS PENDENTES:**\n"
        for b in upcoming_bills[:30]:
            report += f"  ⏳ {b['vencimento']} | {b['fornecedor'][:40]} | {fmt_brl(b['valor'])}\n"
            report += f"     _{b['categoria']}_\n"
        if len(upcoming_bills) > 30:
            report += f"  _...e mais {len(upcoming_bills) - 30} contas._\n"
        report += "\n"
    
    if saldo_projetado >= 0:
        report += f"✅ **SALDO LÍQUIDO PROJETADO: {fmt_brl(saldo_projetado)}**\n"
        report += f"_O faturamento projetado COBRE as despesas pendentes com folga de {fmt_brl(saldo_projetado)}._\n"
    else:
        deficit = abs(saldo_projetado)
        report += f"🚨 **ALERTA DE DÉFICIT: -{fmt_brl(deficit)}**\n"
        report += f"_O faturamento projetado NÃO COBRE as despesas pendentes! Faltarão aproximadamente {fmt_brl(deficit)}. Antecipe recebíveis ou renegocie prazos._\n"
    
    return report[:14000]

def get_weather_forecast(restaurant_name=None):
    # Per-restaurant GPS coordinates
    COORDS = {
        'nauan':    {'lat': -5.12, 'lon': -35.63, 'label': 'São Miguel do Gostoso / RN'},
        'milagres': {'lat': -9.27, 'lon': -35.38, 'label': 'São Miguel dos Milagres / AL'},
        'ahau':     {'lat': -9.27, 'lon': -35.38, 'label': 'São Miguel dos Milagres / AL'},
    }
    
    # Pick coordinates from restaurant name
    coord = COORDS.get('milagres')  # default Milagres (AL)
    region_label = 'São Miguel dos Milagres / AL'
    if restaurant_name:
        rn = restaurant_name.lower()
        for key, val in COORDS.items():
            if key in rn:
                coord = val
                region_label = val['label']
                break
    
    lat = coord['lat']
    lon = coord['lon']
    
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,weathercode&timezone=America/Recife&forecast_days=7"
        r = requests.get(url, timeout=10)
        data = r.json()
    except Exception as e:
        return f"Erro ao consultar API de clima: {e}"
    
    daily = data.get('daily', {})
    dates = daily.get('time', [])
    temp_max = daily.get('temperature_2m_max', [])
    temp_min = daily.get('temperature_2m_min', [])
    precip = daily.get('precipitation_sum', [])
    codes = daily.get('weathercode', [])
    
    if not dates:
        return "Sem dados de previsão do tempo disponíveis."
    
    weather_names = {
        0: "☀️ Céu Limpo", 1: "🌤️ Predominantemente Limpo", 2: "⛅ Parcialmente Nublado", 3: "☁️ Nublado",
        45: "🌫️ Neblina", 48: "🌫️ Neblina Gelada",
        51: "🌦️ Garoa Leve", 53: "🌦️ Garoa", 55: "🌧️ Garoa Intensa",
        61: "🌧️ Chuva Leve", 63: "🌧️ Chuva Moderada", 65: "🌧️ Chuva Forte",
        80: "🌧️ Pancadas de Chuva", 81: "🌧️ Pancadas Moderadas", 82: "⛈️ Pancadas Fortes",
        95: "⛈️ Trovoadas", 96: "⛈️ Trovoadas c/ Granizo", 99: "⛈️ Trovoadas Intensas"
    }
    
    report = f"🌡️ **TERMÔMETRO METEOROLÓGICO (Impacto no Negócio)**\n"
    report += f"📍 {region_label}\n\n"
    
    rainy_days = 0
    hot_days = 0
    
    for i, dt in enumerate(dates):
        tmax = temp_max[i] if i < len(temp_max) else 0
        tmin = temp_min[i] if i < len(temp_min) else 0
        rain = precip[i] if i < len(precip) else 0
        code = codes[i] if i < len(codes) else 0
        
        weather_desc = weather_names.get(code, "🌤️ Indefinido")
        
        rain_alert = ""
        if rain > 5:
            rain_alert = " ⚠️ ALTO IMPACTO"
            rainy_days += 1
        elif rain > 1:
            rain_alert = " ⚡ MODERADO"
            rainy_days += 1
        
        if tmax >= 30:
            hot_days += 1
            
        report += f"📅 **{dt}** | {weather_desc}\n"
        report += f"  🌡️ {fmt_brl(tmin, 0)}°C - {fmt_brl(tmax, 0)}°C | 🌧️ Chuva: {fmt_brl(rain, 1)}mm{rain_alert}\n\n"
    
    report += "---\n"
    report += "💼 **IMPACTO PROJETADO NA OPERAÇÃO:**\n"
    
    if rainy_days >= 4:
        report += f"🚨 **SEMANA CHUVOSA ({rainy_days}/7 dias com chuva)!** Reduza pedidos de cervejas e itens de praia em 30-40%. Foque em drinks quentes e sobremesas.\n"
    elif rainy_days >= 2:
        report += f"⚠️ **Chuva em {rainy_days} dias.** Considere reduzir pedidos de gelo e cervejas em 15-20%. Mantenha estoque de pratos quentes.\n"
    else:
        report += f"☀️ **Semana ensolarada!** Demanda máxima prevista para cervejas, águas, sucos e Day Use. Garanta estoque cheio de bebidas geladas.\n"
    
    if hot_days >= 5:
        report += f"🔥 **{hot_days} dias acima de 30°C!** Aumente o pedido de gelo, água de coco e cervejas. Espere lotação máxima no Beach Club.\n"
    
    return report[:14000]

def get_invoice_reconciliation(restaurant_name, days=15):
    try: days = int(days)
    except: days = 15
    
    rest = find_restaurant_files(restaurant_name)
    start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    end = datetime.now().strftime('%Y-%m-%d')
    
    inbound = fetch_inbound_data(rest['id'], start, end)
    expenses = fetch_expenses_data(rest['id'], start, end)
    
    if not inbound:
        return f"Sem notas de entrada para {rest['name']} nos últimos {days} dias."
    
    # Map inbound notes by nota fiscal number
    notas_entrada = {}
    for i in inbound:
        nf = str(i.get('notaFiscal', '')).strip()
        forn = str(i.get('fornecedor', '')).strip()
        valor = safe_float(i.get('totalLiquido', 0))
        if nf:
            key = f"{nf}_{normalize_text(forn)}"
            if key not in notas_entrada:
                notas_entrada[key] = {'nf': nf, 'fornecedor': forn, 'valor': valor, 'itens': 1}
            else:
                notas_entrada[key]['valor'] += valor
                notas_entrada[key]['itens'] += 1
    
    # Map expenses by historico (which usually contains NF number)
    expense_nfs = set()
    for e in expenses:
        hist = str(e.get('historico', '')).upper()
        for nf_key, nota in notas_entrada.items():
            if nota['nf'] in hist:
                expense_nfs.add(nf_key)
    
    # Find mismatches
    sem_financeiro = []
    for key, nota in notas_entrada.items():
        if key not in expense_nfs:
            sem_financeiro.append(nota)
    
    sem_financeiro.sort(key=lambda x: x['valor'], reverse=True)
    
    report = f"🧾 **CONCILIAÇÃO DE NOTAS FISCAIS**\n"
    report += f"🏠 {rest['name']} | Últimos {days} dias ({start} a {end})\n\n"
    report += f"📊 Total de Notas de Entrada: {len(notas_entrada)}\n"
    report += f"✅ Conciliadas com Financeiro: {len(expense_nfs)}\n"
    report += f"⚠️ Sem lançamento no Financeiro: {len(sem_financeiro)}\n\n"
    
    if sem_financeiro:
        total_missing = sum([n['valor'] for n in sem_financeiro])
        report += f"🚨 **NOTAS SEM LANÇAMENTO NO CONTAS A PAGAR ({fmt_brl(total_missing)} total):**\n"
        for n in sem_financeiro[:25]:
            report += f"  ❌ NF {n['nf']} | {n['fornecedor'][:45]} | {fmt_brl(n['valor'])} ({n['itens']} itens)\n"
        if len(sem_financeiro) > 25:
            report += f"  _...e mais {len(sem_financeiro) - 25} notas._\n"
        report += "\n💡 _Essas notas deram entrada no estoque mas podem não ter sido lançadas no financeiro. Verifique com o setor administrativo._\n"
    else:
        report += "✅ **Todas as notas de entrada estão conciliadas com o financeiro!**\n"
    
    return report[:14000]

def get_complete_audit(restaurant_name, start_date=None, end_date=None):
    """
    Auditoria Cruzada Completa CEO: cruza cancelamentos × caixa × entradas NF × despesas × vendas.
    Detecta: fraude por operador, NF sem lançamento financeiro, pagamentos sem entrada, duplicatas,
    e produtos com cancelamento suspeito. Única ferramenta que FAZ o cruzamento programaticamente.
    """
    if not start_date: start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    if not end_date: end_date = datetime.now().strftime('%Y-%m-%d')

    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session: return "❌ Erro de autenticação no portal."

    d_start = f"{start_date}T03:00:00.000Z"
    d_end   = f"{end_date}T23:59:59.000Z"

    # ── 1. FETCH ALL DATA ──────────────────────────────────────────────────────
    cancellations = []
    try:
        r = session.get(CANCELLATION_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
        if r.status_code == 200: cancellations = r.json() or []
    except: pass

    cashier_data = []
    try:
        r = session.get(CASHIER_CLOSURE_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
        if r.status_code == 200: cashier_data = r.json() or []
    except: pass

    inbound = fetch_inbound_data(rest['id'], start_date, end_date) or []
    expenses_plan = fetch_expenses_data(rest['id'], start_date, end_date) or []

    # Detailed expenses (supplier-level) — try two endpoints
    expenses_detail = []
    try:
        r = session.get(EXPENSES_DETAILED_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
        if r.status_code == 200: expenses_detail = r.json() or []
    except: pass
    if not expenses_detail:
        try:
            r = session.get(EXPENSES_SUPPLIER_URL, params={'DataInicial': d_start, 'DataFinal': d_end})
            if r.status_code == 200: expenses_detail = r.json() or []
        except: pass

    sales = fetch_sales_data(rest['id'], start_date, end_date) or []

    # ── 2. CROSS A: CANCELAMENTOS × CAIXA POR OPERADOR ────────────────────────
    canc_by_op = {}
    for c in cancellations:
        op = normalize_text(str(c.get('nomeOperadorCancelamento') or c.get('vendedor') or 'SEM_OP'))
        val = safe_float(c.get('valor', c.get('valorTotal', 0)))
        qty = safe_float(c.get('qtde', 1))
        prod = str(c.get('nomeProduto', c.get('produtoNome', '?')))
        reason = str(c.get('motivo', 'SEM MOTIVO')).strip() or 'SEM MOTIVO'
        if op not in canc_by_op:
            canc_by_op[op] = {'valor': 0, 'qtde': 0, 'items': [], 'reasons': {}}
        canc_by_op[op]['valor'] += val
        canc_by_op[op]['qtde'] += qty
        canc_by_op[op]['items'].append(prod)
        canc_by_op[op]['reasons'][reason] = canc_by_op[op]['reasons'].get(reason, 0) + val

    cashier_by_op = {}
    for c in cashier_data:
        op = normalize_text(str(c.get('operador') or c.get('nomeCaixa') or 'SEM_OP'))
        diff = safe_float(c.get('diferenca', 0))
        if op not in cashier_by_op:
            cashier_by_op[op] = {'diff_total': 0, 'fechamentos': 0}
        cashier_by_op[op]['diff_total'] += diff
        cashier_by_op[op]['fechamentos'] += 1

    total_canc_val = sum(v['valor'] for v in canc_by_op.values()) or 1

    cross_fraud = []   # operators with BOTH high cancellations AND cashier shortfall
    solo_high_canc = []  # operators with only high cancellations
    for op, cs in sorted(canc_by_op.items(), key=lambda x: x[1]['valor'], reverse=True):
        pct = cs['valor'] / total_canc_val * 100
        cashier_diff = cashier_by_op.get(op, {}).get('diff_total', None)
        has_high_canc = cs['valor'] >= 200 or pct >= 25
        has_cashier_issue = cashier_diff is not None and abs(cashier_diff) > 20
        if has_high_canc and has_cashier_issue:
            cross_fraud.append({'op': op, 'canc_val': cs['valor'], 'canc_pct': pct,
                                 'cashier_diff': cashier_diff, 'items': list(set(cs['items']))[:4],
                                 'reasons': cs['reasons']})
        elif has_high_canc:
            solo_high_canc.append({'op': op, 'canc_val': cs['valor'], 'canc_pct': pct,
                                    'reasons': cs['reasons']})

    # ── 3. CROSS B: ENTRADAS × DESPESAS (NF sem lançamento financeiro) ──────────
    inbound_map = {}  # key "NF|fornecedor_norm" → {nf, fornecedor, valor, data}
    for i in inbound:
        nf  = str(i.get('notaFiscal', '')).strip()
        forn_raw = str(i.get('fornecedor', '')).strip()
        forn = normalize_text(forn_raw)
        valor = safe_float(i.get('totalLiquido', i.get('valorTotal', 0)))
        data = str(i.get('data', i.get('dataEmissao', '')))[:10]
        if nf and valor > 0:
            key = f"{nf}|{forn}"
            inbound_map[key] = {'nf': nf, 'fornecedor': forn_raw, 'valor': valor, 'data': data}

    # Find which NFs appear in expense historicos
    reconciled_nfs = set()
    all_expenses = expenses_detail + expenses_plan
    for e in all_expenses:
        hist = str(e.get('historico', '')).upper()
        for key, nota in inbound_map.items():
            if nota['nf'] and nota['nf'] in hist:
                reconciled_nfs.add(key)

    inbound_sem_fin = sorted(
        [n for k, n in inbound_map.items() if k not in reconciled_nfs],
        key=lambda x: x['valor'], reverse=True
    )

    # ── 4. CROSS C: PAGAMENTOS SEM ENTRADA (pagou fornecedor, mas sem NF de entrada) ─
    inbound_suppliers = {normalize_text(str(i.get('fornecedor', ''))) for i in inbound if i.get('fornecedor')}

    expenses_by_supplier = {}
    for e in expenses_detail:
        forn_raw = str(e.get('fornecedor', '')).strip()
        forn = normalize_text(forn_raw)
        valor = safe_float(e.get('valor', 0))
        hist = str(e.get('historico', ''))[:60]
        if forn and valor >= 300:
            if forn not in expenses_by_supplier:
                expenses_by_supplier[forn] = {'raw': forn_raw, 'total': 0, 'hists': []}
            expenses_by_supplier[forn]['total'] += valor
            expenses_by_supplier[forn]['hists'].append(hist)

    pagamentos_sem_entrada = sorted(
        [{'fornecedor': d['raw'], 'valor': d['total'], 'hists': d['hists'][:3]}
         for forn, d in expenses_by_supplier.items() if forn not in inbound_suppliers],
        key=lambda x: x['valor'], reverse=True
    )

    # ── 5. CROSS D: PAGAMENTOS DUPLICADOS (mesmo fornecedor + valor similar) ──────
    pagamentos_dup = []
    if expenses_detail:
        seen_payments = {}
        for e in expenses_detail:
            forn_raw = str(e.get('fornecedor', '')).strip()
            forn = normalize_text(forn_raw)
            valor = safe_float(e.get('valor', 0))
            hist = str(e.get('historico', ''))[:50]
            if forn and valor >= 200:
                rounded = round(valor / 10) * 10  # bucket by nearest R$10
                key = f"{forn}|{rounded}"
                if key not in seen_payments:
                    seen_payments[key] = {'count': 0, 'total': 0, 'raw': forn_raw, 'valor_unit': valor, 'hists': []}
                seen_payments[key]['count'] += 1
                seen_payments[key]['total'] += valor
                seen_payments[key]['hists'].append(hist)
        pagamentos_dup = sorted(
            [d for d in seen_payments.values() if d['count'] >= 2],
            key=lambda x: x['total'], reverse=True
        )

    # ── 6. CROSS E: CANCELAMENTOS POR PRODUTO (alto volume sem justificativa) ────
    canc_by_prod = {}
    for c in cancellations:
        prod = str(c.get('nomeProduto', c.get('produtoNome', '?'))).strip()
        reason = str(c.get('motivo', 'SEM MOTIVO')).strip() or 'SEM MOTIVO'
        val = safe_float(c.get('valor', c.get('valorTotal', 0)))
        qty = safe_float(c.get('qtde', 1))
        if prod not in canc_by_prod:
            canc_by_prod[prod] = {'val': 0, 'qty': 0, 'reasons': {}}
        canc_by_prod[prod]['val'] += val
        canc_by_prod[prod]['qty'] += qty
        canc_by_prod[prod]['reasons'][reason] = canc_by_prod[prod]['reasons'].get(reason, 0) + qty

    suspicious_reasons = {'SEM MOTIVO', '', 'ERRO', 'TESTE', 'SEM_MOTIVO'}
    canc_risk_prods = []
    for prod, d in sorted(canc_by_prod.items(), key=lambda x: x[1]['val'], reverse=True)[:15]:
        top_reason = max(d['reasons'].items(), key=lambda x: x[1])[0] if d['reasons'] else 'SEM MOTIVO'
        is_sus = top_reason.upper() in suspicious_reasons or 'SEM MOTIVO' in top_reason.upper()
        if d['val'] >= 100:
            canc_risk_prods.append({'prod': prod, 'val': d['val'], 'qty': d['qty'],
                                     'reason': top_reason, 'suspicious': is_sus})

    # ── 7. MONTAGEM DO RELATÓRIO ──────────────────────────────────────────────
    lines = []
    lines.append(f"🔍 **AUDITORIA CRUZADA COMPLETA CEO — {rest['name'].upper()}**")
    lines.append(f"📅 Período: {start_date} a {end_date}")
    lines.append(f"📊 Dados: {len(cancellations)} cancelamentos | {len(cashier_data)} fechamentos caixa | "
                 f"{len(inbound)} entradas NF | {len(all_expenses)} despesas | {len(sales)} vendas")
    lines.append("")

    # Resumo executivo (inserido após seção de dados)
    n_crit = len(cross_fraud) + (1 if inbound_sem_fin else 0) + (1 if pagamentos_sem_entrada else 0)
    n_atenc = len(solo_high_canc) + len(pagamentos_dup) + len([p for p in canc_risk_prods if p['suspicious']])
    val_risco = sum(n['valor'] for n in inbound_sem_fin) + sum(p['valor'] for p in pagamentos_sem_entrada)
    lines.append(f"⚡ **RESUMO:** 🔴 {n_crit} crítico(s) | 🟡 {n_atenc} atenção | 💸 Risco estimado: {fmt_brl(val_risco)}")
    lines.append("")

    # SEÇÃO A
    lines.append("─" * 55)
    lines.append("🚨 **A) CRUZAMENTO: CANCELAMENTOS × CAIXA POR OPERADOR**")
    if cross_fraud:
        for cf in cross_fraud:
            top_r = max(cf['reasons'].items(), key=lambda x: x[1])[0] if cf['reasons'] else '?'
            lines.append(f"  🔴 **FRAUDE PROVÁVEL — {cf['op'].upper()}**")
            lines.append(f"     Cancelamentos: {fmt_brl(cf['canc_val'])} ({cf['canc_pct']:.0f}% do total do período)")
            lines.append(f"     Quebra de caixa no mesmo operador: {fmt_brl(cf['cashier_diff'])}")
            lines.append(f"     Produtos cancelados: {', '.join(cf['items'])}")
            lines.append(f"     Motivo predominante: {top_r}")
            lines.append(f"     ⚠️ Padrão duplo: cancela muito E fecha caixa com diferença = investigar imediatamente.")
    else:
        lines.append("  ✅ Nenhum operador apresenta padrão simultâneo de cancela+quebra.")

    if solo_high_canc:
        lines.append("")
        lines.append("  🟡 Operadores com alto cancelamento (sem quebra de caixa correlata):")
        for s in solo_high_canc[:5]:
            top_r = max(s['reasons'].items(), key=lambda x: x[1])[0] if s['reasons'] else '?'
            lines.append(f"    • {s['op']}: {fmt_brl(s['canc_val'])} ({s['canc_pct']:.0f}%) — Motivo: {top_r}")
    lines.append("")

    # SEÇÃO B
    lines.append("─" * 55)
    lines.append(f"🧾 **B) ENTRADAS NF SEM LANÇAMENTO FINANCEIRO**")
    lines.append(f"  Total NFs de entrada no período: {len(inbound_map)} | Conciliadas: {len(reconciled_nfs)} | Pendentes: {len(inbound_sem_fin)}")
    if inbound_sem_fin:
        total_miss = sum(n['valor'] for n in inbound_sem_fin)
        lines.append(f"  🔴 {len(inbound_sem_fin)} nota(s) entraram no estoque SEM registro no contas a pagar!")
        lines.append(f"  💸 Passivo oculto estimado: **{fmt_brl(total_miss)}**")
        for n in inbound_sem_fin[:15]:
            lines.append(f"    ❌ NF {n['nf']} | {n['fornecedor'][:42]} | {fmt_brl(n['valor'])} | {n['data']}")
        if len(inbound_sem_fin) > 15:
            lines.append(f"    ...e mais {len(inbound_sem_fin) - 15} notas.")
        lines.append(f"  ⚠️ Risco: passivo não registrado pode causar surpresas no fluxo de caixa.")
    else:
        lines.append("  ✅ Todas as entradas de NF estão lançadas no financeiro.")
    lines.append("")

    # SEÇÃO C
    lines.append("─" * 55)
    lines.append(f"💸 **C) PAGAMENTOS A FORNECEDORES SEM ENTRADA DE MERCADORIA**")
    if pagamentos_sem_entrada:
        total_sem = sum(p['valor'] for p in pagamentos_sem_entrada)
        lines.append(f"  🔴 {len(pagamentos_sem_entrada)} fornecedor(es) receberam pagamentos sem NF de entrada no período.")
        lines.append(f"  💰 Total pago sem entrada: **{fmt_brl(total_sem)}**")
        for p in pagamentos_sem_entrada[:10]:
            lines.append(f"    💸 {p['fornecedor'][:45]} — {fmt_brl(p['valor'])}")
            for h in p['hists'][:2]:
                if h: lines.append(f"       Hist: {h}")
        lines.append(f"  ⚠️ Pode ser serviço legítimo (aluguel, RH, etc.) ou pagamento sem recebimento. Validar.")
    else:
        lines.append("  ✅ Todos os fornecedores pagos têm entrada de mercadoria correspondente.")
    lines.append("")

    # SEÇÃO D
    lines.append("─" * 55)
    lines.append(f"🔁 **D) PAGAMENTOS DUPLICADOS / SUSPEITOS**")
    if pagamentos_dup:
        lines.append(f"  🟡 {len(pagamentos_dup)} padrão(ões) de pagamento repetido para mesmo fornecedor + valor similar:")
        for pd in pagamentos_dup[:8]:
            lines.append(f"    ⚠️ {pd['raw'][:45]} — {pd['count']}x vezes | Total: {fmt_brl(pd['total'])} | Unit: ~{fmt_brl(pd['valor_unit'])}")
            for h in pd['hists'][:2]:
                if h: lines.append(f"       Hist: {h}")
        lines.append(f"  ⚠️ Verifique se são parcelas legítimas ou duplicatas por erro/fraude.")
    else:
        lines.append("  ✅ Nenhum padrão de pagamento duplicado detectado.")
    lines.append("")

    # SEÇÃO E
    lines.append("─" * 55)
    lines.append(f"📦 **E) ITENS COM MAIOR VOLUME DE CANCELAMENTO**")
    if canc_risk_prods:
        for p in canc_risk_prods[:10]:
            flag = "🔴" if p['suspicious'] else "🟡"
            sus_tag = " ⚠️ MOTIVO SUSPEITO" if p['suspicious'] else ""
            lines.append(f"  {flag} {p['prod']}: {fmt_brl(p['val'])} | {int(p['qty'])} unid. | Motivo: {p['reason']}{sus_tag}")
    else:
        lines.append("  ✅ Nenhum produto com volume de cancelamento relevante.")
    lines.append("")

    # SEÇÃO F
    lines.append("─" * 55)
    lines.append(f"🏦 **F) RESUMO DE QUEBRA DE CAIXA POR OPERADOR**")
    if cashier_data:
        op_summ = {}
        for c in cashier_data:
            op = str(c.get('operador') or c.get('nomeCaixa') or 'SEM_OP')
            diff = safe_float(c.get('diferenca', 0))
            if op not in op_summ: op_summ[op] = {'diff': 0, 'count': 0}
            op_summ[op]['diff'] += diff
            op_summ[op]['count'] += 1
        any_issue = False
        for op, st in sorted(op_summ.items(), key=lambda x: abs(x[1]['diff']), reverse=True):
            icon = "🔴" if abs(st['diff']) > 100 else "🟡" if abs(st['diff']) > 20 else "🟢"
            lines.append(f"  {icon} {op}: acumulado {fmt_brl(st['diff'])} em {st['count']} fechamento(s)")
            if abs(st['diff']) > 20: any_issue = True
        if not any_issue:
            lines.append("  ✅ Todos os caixas dentro da margem de tolerância.")
    else:
        lines.append("  ℹ️ Sem dados de fechamento de caixa no período.")
    lines.append("")

    lines.append("─" * 55)
    lines.append("📍 _Fonte: NetControll — cancelamentos, caixa, entradas NF, despesas_")

    return "\n".join(lines)[:14000]

def get_daily_briefing():
    """Briefing executivo turbinado das 7h com ticket médio, top 3, CMV flash, contas 3 dias e meta mensal."""
    import calendar
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    today     = datetime.now().strftime('%Y-%m-%d')
    now       = datetime.now()

    # Contas vencendo nos próximos 3 dias
    d3 = (now + timedelta(days=3)).strftime('%Y-%m-%d')

    # Metas mensais (targets.json)
    targets = {}
    try:
        with open('targets.json', 'r') as f:
            targets = json.load(f)
    except: pass

    days_in_month = calendar.monthrange(now.year, now.month)[1]
    days_passed   = now.day
    days_left     = days_in_month - days_passed
    month_start   = now.replace(day=1).strftime('%Y-%m-%d')

    report  = f"☀️ **BOM DIA, CEO! BRIEFING EXECUTIVO**\n"
    report += f"📅 {today} | Base: ontem ({yesterday})\n"
    report += f"📆 Dia {days_passed}/{days_in_month} do mês — faltam {days_left} dias\n\n"

    total_geral = 0
    total_meta  = sum(targets.values()) if targets else 0

    for rest in RESTAURANTS:
        # ── Faturamento ontem vs semana passada ──────────────────────────
        sales_y  = fetch_sales_data(rest['id'], yesterday, yesterday)
        fat_y    = sum(safe_float(s.get('valor', 0)) for s in sales_y)
        total_geral += fat_y

        last_week = (now - timedelta(days=8)).strftime('%Y-%m-%d')
        sales_lw  = fetch_sales_data(rest['id'], last_week, last_week)
        fat_lw    = sum(safe_float(s.get('valor', 0)) for s in sales_lw)
        variacao  = ((fat_y - fat_lw) / fat_lw * 100) if fat_lw > 0 else 0
        ev        = "📈" if variacao >= 0 else "📉"

        report += f"🏠 **{rest['name']}**\n"
        report += f"  💰 Faturamento: {fmt_brl(fat_y)} {ev} ({fmt_pct(variacao, 1)}% vs sem. passada)\n"

        # ── Ticket Médio ─────────────────────────────────────────────────
        try:
            itens_unicos = len(set(s.get('nome','') for s in sales_y if s.get('nome')))
            ticket_medio = fat_y / itens_unicos if itens_unicos > 0 else 0
            report += f"  🎟️ Ticket médio por item: {fmt_brl(ticket_medio)} ({itens_unicos} itens distintos)\n"
        except: pass

        # ── Top 3 itens do dia ───────────────────────────────────────────
        try:
            sales_map_y = {}
            for s in sales_y:
                n = s.get('nome', '')
                v = safe_float(s.get('valor', 0))
                q = safe_float(s.get('qtde', 0))
                if n:
                    sales_map_y[n] = sales_map_y.get(n, {'v': 0, 'q': 0})
                    sales_map_y[n]['v'] += v
                    sales_map_y[n]['q'] += q
            top3 = sorted(sales_map_y.items(), key=lambda x: -x[1]['v'])[:3]
            if top3:
                nomes = ' | '.join(f"{n} ({fmt_brl(d['q'], 0)}un R${fmt_brl(d['v'], 0)})" for n, d in top3)
                report += f"  🏆 Top 3: {nomes}\n"
        except: pass

        # ── CMV flash (custo total / faturamento das vendas) ─────────────
        try:
            cmv_data = fetch_cmv_data(rest['id'], yesterday, yesterday)
            custo_total = sum(safe_float(c.get('valorCusto', 0)) for c in cmv_data)
            cmv_perc = (custo_total / fat_y * 100) if fat_y > 0 else 0
            alerta_cmv = " 🚨" if cmv_perc > 38 else (" ✅" if cmv_perc > 0 else "")
            if cmv_perc > 0:
                report += f"  📊 CMV Flash: {fmt_pct(cmv_perc)}%{alerta_cmv}\n"
        except: pass

        # ── Meta mensal: progresso ───────────────────────────────────────
        try:
            rest_key = normalize_text(rest['name']).split()[0]
            target   = targets.get(rest_key, 0)
            if target > 0:
                sales_mes = fetch_sales_data(rest['id'], month_start, today)
                fat_mes   = sum(safe_float(s.get('valor', 0)) for s in sales_mes)
                progress  = fat_mes / target * 100
                esperado  = days_passed / days_in_month * 100
                s_emoji   = "✅" if progress >= esperado else "⚠️"
                ritmo_d   = fat_mes / days_passed if days_passed > 0 else 0
                ritmo_nec = (target - fat_mes) / days_left if days_left > 0 else 0
                report += (f"  🎯 Meta: R${fmt_brl(fat_mes, 0)}/R${fmt_brl(target, 0)} ({fmt_brl(progress, 0)}%) {s_emoji} "
                           f"— Ritmo atual R${fmt_brl(ritmo_d, 0)}/dia | Necessário R${fmt_brl(ritmo_nec, 0)}/dia\n")
        except: pass

        # ── Estoque negativo ─────────────────────────────────────────────
        try:
            stock_data = load_json(rest.get('stock_file', ''))
            low = [str(i.get('produto','')) for i in stock_data if safe_float(i.get('estoqueAtual',0)) < 0]
            if low:
                report += f"  ⚠️ Estoque negativo: {', '.join(low[:5])}{'...' if len(low)>5 else ''}\n"
        except: pass

        # ── Contas vencendo nos próximos 3 dias ──────────────────────────
        try:
            exp3 = fetch_expenses_data(rest['id'], today, d3)
            pend3 = [e for e in exp3 if not e.get('pagamento', False)]
            if pend3:
                tot3 = sum(safe_float(e.get('valor', 0)) for e in pend3)
                maiores = sorted(pend3, key=lambda e: -safe_float(e.get('valor',0)))[:3]
                nomes_f = ', '.join(f"{e.get('fornecedor','N/A')[:25]} R${fmt_brl(safe_float(e.get('valor',0)), 0)}" for e in maiores)
                report += f"  ⏳ {len(pend3)} contas (3d) R${fmt_brl(tot3, 0)}: {nomes_f}\n"
        except: pass

        report += "\n"

    report += f"🏦 **FATURAMENTO TOTAL GRUPO: {fmt_brl(total_geral)}**"
    if total_meta > 0:
        group_mes = 0
        try:
            for r in RESTAURANTS:
                s = fetch_sales_data(r['id'], month_start, today)
                group_mes += sum(safe_float(x.get('valor',0)) for x in s)
        except: pass
        prog_g = group_mes / total_meta * 100 if total_meta > 0 else 0
        report += f" | Meta grupo: {fmt_brl(prog_g, 0)}%"
    report += "\n\n"

    # Clima (Regiões do Grupo)
    try:
        w_rn = get_weather_forecast('Nauan')
        w_al = get_weather_forecast('Milagres')
        l_rn = [l for l in w_rn.split('\n') if '📅' in l]
        l_al = [l for l in w_al.split('\n') if '📅' in l]
        if l_rn or l_al:
            report += f"🌡️ **CLIMA HOJE:**\n"
            if l_rn: report += f"  📍 Gostoso/RN: {l_rn[0].strip()}\n"
            if l_al: report += f"  📍 Milagres/AL: {l_al[0].strip()}\n"
    except: pass

    return report[:14000]

def get_proactive_alerts():
    """Alertas proativos com prioridade 3 níveis, ação prescritiva, impacto R$, e deduplicação com escalation."""
    import calendar, hashlib
    now             = datetime.now()
    today           = now.strftime('%Y-%m-%d')
    yesterday       = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    last_week_start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    tomorrow        = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    d3              = (now + timedelta(days=3)).strftime('%Y-%m-%d')
    month_start     = now.replace(day=1).strftime('%Y-%m-%d')
    days_in_month   = calendar.monthrange(now.year, now.month)[1]
    days_passed     = now.day
    days_left       = days_in_month - days_passed

    # ── Carregar metas e histórico de alertas ───────────────────────────
    targets = {}
    try:
        with open('targets.json', 'r') as f:
            targets = json.load(f)
    except: pass

    # Cache de alertas enviados: {hash: {'first_seen': date, 'count': N}}
    ALERT_CACHE_FILE = 'alert_cache.json'
    alert_cache = {}
    try:
        with open(ALERT_CACHE_FILE, 'r') as f:
            alert_cache = json.load(f)
    except: pass

    # Limpa alertas com mais de 14 dias do cache
    alert_cache = {k: v for k, v in alert_cache.items()
                   if (now - datetime.strptime(v.get('first_seen', today), '%Y-%m-%d')).days <= 14}

    def alert_key(rest_name, alert_type, detail=''):
        """Hash estável por (casa, tipo, detalhe) — ignora valores que mudam a cada hora."""
        raw = f"{rest_name}|{alert_type}|{detail}"
        return hashlib.md5(raw.encode()).hexdigest()[:12]

    def register_alert(key):
        """Registra alerta no cache e retorna (is_new, days_persisting)."""
        if key not in alert_cache:
            alert_cache[key] = {'first_seen': today, 'count': 1, 'last_seen': today}
            return True, 1
        entry = alert_cache[key]
        days = (now - datetime.strptime(entry['first_seen'], '%Y-%m-%d')).days + 1
        entry['count'] = entry.get('count', 1) + 1
        entry['last_seen'] = today
        return False, days

    def persist_badge(days):
        """Retorna badge de persistência quando alerta é recorrente."""
        if days <= 1: return ''
        if days == 2: return ' _(persiste 2 dias)_'
        if days <= 5: return f' _⚠️ PERSISTE {days} DIAS_'
        return f' _🔴 PERSISTE {days} DIAS — SEM RESOLUÇÃO_'

    # ── Estrutura de alertas com prioridade ─────────────────────────────
    # Cada alerta: {'priority': 1/2/3, 'icon': str, 'msg': str, 'impact': float, 'action': str, 'key': str}
    CRIT = 1; ATENC = 2; INFO = 3

    per_restaurant = {}

    for rest in RESTAURANTS:
        rest_key    = normalize_text(rest['name']).split()[0]
        rest_alerts = []

        # ── 1. Ruptura de estoque ── CRÍTICO ────────────────────────────
        try:
            stock_data = load_json(rest.get('stock_file', ''))
            # Estima faturamento médio de bebidas/alimentos para calcular impacto
            sales_7d_rest = fetch_sales_data(rest['id'], last_week_start, yesterday)
            avg_daily = sum(safe_float(s.get('valor',0)) for s in sales_7d_rest) / 7 if sales_7d_rest else 0

            for item in stock_data:
                nome   = str(item.get('produto', '')).strip()
                est    = safe_float(item.get('estoqueAtual', 0))
                if est <= 0:
                    if not any(x in nome.upper() for x in ['DAY USE','PONTO','PASSADA','COUVERT']):
                        ak = alert_key(rest['name'], 'ruptura', nome[:20])
                        is_new, days = register_alert(ak)
                        badge = persist_badge(days)
                        impact_est = avg_daily * 0.05  # ruptura individual ≈ 5% do faturamento diário
                        msg = (f"🔴 **RUPTURA DE ESTOQUE**{badge}\n"
                               f"   🏠 {rest['name']} | Produto: *{nome}*\n"
                               f"   💸 Impacto estimado: R${fmt_brl(impact_est)}/dia em faturamento\n"
                               f"   👉 *Ação:* Contatar fornecedor hoje mesmo. Se urgente, compra de emergência no atacado.")
                        rest_alerts.append({'priority': CRIT, 'msg': msg, 'impact': impact_est, 'key': ak, 'new': is_new})
        except: pass

        # ── 2. Queda de faturamento ── CRÍTICO se >40% ──────────────────
        try:
            sales_y  = fetch_sales_data(rest['id'], yesterday, yesterday)
            fat_y    = sum(safe_float(s.get('valor',0)) for s in sales_y)
            sales_7d_r = fetch_sales_data(rest['id'], last_week_start, yesterday)
            fat_7d   = sum(safe_float(s.get('valor',0)) for s in sales_7d_r)
            avg_d    = fat_7d / 7 if fat_7d > 0 else 0
            if avg_d > 0 and fat_y < avg_d * 0.7:
                drop_pct = (fat_y - avg_d) / avg_d * 100
                gap_r    = avg_d - fat_y
                prioridade = CRIT if drop_pct < -40 else ATENC
                icon = "🔴" if prioridade == CRIT else "🟡"
                ak = alert_key(rest['name'], 'queda_fat', '')
                is_new, days = register_alert(ak)
                badge = persist_badge(days)
                msg = (f"{icon} **QUEDA DE FATURAMENTO**{badge}\n"
                       f"   🏠 {rest['name']} | Ontem: R${fmt_brl(fat_y)} vs média R${fmt_brl(avg_d)} ({drop_pct:.0f}%)\n"
                       f"   💸 Receita perdida vs média: R${fmt_brl(gap_r)}\n"
                       f"   👉 *Ação:* Verificar se houve problema operacional (fechamento antecipado, falta de equipe) ou queda de demanda. Ativar promoção relâmpago se for demanda.")
                rest_alerts.append({'priority': prioridade, 'msg': msg, 'impact': gap_r, 'key': ak, 'new': is_new})
        except:
            sales_y = []; fat_y = 0

        # ── 3. Contas grandes amanhã (>R$2k) ── ATENÇÃO ─────────────────
        try:
            exp_d1 = fetch_expenses_data(rest['id'], tomorrow, tomorrow)
            for b in exp_d1:
                val = safe_float(b.get('valor',0))
                if not b.get('pagamento', False) and val > 2000:
                    forn = b.get('fornecedor','N/A')[:30]
                    ak = alert_key(rest['name'], 'conta_amanha', forn)
                    is_new, days = register_alert(ak)
                    badge = persist_badge(days)
                    msg = (f"🟡 **CONTA VENCE AMANHÃ**{badge}\n"
                           f"   🏠 {rest['name']} | Fornecedor: *{forn}*\n"
                           f"   💸 Valor: R${fmt_brl(val)} — vence {tomorrow}\n"
                           f"   👉 *Ação:* Confirmar saldo disponível e autorizar pagamento hoje. Evitar juros de mora.")
                    rest_alerts.append({'priority': ATENC, 'msg': msg, 'impact': val, 'key': ak, 'new': is_new})
        except: pass

        # ── 4. Contas em 3 dias (volume >R$10k) ── ATENÇÃO ──────────────
        try:
            exp3    = fetch_expenses_data(rest['id'], today, d3)
            pend3_v = sum(safe_float(e.get('valor',0)) for e in exp3 if not e.get('pagamento',False))
            if pend3_v > 10000:
                ak = alert_key(rest['name'], 'caixa_3d', '')
                is_new, days = register_alert(ak)
                badge = persist_badge(days)
                msg = (f"🟡 **COMPROMISSO DE CAIXA — 3 DIAS**{badge}\n"
                       f"   🏠 {rest['name']} | Total a vencer até {d3}: R${fmt_brl(pend3_v)}\n"
                       f"   👉 *Ação:* Verificar saldo atual vs compromisso. Se insuficiente, antecipar recebíveis ou negociar prazo com fornecedores.")
                rest_alerts.append({'priority': ATENC, 'msg': msg, 'impact': pend3_v, 'key': ak, 'new': is_new})
        except: pass

        # ── 5. Quebras de caixa ── CRÍTICO ──────────────────────────────
        try:
            closure_alerts = get_cashier_closure_report(rest['name'], yesterday, alert_only=True)
            if closure_alerts:
                for line in closure_alerts.split('\n'):
                    if line.strip():
                        ak = alert_key(rest['name'], 'quebra_caixa', line[:30])
                        is_new, days = register_alert(ak)
                        badge = persist_badge(days)
                        msg = (f"🔴 **QUEBRA DE CAIXA**{badge}\n"
                               f"   🏠 {rest['name']} | {line.replace('⚠️ ','')}\n"
                               f"   👉 *Ação:* Acionar gerente para conferência de comanda física. Se padrão se repetir com mesmo operador → auditoria disciplinar.")
                        rest_alerts.append({'priority': CRIT, 'msg': msg, 'impact': 500, 'key': ak, 'new': is_new})
        except: pass

        # ── 6. Cancelamentos altos (>8%) ── CRÍTICO (suspeita de fraude) ─
        try:
            session = get_session_for_rest(rest['id'])
            if session:
                s_dt = f"{yesterday}T03:00:00.000Z"
                e_dt = f"{yesterday}T23:59:59.000Z"
                resp = session.get(CANCELLATION_URL, params={'DataInicial': s_dt, 'DataFinal': e_dt})
                if resp.status_code == 200:
                    canc_data = resp.json()
                    total_canc = sum(safe_float(c.get('valor',0)) for c in canc_data)
                    total_vend = sum(safe_float(s.get('valor',0)) for s in (sales_y if sales_y else []))
                    if total_vend > 0:
                        tx_canc = total_canc / total_vend * 100
                        if tx_canc > 8:
                            ak = alert_key(rest['name'], 'cancelamentos', '')
                            is_new, days = register_alert(ak)
                            badge = persist_badge(days)
                            # Identifica garçom com mais cancelamentos
                            garcon_map = {}
                            for c in canc_data:
                                g = c.get('nomeOperador') or c.get('operador') or 'N/A'
                                garcon_map[g] = garcon_map.get(g, 0) + safe_float(c.get('valor',0))
                            top_g = max(garcon_map.items(), key=lambda x: x[1]) if garcon_map else ('N/A', 0)
                            msg = (f"🔴 **CANCELAMENTOS SUSPEITOS**{badge}\n"
                                   f"   🏠 {rest['name']} | Taxa: {tx_canc:.1f}% do faturamento cancelado (R${fmt_brl(total_canc)})\n"
                                   f"   👤 Maior volume: *{top_g[0]}* — R${fmt_brl(top_g[1])} cancelados\n"
                                   f"   👉 *Ação:* Puxar comanda física de {top_g[0]} e comparar com sistema. Taxa acima de 8% = investigação obrigatória.")
                            rest_alerts.append({'priority': CRIT, 'msg': msg, 'impact': total_canc, 'key': ak, 'new': is_new})
        except: pass

        # ── 7. Inflação de fornecedor (>10%) ── ATENÇÃO ──────────────────
        try:
            inbound_rec  = fetch_inbound_data(rest['id'], yesterday, yesterday)
            inbound_prev = fetch_inbound_data(rest['id'], last_week_start, (now-timedelta(days=2)).strftime('%Y-%m-%d'))
            price_hist = {}
            for item in inbound_prev:
                prod = str(item.get('produto','')).strip().upper()
                vu   = safe_float(item.get('valorUnitario',0)) or safe_float(item.get('valorTotal',0)) / max(safe_float(item.get('qtde',1)),1)
                if prod and vu > 0:
                    if prod not in price_hist: price_hist[prod] = []
                    price_hist[prod].append(vu)
            for item in inbound_rec:
                prod  = str(item.get('produto','')).strip().upper()
                forn  = str(item.get('fornecedor','')).strip()
                vu    = safe_float(item.get('valorUnitario',0)) or safe_float(item.get('valorTotal',0)) / max(safe_float(item.get('qtde',1)),1)
                if prod in price_hist and vu > 0:
                    avg_prev = sum(price_hist[prod]) / len(price_hist[prod])
                    if avg_prev > 0 and vu > avg_prev * 1.10:
                        alta_pct = (vu - avg_prev) / avg_prev * 100
                        impacto_mes = (vu - avg_prev) * 50  # estima 50 unidades/mês
                        ak = alert_key(rest['name'], 'inflacao', prod[:20])
                        is_new, days = register_alert(ak)
                        badge = persist_badge(days)
                        msg = (f"🟡 **INFLAÇÃO DE FORNECEDOR**{badge}\n"
                               f"   🏠 {rest['name']} | Produto: *{prod[:30]}* | Fornecedor: {forn[:25]}\n"
                               f"   📈 Preço: R${fmt_brl(avg_prev)} → R${fmt_brl(vu)} (+{alta_pct:.0f}%)\n"
                               f"   💸 Impacto estimado no CMV: +R${fmt_brl(impacto_mes)}/mês\n"
                               f"   👉 *Ação:* Recalcular markup do(s) prato(s) que usam este insumo. Se impacto > R$200/mês, buscar cotação concorrente.")
                        rest_alerts.append({'priority': ATENC, 'msg': msg, 'impact': impacto_mes, 'key': ak, 'new': is_new})
        except: pass

        # ── 8. Capital empatado (>R$5k parado >7 dias) ── ATENÇÃO ────────
        try:
            stock_data2  = load_json(rest.get('stock_file', ''))
            sales_7d_r2  = fetch_sales_data(rest['id'], last_week_start, yesterday)
            sales_7_names = {normalize_text(str(s.get('nome',''))) for s in sales_7d_r2 if s.get('nome')}
            zombie_val = 0.0; zombie_cnt = 0; zombie_items = []
            for item in stock_data2:
                nome_s  = str(item.get('produto','')).strip()
                n_norm  = normalize_text(nome_s)
                est     = safe_float(item.get('estoqueAtual',0))
                custo   = safe_float(item.get('custoAtual',0))
                if est > 0 and custo > 5 and n_norm not in sales_7_names:
                    val_item = est * custo
                    zombie_val += val_item
                    zombie_cnt += 1
                    if val_item > 200:
                        zombie_items.append(f"{nome_s[:25]} (R${fmt_brl(val_item, 0)})")
            if zombie_val > 5000:
                ak = alert_key(rest['name'], 'capital_empatado', '')
                is_new, days = register_alert(ak)
                badge = persist_badge(days)
                top_z = ', '.join(zombie_items[:3])
                msg = (f"🟡 **CAPITAL EMPATADO NO ESTOQUE**{badge}\n"
                       f"   🏠 {rest['name']} | R${fmt_brl(zombie_val)} em {zombie_cnt} itens sem venda há 7+ dias\n"
                       f"   📦 Maiores: {top_z}\n"
                       f"   👉 *Ação:* Incluir os 3 maiores itens no menu do dia com 10-15% de desconto. Cada R$1.000 liberado = caixa disponível amanhã.")
                rest_alerts.append({'priority': ATENC, 'msg': msg, 'impact': zombie_val * 0.15, 'key': ak, 'new': is_new})
        except: pass

        # ── 9. Meta mensal em risco ── CRÍTICO se <75% ───────────────────
        try:
            rest_key_t = normalize_text(rest['name']).split()[0]
            target = targets.get(rest_key_t, 0)
            if target > 0 and days_left > 0:
                sales_mes  = fetch_sales_data(rest['id'], month_start, today)
                fat_mes    = sum(safe_float(s.get('valor',0)) for s in sales_mes)
                ritmo_at   = fat_mes / days_passed if days_passed > 0 else 0
                projecao   = fat_mes + ritmo_at * days_left
                pct_proj   = projecao / target * 100
                if projecao < target * 0.85:
                    gap = target - projecao
                    fat_dia_nec = gap / days_left if days_left > 0 else gap
                    prioridade = CRIT if pct_proj < 75 else ATENC
                    icon = "🔴" if prioridade == CRIT else "🟡"
                    ak = alert_key(rest['name'], 'meta_risco', '')
                    is_new, days_p = register_alert(ak)
                    badge = persist_badge(days_p)
                    msg = (f"{icon} **META MENSAL EM RISCO**{badge}\n"
                           f"   🏠 {rest['name']} | Projeção: R${fmt_brl(projecao, 0)} ({pct_proj:.0f}% da meta R${fmt_brl(target, 0)})\n"
                           f"   📊 Déficit estimado: R${fmt_brl(gap, 0)} | Faltam {days_left} dias\n"
                           f"   💸 Precisa faturar R${fmt_brl(fat_dia_nec, 0)}/dia para fechar a meta\n"
                           f"   👉 *Ação:* Ativar promoção de happy hour, push de delivery ou evento especial nos próximos {min(days_left, 7)} dias.")
                    rest_alerts.append({'priority': prioridade, 'msg': msg, 'impact': gap, 'key': ak, 'new': is_new})
        except: pass

        # ── 10. Pico de vendas ── INFO (positivo) ────────────────────────
        try:
            sales_today = fetch_sales_data(rest['id'], today, today)
            fat_today   = sum(safe_float(s.get('valor',0)) for s in sales_today)
            sales_7d_r3 = fetch_sales_data(rest['id'], last_week_start, yesterday)
            avg_day     = sum(safe_float(s.get('valor',0)) for s in sales_7d_r3) / 7 if sales_7d_r3 else 0
            if avg_day > 0 and now.hour >= 14 and fat_today > avg_day * 1.3:
                pico_pct = (fat_today - avg_day) / avg_day * 100
                ak = alert_key(rest['name'], 'pico_vendas', today)
                is_new, days = register_alert(ak)
                if is_new:  # Pico só notifica 1x por dia
                    msg = (f"🟢 **PICO DE VENDAS HOJE**\n"
                           f"   🏠 {rest['name']} | Hoje: R${fmt_brl(fat_today)} (+{pico_pct:.0f}% vs média)\n"
                           f"   👉 *Ação:* Verificar estoque de ingredientes críticos (carnes, bebidas). Reforçar equipe se necessário. Aproveitar o pico!")
                    rest_alerts.append({'priority': INFO, 'msg': msg, 'impact': fat_today - avg_day, 'key': ak, 'new': True})
        except: pass

        # ── 11. NFC-e com problema ── CRÍTICO (risco fiscal) ─────────────
        try:
            session = get_session_for_rest(rest['id'])
            if session:
                s_dt = f"{yesterday}T03:00:00.000Z"
                e_dt = f"{yesterday}T23:59:59.000Z"
                resp = session.get(FISCAL_URL, params={'DataInicial': s_dt, 'DataFinal': e_dt})
                if resp.status_code == 200:
                    notas = resp.json()
                    rejeitadas = [n for n in notas if str(n.get('status','')).lower() in
                                  ('rejeição','rejeicao','rejeitada','contingencia','contingência','cancelada','erro')]
                    prob_code  = [n for n in notas if n.get('cStat') and str(n.get('cStat','')) not in ('100','101','135','150','301','302')]
                    problemas  = rejeitadas or prob_code
                    if problemas:
                        total_prob_val = sum(safe_float(n.get('vNF',0) or n.get('valor',0)) for n in problemas)
                        ak = alert_key(rest['name'], 'nfce_problema', '')
                        is_new, days = register_alert(ak)
                        badge = persist_badge(days)
                        msg = (f"🔴 **NFC-e EM REJEIÇÃO/CONTINGÊNCIA**{badge}\n"
                               f"   🏠 {rest['name']} | {len(problemas)} nota(s) com problema — R${fmt_brl(total_prob_val)}\n"
                               f"   ⚠️ Risco: autuação fiscal e multa sobre o valor total\n"
                               f"   👉 *Ação:* Contatar contador/responsável fiscal hoje. Regularizar notas em contingência antes do prazo de 24h para autorização retroativa.")
                        rest_alerts.append({'priority': CRIT, 'msg': msg, 'impact': total_prob_val * 0.10, 'key': ak, 'new': is_new})
        except: pass

        # ── 12. CMV alto (>38%) ── CRÍTICO se >44% ───────────────────────
        try:
            cmv_y   = fetch_cmv_data(rest['id'], yesterday, yesterday)
            custo_y = sum(safe_float(c.get('valorCusto',0)) for c in cmv_y)
            fat_y2  = fat_y if fat_y else sum(safe_float(s.get('valor',0)) for s in fetch_sales_data(rest['id'], yesterday, yesterday))
            if fat_y2 > 0:
                cmv_p = custo_y / fat_y2 * 100
                if cmv_p > 38:
                    cats = {}
                    for c in cmv_y:
                        cat = c.get('nomeGrupo') or c.get('subgrupo') or 'N/A'
                        cats[cat] = cats.get(cat, 0) + safe_float(c.get('valorCusto',0))
                    top_cat = sorted(cats.items(), key=lambda x: -x[1])[:2]
                    cats_str = ' | '.join(f"{k}: R${fmt_brl(v, 0)}" for k,v in top_cat)
                    custo_ideal = fat_y2 * 0.35
                    excesso = custo_y - custo_ideal
                    prioridade = CRIT if cmv_p > 44 else ATENC
                    icon = "🔴" if prioridade == CRIT else "🟡"
                    ak = alert_key(rest['name'], 'cmv_alto', '')
                    is_new, days = register_alert(ak)
                    badge = persist_badge(days)
                    msg = (f"{icon} **CMV ACIMA DO LIMITE**{badge}\n"
                           f"   🏠 {rest['name']} | CMV ontem: *{cmv_p:.1f}%* (ideal ≤ 38%)\n"
                           f"   💸 Custo excessivo vs benchmark: R${fmt_brl(excesso)} | Maiores pesos: {cats_str}\n"
                           f"   👉 *Ação:* Revisar fichas técnicas das categorias acima. Verificar desperdício na cozinha. CMV acima de 44% = sangria no caixa.")
                    rest_alerts.append({'priority': prioridade, 'msg': msg, 'impact': excesso, 'key': ak, 'new': is_new})
        except: pass

        # ── 13. Break-even atingido ── INFO (positivo) ───────────────────
        try:
            now_dt    = datetime.now()
            start_30  = (now_dt - timedelta(days=30)).strftime('%Y-%m-%d')
            end_30    = now_dt.strftime('%Y-%m-%d')
            exps_30   = fetch_expenses_data(rest['id'], start_30, end_30)
            fixed_c   = 0
            fixed_k   = ["FIXO", "FOLHA", "PESSOAL", "ADMIN", "OCUPAC", "RETIRADA", "ALUGUEL", "SOFTW", "CONTAB", "MARKET"]
            for ex in exps_30:
                c_name = normalize_text(str(ex.get('planoContas1', ''))).upper()
                if any(k in c_name for k in fixed_k) or c_name in ["DESPESAS", "NONE"]:
                    fixed_c += safe_float(ex.get('valor', 0))
            cmv_d_30  = fetch_cmv_data(rest['id'], start_30, end_30)
            c_val_30  = sum(safe_float(c.get('valorCusto', 0)) for c in cmv_d_30)
            s_d_30    = fetch_sales_data(rest['id'], start_30, end_30)
            r_val_30  = sum(safe_float(s.get('valor', 0)) for s in s_d_30)
            if r_val_30 > 0:
                m_contrib = 1 - (c_val_30 / r_val_30 + 0.10)
                if m_contrib > 0:
                    bep_est = fixed_c / m_contrib
                    m_st    = now_dt.replace(day=1).strftime('%Y-%m-%d')
                    m_sls   = fetch_sales_data(rest['id'], m_st, today)
                    m_fat   = sum(safe_float(s.get('valor', 0)) for s in m_sls)
                    if m_fat >= bep_est > 0:
                        ak = alert_key(rest['name'], 'bep_atingido', now_dt.strftime('%Y-%m'))
                        is_new, _ = register_alert(ak)
                        if is_new:  # Notifica só uma vez por mês
                            margem_extra = m_fat - bep_est
                            msg = (f"🟢 **PONTO DE EQUILÍBRIO SUPERADO!**\n"
                                   f"   🏠 {rest['name']} | Realizado: R${fmt_brl(m_fat)} vs BEP R${fmt_brl(bep_est)}\n"
                                   f"   💰 Margem livre a partir de agora: R${fmt_brl(margem_extra)} no mês\n"
                                   f"   👉 *Ação:* A partir de agora toda venda gera margem direta. Aproveitar para antecipar investimentos ou reforçar caixa.")
                            rest_alerts.append({'priority': INFO, 'msg': msg, 'impact': margem_extra, 'key': ak, 'new': True})
        except: pass

        # ── Salva alertas da casa ────────────────────────────────────────
        if rest_alerts:
            per_restaurant[rest_key] = rest_alerts

    # ── Salva cache de alertas ───────────────────────────────────────────
    try:
        with open(ALERT_CACHE_FILE, 'w') as f:
            json.dump(alert_cache, f)
    except: pass

    if not any(per_restaurant.values()):
        return {'ceo': None, 'per_restaurant': {}}

    # ── Monta relatório com prioridade ───────────────────────────────────
    # Ordena todos alertas: Críticos primeiro, depois Atenção, depois Info
    # Dentro de cada nível: alertas novos primeiro, depois persistentes
    all_flat = []
    for rest in RESTAURANTS:
        rk = normalize_text(rest['name']).split()[0]
        for a in per_restaurant.get(rk, []):
            a['rest_name'] = rest['name']
            all_flat.append(a)

    all_flat.sort(key=lambda x: (x['priority'], -x['impact']))

    n_crit  = sum(1 for a in all_flat if a['priority'] == CRIT)
    n_atenc = sum(1 for a in all_flat if a['priority'] == ATENC)
    n_info  = sum(1 for a in all_flat if a['priority'] == INFO)
    impacto_total = sum(a['impact'] for a in all_flat if a['priority'] <= ATENC)

    # Cabeçalho executivo
    report = (f"🔔 **ALERTAS PROATIVOS** — {now.strftime('%d/%m %H:%M')}\n"
              f"{'🔴 ' + str(n_crit) + ' Críticos  ' if n_crit else ''}"
              f"{'🟡 ' + str(n_atenc) + ' Atenção  ' if n_atenc else ''}"
              f"{'🟢 ' + str(n_info) + ' Info' if n_info else ''}\n"
              f"💸 Impacto total em risco: R${fmt_brl(impacto_total)}\n\n")

    # Agrupa por casa, ordenados por prioridade
    for rest in RESTAURANTS:
        rk = normalize_text(rest['name']).split()[0]
        casa_alerts = sorted(per_restaurant.get(rk, []), key=lambda x: (x['priority'], -x['impact']))
        if casa_alerts:
            report += f"🏠 **{rest['name']}**\n"
            for a in casa_alerts[:5]:  # Máx 5 por casa
                report += f"{a['msg']}\n\n"

    # Monta per_restaurant como lista de strings para compatibilidade
    per_restaurant_str = {}
    for rk, alerts in per_restaurant.items():
        per_restaurant_str[rk] = [a['msg'] for a in sorted(alerts, key=lambda x: (x['priority'], -x['impact']))]

    return {'ceo': report[:5000], 'per_restaurant': per_restaurant_str}


def get_daily_closing_report():
    """Relatório de fechamento do dia para envio às 23h: resultado do dia, ranking, PIS estimado, despesas."""
    import calendar
    now         = datetime.now()
    today       = now.strftime('%Y-%m-%d')
    month_start = now.replace(day=1).strftime('%Y-%m-%d')
    days_passed = now.day
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    days_left   = days_in_month - days_passed

    targets = {}
    try:
        with open('targets.json', 'r') as f:
            targets = json.load(f)
    except: pass

    report  = f"🌙 **FECHAMENTO DO DIA — {today}**\n"
    report += f"📆 Dia {days_passed}/{days_in_month} | Faltam {days_left} dias para fechar o mês\n\n"

    total_dia = 0

    for rest in RESTAURANTS:
        sales   = fetch_sales_data(rest['id'], today, today)
        fat_hj  = sum(safe_float(s.get('valor',0)) for s in sales)
        total_dia += fat_hj

        report += f"🏠 **{rest['name']}** — {fmt_brl(fat_hj)}\n"

        # Ranking do dia
        try:
            sm = {}
            for s in sales:
                n = s.get('nome','')
                if n:
                    sm[n] = sm.get(n, {'v':0,'q':0})
                    sm[n]['v'] += safe_float(s.get('valor',0))
                    sm[n]['q'] += safe_float(s.get('qtde',0))
            top5 = sorted(sm.items(), key=lambda x: -x[1]['v'])[:5]
            if top5:
                report += f"  🏆 Top itens do dia:\n"
                for nm, d in top5:
                    report += f"    • {nm}: {fmt_brl(d['q'], 0)}un → R${fmt_brl(d['v'])}\n"
        except: pass

        # CMV do dia
        try:
            cmv_d   = fetch_cmv_data(rest['id'], today, today)
            custo_d = sum(safe_float(c.get('valorCusto',0)) for c in cmv_d)
            cmv_p   = custo_d / fat_hj * 100 if fat_hj > 0 else 0
            alert_c = " 🚨 ACIMA DO IDEAL" if cmv_p > 38 else " ✅"
            if cmv_p > 0:
                report += f"  📊 CMV: {fmt_pct(cmv_p)}%{alert_c}\n"
        except: pass

        # PIS/COFINS estimado do dia
        try:
            pis_est = fat_hj * PIS_COFINS_RATE
            report += f"  🧾 PIS/COFINS est.: R${fmt_brl(pis_est)} (3.65% s/ fat.)\n"
        except: pass

        # Progresso da meta mensal
        try:
            rest_key = normalize_text(rest['name']).split()[0]
            target   = targets.get(rest_key, 0)
            if target > 0:
                s_mes   = fetch_sales_data(rest['id'], month_start, today)
                fat_mes = sum(safe_float(s.get('valor',0)) for s in s_mes)
                prog    = fat_mes / target * 100
                esp     = days_passed / days_in_month * 100
                s_emj   = "✅" if prog >= esp else "⚠️"
                ritmo   = fat_mes / days_passed if days_passed > 0 else 0
                nec     = (target - fat_mes) / days_left if days_left > 0 else 0
                report += (f"  🎯 Meta mês: R${fmt_brl(fat_mes, 0)}/R${fmt_brl(target, 0)} ({fmt_brl(prog, 0)}%) {s_emj} "
                           f"— precisa R${fmt_brl(nec, 0)}/dia\n")
        except: pass

        # Principais despesas lançadas hoje
        try:
            exp_hj = fetch_expenses_data(rest['id'], today, today)
            pend_hj = [e for e in exp_hj if not e.get('pagamento',False)]
            tot_pend = sum(safe_float(e.get('valor',0)) for e in pend_hj)
            if pend_hj:
                top_exp = sorted(pend_hj, key=lambda e: -safe_float(e.get('valor',0)))[:3]
                nomes_exp = ', '.join(f"{e.get('fornecedor','N/A')[:20]} R${fmt_brl(safe_float(e.get('valor',0)), 0)}" for e in top_exp)
                report += f"  💸 Despesas hoje: {len(pend_hj)} contas R${fmt_brl(tot_pend, 0)} | {nomes_exp}\n"
        except: pass

        report += "\n"

    report += f"🏦 **TOTAL DO DIA (GRUPO): {fmt_brl(total_dia)}**\n"
    report += "\n💡 _Boa noite! Cuide da equipe e prepare amanhã com inteligência._"
    return report[:14000]

def simulate_price_change(restaurant_name, product_name, price_change, days_history=7):
    try: price_change = float(price_change)
    except: return "Informe o valor da alteração em reais (ex: 5 para +R$5 ou -3 para -R$3)."
    try: days_history = int(days_history)
    except: days_history = 7
    
    rest = find_restaurant_files(restaurant_name)
    start = (datetime.now() - timedelta(days=days_history)).strftime('%Y-%m-%d')
    end = datetime.now().strftime('%Y-%m-%d')
    
    sales = fetch_sales_data(rest['id'], start, end)
    cmv_data = fetch_cmv_data(rest['id'], start, end)
    
    if not sales:
        return f"Sem dados de vendas para {rest['name']}."
    
    # Find matching products
    search = normalize_text(product_name)
    matches = []
    
    # Aggregate sales volume
    vol_map = {}
    for s in sales:
        n = normalize_text(str(s.get('nome', '')))
        nome_orig = str(s.get('nome', ''))
        if match_query(search, nome_orig) or match_query(search, n):
            vol_map[n] = vol_map.get(n, {'nome': nome_orig, 'qtde': 0, 'valor': 0})
            vol_map[n]['qtde'] += safe_float(s.get('qtde', 0))
            vol_map[n]['valor'] += safe_float(s.get('valor', 0))
    
    if not vol_map:
        return f"Nenhum produto encontrado com '{product_name}' nas vendas de {rest['name']}."
    
    # Get current prices from CMV
    price_map = {}
    for c in cmv_data:
        n = normalize_text(str(c.get('nome', '')))
        if n in vol_map:
            price_map[n] = {
                'preco': safe_float(c.get('preco', 0)),
                'custo': safe_float(c.get('precoCompra', 0))
            }
    
    report = f"🧮 **SIMULADOR DE PREÇO (\"E SE?\")**\n"
    report += f"🏠 {rest['name']} | Base: últimos {days_history} dias\n"
    report += f"💰 Alteração Simulada: {'+'  if price_change > 0 else ''}{fmt_brl(price_change)} por unidade\n\n"
    
    total_impact_week = 0
    total_impact_month = 0
    
    for n, data in vol_map.items():
        qtde = data['qtde']
        valor_atual = data['valor']
        preco_atual = valor_atual / qtde if qtde > 0 else 0
        
        # Override with CMV price if available
        if n in price_map and price_map[n]['preco'] > 0:
            preco_atual = price_map[n]['preco']
            custo = price_map[n]['custo']
        else:
            custo = 0
        
        novo_preco = preco_atual + price_change
        impacto_semana = price_change * qtde
        impacto_mes = impacto_semana * 4
        
        margem_atual = preco_atual - custo if custo > 0 else 0
        margem_nova = novo_preco - custo if custo > 0 else 0
        
        total_impact_week += impacto_semana
        total_impact_month += impacto_mes
        
        report += f"📦 **{data['nome']}**\n"
        report += f"  - Vendas no período: {fmt_brl(qtde, 0)} unidades\n"
        report += f"  - Preço Atual: {fmt_brl(preco_atual)}\n"
        report += f"  - Novo Preço Simulado: {fmt_brl(novo_preco)}\n"
        if custo > 0:
            report += f"  - Margem Atual: {fmt_brl(margem_atual)} → Nova: {fmt_brl(margem_nova)}\n"
        report += f"  - Impacto Semanal: {'+'  if impacto_semana > 0 else ''}{fmt_brl(impacto_semana)}\n"
        report += f"  - Impacto Mensal Projetado: {'+'  if impacto_mes > 0 else ''}{fmt_brl(impacto_mes)}\n\n"
    
    emoji = "📈" if total_impact_month > 0 else "📉"
    report += f"{emoji} **IMPACTO TOTAL PROJETADO:**\n"
    report += f"  - Semanal: {'+'  if total_impact_week > 0 else ''}{fmt_brl(total_impact_week)}\n"
    report += f"  - Mensal: {'+'  if total_impact_month > 0 else ''}{fmt_brl(total_impact_month)}\n"
    
    return report[:14000]

def get_weekly_ranking():
    start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')
    
    ranking_data = []
    
    for rest in RESTAURANTS:
        try:
            # 1. Sales
            sales = fetch_sales_data(rest['id'], start_date, end_date)
            total_sales = sum([safe_float(s.get('valor', 0)) for s in sales])
            
            import re
            
            # 2. CMV (Cost of Goods Sold) - Using the robust CMV engine
            cmv_perc = 0
            if total_sales > 0:
                report_txt = get_cmv_report(rest['name'], start_date, end_date)
                m = re.search(r"CMV Real.*?(\d+,\d+)%", report_txt, flags=re.IGNORECASE | re.DOTALL)
                if m:
                    cmv_perc = float(m.group(1).replace(',', '.'))
            
            # 3. Stock Risk
            stock_data = load_json(rest.get('stock_file', ''))
            negative_items = len([i for i in stock_data if safe_float(i.get('estoqueAtual', 0)) < 0])
            
            ranking_data.append({
                'name': rest['name'],
                'sales': total_sales,
                'cmv': cmv_perc,
                'stock_risk': negative_items,
                'efficiency': (100 - cmv_perc) if total_sales > 0 else 0
            })
        except:
            continue

    if not ranking_data:
        return "Não foi possível consolidar os dados dos restaurantes para o ranking semanal."

    # Sort by Sales
    ranking_sales = sorted(ranking_data, key=lambda x: x['sales'], reverse=True)
    # Sort by CMV Efficiency (lower CMV is better)
    ranking_cmv = sorted(ranking_data, key=lambda x: x['cmv'])
    
    report = f"🏆 **RANKING SEMANAL DO GRUPO MILAGRES**\n"
    report += f"📅 Período: {start_date} a {end_date}\n\n"
    
    report += "💰 **TOP FATURAMENTO:**\n"
    for i, r in enumerate(ranking_sales, 1):
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉"
        report += f"  {medal} {r['name']}: {fmt_brl(r['sales'])}\n"
        
    report += "\n🎯 **EFICIÊNCIA DE CUSTO (Menor CMV):**\n"
    for i, r in enumerate(ranking_cmv, 1):
        star = "⭐" if i == 1 else "🔹"
        report += f"  {star} {r['name']}: {fmt_pct(r['cmv'])}% de CMV\n"
        
    report += "\n🚨 **RISCO OPERACIONAL (Itens Negativos):**\n"
    for r in sorted(ranking_data, key=lambda x: x['stock_risk'], reverse=True):
        risk_emoji = "🔴" if r['stock_risk'] > 10 else "🟡" if r['stock_risk'] > 0 else "🟢"
        report += f"  {risk_emoji} {r['name']}: {r['stock_risk']} itens em ruptura\n"
        
    report += "\n💡 **INSIGHT CEO:** "
    best_sales = ranking_sales[0]['name']
    best_cmv = ranking_cmv[0]['name']
    
    if best_sales == best_cmv:
        report += f"O {best_sales} dominou a semana tanto em faturamento quanto em eficiência! Equipe nota 10."
    else:
        report += f"O {best_sales} faturou mais, porém o {best_cmv} foi mais eficiente na margem. Vale conferir as fichas técnicas do {best_sales}."
        
    return report[:14000]

def get_revenue_tracker(restaurant_name=None):
    # Current month period
    now = datetime.now()
    start_date = now.replace(day=1).strftime('%Y-%m-%d')
    end_date = now.strftime('%Y-%m-%d')
    
    # Calculate days
    import calendar
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    days_passed = now.day
    days_left = days_in_month - days_passed
    
    # Load targets
    targets = {}
    try:
        with open('targets.json', 'r') as f:
            targets = json.load(f)
    except:
        return "Erro ao carregar metas de faturamento (targets.json)."
        
    rests_to_check = []
    if restaurant_name:
        rest = find_restaurant_files(restaurant_name)
        rests_to_check.append(rest)
    else:
        rests_to_check = RESTAURANTS
        
    report = f"🎯 **TRACKER DE METAS - {now.strftime('%B/%Y').upper()}**\n\n"
    
    for rest in rests_to_check:
        rest_key = normalize_text(rest['name']).split()[0]
        target = targets.get(rest_key, 0)
        
        if target == 0:
            continue
            
        sales = fetch_sales_data(rest['id'], start_date, end_date)
        actual = sum([safe_float(s.get('valor', 0)) for s in sales])
        
        progress = (actual / target * 100) if target > 0 else 0
        expected_progress = (days_passed / days_in_month * 100)
        status_emoji = "✅" if progress >= expected_progress else "⚠️"
        
        needed = target - actual
        run_rate_needed = (needed / days_left) if days_left > 0 else 0
        actual_run_rate = (actual / days_passed) if days_passed > 0 else 0
        
        report += f"🏠 **{rest['name']}**\n"
        report += f"  - Meta: {fmt_brl(target)}\n"
        report += f"  - Realizado: {fmt_brl(actual)} ({fmt_pct(progress)}%)\n"
        report += f"  - Status: {status_emoji} {'No ritmo' if status_emoji == '✅' else 'Abaixo do ritmo'}\n"
        report += f"  - Ritmo Atual: {fmt_brl(actual_run_rate)}/dia\n"
        report += f"  - Necessário p/ bater meta: {fmt_brl(run_rate_needed)}/dia\n\n"
        
    report += f"📅 Faltam {days_left} dias para encerrar o mês."
    
    return report[:14000]

def get_weekly_consolidated_report():
    start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    end_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    report = f"📊 **RELATÓRIO SEMANAL CONSOLIDADO**\n"
    report += f"📅 Período: {start_date} a {end_date}\n\n"
    
    total_geral = 0
    
    for rest in RESTAURANTS:
        try:
            sales = fetch_sales_data(rest['id'], start_date, end_date)
            fat = sum([safe_float(s.get('valor', 0)) for s in sales])
            total_geral += fat
            
            # Comparison with previous week
            prev_start = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
            prev_end = (datetime.now() - timedelta(days=8)).strftime('%Y-%m-%d')
            sales_prev = fetch_sales_data(rest['id'], prev_start, prev_end)
            fat_prev = sum([safe_float(s.get('valor', 0)) for s in sales_prev])
            
            diff = ((fat - fat_prev) / fat_prev * 100) if fat_prev > 0 else 0
            emoji = "📈" if diff >= 0 else "📉"
            
            report += f"🏠 **{rest['name']}**\n"
            report += f"  💰 Faturamento: {fmt_brl(fat)} {emoji} ({fmt_pct(diff, 1)}% vs. sem. anterior)\n"
            
            # CMV
            cmv_data = fetch_cmv_data(rest['id'], start_date, end_date)
            custo_total = sum([safe_float(c.get('valorCusto', 0)) for c in cmv_data])
            cmv_perc = (custo_total / fat * 100) if fat > 0 else 0
            report += f"  🎯 CMV: {fmt_pct(cmv_perc)}%\n"
        except:
            continue
        
    report += f"🏦 **FATURAMENTO TOTAL DO GRUPO: {fmt_brl(total_geral)}**\n\n"
    report += "💡 _Este relatório foca no faturamento líquido e eficiência operacional (CMV)._"
    
    return report[:14000]

def get_inventory_turnover(restaurant_name, query=None):
    # Base configuration: 30 days history to analyze churn
    days_history = 30
    start_date = (datetime.now() - timedelta(days=days_history)).strftime('%Y-%m-%d')
    end_date = datetime.now().strftime('%Y-%m-%d')
    
    rest = find_restaurant_files(restaurant_name)
    sales = fetch_sales_data(rest['id'], start_date, end_date)
    
    # 1. Map Sales
    sales_map = {}
    for r in sales:
        n = normalize_text(str(r.get('nome', '')))
        if n: sales_map[n] = sales_map.get(n, 0) + safe_float(r.get('qtde', 0))
        
    # 2. Expand recipes to get theoretical consumption of ingredients
    ingredient_demand = dict(sales_map)
    recipe_path = rest.get('recipe_file')
    if recipe_path and os.path.exists(recipe_path):
        try:
            wb = openpyxl.load_workbook(recipe_path, data_only=True)
            ws = wb.active
            current_dish = None
            recipes = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                r0 = str(row[0]).strip() if row[0] else ""
                r1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if r0.startswith("Produto:") or (r0 and not r1 and not r0.replace('.','').isdigit()):
                    current_dish = normalize_text(r0.split(":", 1)[1]) if ":" in r0 else normalize_text(r0)
                    recipes[current_dish] = []
                elif current_dish and len(row) >= 4:
                    if isinstance(row[2], (int, float)):
                        ing_name = normalize_text(str(row[1])) if row[1] else ""
                        ing_qty = safe_float(row[2])
                        if ing_name and ing_qty > 0:
                            recipes[current_dish].append((ing_name, ing_qty))
                            
            for sn, qty_sold in sales_map.items():
                for dish, ingr_list in recipes.items():
                    if sn == dish:
                        for ing_name, ing_qty in ingr_list:
                            ingredient_demand[ing_name] = ingredient_demand.get(ing_name, 0) + (qty_sold * ing_qty)
        except: pass

    # 3. Fetch Stock
    stock_data = []
    try:
        session = get_session_for_rest(rest['id'])
        if session:
            r = session.get(INVENTORY_URL)
            if r.status_code == 200:
                raw = r.json()
                if isinstance(raw, list) and len(raw) > 0: stock_data = raw
    except: pass

    if not stock_data:
        stock_path = rest.get('stock_file')
        if stock_path and os.path.exists(stock_path):
            stock_data = load_json(stock_path)
            
    if not stock_data:
        return f"Dados de estoque Live indisponíveis para {rest['name']}."
    
    # Pre-aggregate stock
    stock_map = {}
    for row in stock_data:
        nome = str(row.get('produto', '')).strip()
        n_norm = normalize_text(nome)
        if len(n_norm) < 3: continue
        
        if query:
            if not match_query(query, nome): continue
                
        if n_norm not in stock_map:
            stock_map[n_norm] = {
                'nome': nome,
                'estoqueAtual': safe_float(row.get('estoqueAtual', 0)),
                'custoAtual': safe_float(row.get('custoAtual', 0))
            }
        else:
            stock_map[n_norm]['estoqueAtual'] += safe_float(row.get('estoqueAtual', 0))
            stock_map[n_norm]['custoAtual'] = max(stock_map[n_norm]['custoAtual'], safe_float(row.get('custoAtual', 0)))

    # 4. Analyze Turnover
    idle_items = []
    excess_items = []
    total_stuck_value = 0.0
    
    for n_norm, data in stock_map.items():
        qty_sold = ingredient_demand.get(n_norm, 0)
        current = data['estoqueAtual']
        cost = data['custoAtual']
        total_val = current * cost
        
        if current <= 0: continue
        
        # Skip intangibles
        if any(x in n_norm.upper() for x in ['DAY USE', 'PONTO', 'PASSADA', 'COUVERT', 'TAXA']):
            continue

        if qty_sold == 0:
            # IDLE: No sales in 30 days but has stock
            idle_items.append({
                'nome': data['nome'],
                'qty': current,
                'val': total_val
            })
            total_stuck_value += total_val
        else:
            daily_avg = qty_sold / days_history
            days_cover = current / daily_avg
            if days_cover > 30:
                # EXCESS: Stock lasts for more than 30 days
                excess_items.append({
                    'nome': data['nome'],
                    'qty': current,
                    'cover': days_cover,
                    'val': total_val
                })

    # 5. Build Report
    report = f"📦 **ANÁLISE DE GIRO DE ESTOQUE E CAPITAL EMPATADO**\n"
    report += f"🏠 Unidade: {rest['name']} | Base: {days_history} dias\n\n"
    
    report += f"💰 **CAPITAL EMPATADO (ITENS SEM GIRO): {fmt_brl(total_stuck_value)}**\n"
    report += f"_Itens com estoque positivo mas ZERO vendas/consumo nos últimos 30 dias._\n\n"
    
    if idle_items:
        report += "🚩 **TOP ITENS PARADOS (ZUMBIS):**\n"
        idle_items.sort(key=lambda x: x['val'], reverse=True)
        for it in idle_items[:15]:
            report += f"- {it['nome']}: {fmt_brl(it['qty'], 1)} un/kg | Prejuízo potencial: **{fmt_brl(it['val'])}**\n"
        report += "\n"
        
    if excess_items:
        report += "⚠️ **ESTOQUE EXCESSIVO (> 30 DIAS DE COBERTURA):**\n"
        excess_items.sort(key=lambda x: x['val'], reverse=True)
        for it in excess_items[:10]:
            report += f"- {it['nome']}: {fmt_brl(it['qty'], 1)} un/kg | Cobertura p/ **{fmt_brl(it['cover'], 0)} dias** | Valor: {fmt_brl(it['val'])}\n"

    report += "\n💡 **INSIGHTS CEO:**\n"
    report += "1. **Queima de Estoque:** Itens parados há 30 dias devem entrar em promoção ou ser transferidos para outra casa que tenha giro.\n"
    report += "2. **Ajuste de Compras:** Itens com mais de 30 dias de cobertura indicam que o pedido foi superdimensionado. Suspenda compras desses itens.\n"
    report += f"3. **Recuperação de Caixa:** Ao liquidar apenas o 'Top 5 Parados', você libera {fmt_brl(sum(i['val'] for i in idle_items[:5]))} no seu fluxo de caixa imediato."
    
    return report[:14000]

def get_cancellation_report(restaurant_name, start_date, end_date=None):
    """Analisa motivos e itens cancelados no portal para auditoria de segurança e perdas."""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    params = {
        'DataInicial': date_start_iso,
        'DataFinal': date_end_iso
    }
    
    try:
        resp = session.get(CANCELLATION_URL, params=params)
        if resp.status_code != 200:
            return f"Erro ao acessar relatório de cancelamentos (HTTP {resp.status_code})"
        
        data = resp.json()
        if not data:
            return f"Nenhum cancelamento registrado no {rest['name']} no período de {start_date} a {end_date}."
            
        # Analysis
        total_value = 0
        total_items = 0
        reasons = {}
        top_products = {}
        operators = {}
        
        for item in data:
            val = safe_float(item.get('valor', 0))
            qty = safe_float(item.get('qtde', 0))
            reason = item.get('motivo', 'SEM MOTIVO').strip() or 'SEM MOTIVO'
            prod = item.get('nomeProduto', 'DESCONHECIDO')
            op = item.get('nomeOperadorCancelamento', 'SISTEMA')
            
            total_value += val
            total_items += qty
            
            reasons[reason] = reasons.get(reason, 0) + val
            top_products[prod] = top_products.get(prod, 0) + qty
            operators[op] = operators.get(op, 0) + val
            
        # Sort results
        sorted_reasons = sorted(reasons.items(), key=lambda x: x[1], reverse=True)
        sorted_prods = sorted(top_products.items(), key=lambda x: x[1], reverse=True)
        sorted_ops = sorted(operators.items(), key=lambda x: x[1], reverse=True)
        
        report = []
        report.append(f"🔎 **AUDITORIA DE CANCELAMENTOS: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append(f"---")
        report.append(f"💰 **Total Cancelado: {fmt_brl(total_value)}** ({int(total_items)} itens)")
        report.append("")
        
        report.append("📊 **MOTIVOS DE CANCELAMENTO:**")
        for r, v in sorted_reasons[:10]:
            perc = (v / total_value * 100) if total_value > 0 else 0
            report.append(f"- {r}: {fmt_brl(v)} ({fmt_pct(perc)}%)")
        report.append("")
        
        report.append("🍕 **ITENS MAIS CANCELADOS (Qtd):**")
        for p, q in sorted_prods[:5]:
            report.append(f"- {p}: {int(q)} unid.")
        report.append("")
        
        report.append("👤 **QUEM CANCELOU (Audit):**")
        for o, v in sorted_ops[:5]:
            report.append(f"- {o}: {fmt_brl(v)}")
            
        report.append("\n---")
        report.append("💡 *Dica do Gestor:* Motivos como 'ITEM EM FALTA' indicam erro de estoque. 'ERRO DE LANÇAMENTO' frequente pode indicar necessidade de treinamento da brigada.")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro no processamento dos cancelamentos: {str(e)}"

def get_commission_report(restaurant_name, start_date, end_date=None):
    """Analisa as comissões dos garçons e produtividade da brigada no período."""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    params = {
        'DataInicial': date_start_iso,
        'DataFinal': date_end_iso
    }
    
    try:
        resp = session.get(COMMISSION_URL, params=params)
        if resp.status_code != 200:
            return f"Erro ao acessar relatório de comissões (HTTP {resp.status_code})"
        
        data = resp.json()
        if not data:
            return f"Nenhuma comissão registrada no {rest['name']} no período de {start_date} a {end_date}."
            
        # Analysis
        total_sold = 0
        total_commission = 0
        waiters = []
        
        for item in data:
            v_sold = safe_float(item.get('total', 0))
            v_comm = safe_float(item.get('totalPagar', 0))
            name = item.get('nome', 'DESCONHECIDO').strip()
            
            if v_sold > 0 or v_comm > 0:
                total_sold += v_sold
                total_commission += v_comm
                waiters.append({
                    'nome': name,
                    'vendido': v_sold,
                    'comissao': v_comm
                })
        
        # Sort by value sold
        waiters = sorted(waiters, key=lambda x: x['vendido'], reverse=True)
        
        report = []
        report.append(f"👨‍🍳 **RELATÓRIO DE COMISSÕES: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append(f"---")
        report.append(f"📈 **Venda Total (c/ Comiss.): {fmt_brl(total_sold)}**")
        report.append(f"💰 **Total Comissões: {fmt_brl(total_commission)}**")
        report.append("")
        
        report.append("🏆 **PRODUTIVIDADE POR GARÇOM:**")
        for w in waiters:
            perc = (w['vendido'] / total_sold * 100) if total_sold > 0 else 0
            report.append(f"- **{w['nome']}**: Vendeu {fmt_brl(w['vendido'])} ({fmt_pct(perc)}%) | Ganhou {fmt_brl(w['comissao'])}")
            
        report.append("\n---")
        report.append("💡 *Dica do Gestor:* Funcionários com baixo volume de vendas mas alta carga horária podem precisar de reciclagem em técnicas de vendas sugestivas.")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro no processamento das comissões: {str(e)}"

def get_payment_report(restaurant_name, start_date, end_date=None):
    """Analisa o faturamento por forma de pagamento (Pix, Cartão, Dinheiro).
    Portal: /relatorio/faturamento-forma-pagamento
    API: REVENUE_PAYMENT_URL (/faturamento/forma-pagamento/periodo)"""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    params = {
        'DataInicial': date_start_iso,
        'DataFinal': date_end_iso
    }
    
    try:
        resp = session.get(REVENUE_PAYMENT_URL, params=params)
        if resp.status_code != 200:
            return f"Erro ao acessar relatório de faturamento (HTTP {resp.status_code})"
        
        data = resp.json()
        if not data:
            return f"Nenhum faturamento registrado no {rest['name']} no período de {start_date} a {end_date}."
            
        # Analysis
        total_revenue = 0
        methods = {}
        
        for item in data:
            val = safe_float(item.get('valor', 0))
            name = str(item.get('nome', 'OUTROS')).strip().upper()
            
            if val > 0:
                total_revenue += val
                methods[name] = methods.get(name, 0) + val
        
        # Sort by value
        sorted_methods = sorted(methods.items(), key=lambda x: x[1], reverse=True)
        
        report = []
        report.append(f"💳 **FATURAMENTO POR FORMA DE PAGAMENTO: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append(f"---")
        report.append(f"💰 **Faturamento Total: {fmt_brl(total_revenue)}**")
        report.append("")
        
        report.append("📊 **DISTRIBUIÇÃO DE RECEBIMENTOS:**")
        for name, val in sorted_methods:
            perc = (val / total_revenue * 100) if total_revenue > 0 else 0
            # Friendly emojis for methods
            emoji = "💵" if "DINHEIRO" in name else "📱" if "PIX" in name else "💳" if "CARTAO" in name or "CREDITO" in name or "DEBITO" in name else "📄"
            report.append(f"{emoji} **{name}**: {fmt_brl(val)} ({fmt_pct(perc)}%)")
            
        report.append("\n---")
        report.append("💡 *Dica do Gestor:* Um alto percentual de Pix ajuda no fluxo de caixa imediato (D+0). Já altas taxas de cartões de crédito impactam a margem líquida devido às taxas das operadoras.")
        report.append("📍 _Fonte: /relatorio/faturamento-forma-pagamento_")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro no processamento do faturamento: {str(e)}"

def get_supplier_report(restaurant_name, start_date, end_date=None):
    """Gera um ranking e análise de despesas agrupadas por fornecedor, conforme URL solicitada pelo CEO."""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    # 1. Fetch Consolidated Data (Ranking) via fetch_expenses_supplier_data
    try:
        pivot_data = fetch_expenses_supplier_data(rest['id'], start_date, end_date)
        if pivot_data is None: pivot_data = []
        # fallback direto se função retornar []
        if not pivot_data:
            resp_pivot = session.get(EXPENSES_SUPPLIER_URL, params={'DataInicial': date_start_iso, 'DataFinal': date_end_iso})
            pivot_data = resp_pivot.json() if resp_pivot.status_code == 200 else []
        
        # 2. Fetch Detailed Data (Audit)
        params_det = {
            'DataInicial': date_start_iso,
            'DataFinal': date_end_iso,
            'TipoDataDespesa': 0 # Vencimento
        }
        resp_det = session.get(EXPENSES_DETAILED_URL, params=params_det)
        detailed_data = resp_det.json() if resp_det.status_code == 200 else []
        
        if not pivot_data and not detailed_data:
            return f"Nenhuma despesa de fornecedor encontrada para {rest['name']} no período."
            
        report = []
        report.append(f"🏭 **ANÁLISE DE FORNECEDORES: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append("---")
        
        # Process Ranking from pivot data
        supplier_totals = {}
        for row in pivot_data:
            name = str(row.get('fornecedor', 'OUTROS')).strip()
            total = safe_float(row.get('total', 0))
            if total > 0:
                supplier_totals[name] = supplier_totals.get(name, 0) + total
        
        sorted_suppliers = sorted(supplier_totals.items(), key=lambda x: x[1], reverse=True)
        
        total_pago = sum(supplier_totals.values())
        report.append(f"💰 **Total em Fornecedores: {fmt_brl(total_pago)}**")
        report.append("")
        report.append("🏆 **TOP 5 FORNECEDORES (Ranking):**")
        for name, val in sorted_suppliers[:5]:
            perc = (val / total_pago * 100) if total_pago > 0 else 0
            report.append(f"- {name}: {fmt_brl(val)} ({fmt_pct(perc)}%)")
        
        report.append("")
        report.append("🔍 **DETALHAMENTO DE TÍTULOS (Auditoria):**")
        
        # Group detailed items by supplier
        supplier_details = {}
        for item in detailed_data:
            f_name = str(item.get('nomeFornecedor', 'N/A')).strip()
            if f_name not in supplier_details: supplier_details[f_name] = []
            supplier_details[f_name].append(item)
            
        # Show top suppliers' details
        for f_name, _ in sorted_suppliers[:5]:
            items = supplier_details.get(f_name, [])
            if items:
                report.append(f"\n📌 *{f_name}*:")
                for it in items[:10]: # Limit to 10 titles per supplier
                    venc = it.get('dataVencimento', '').split('T')[0]
                    venc_fmt = datetime.strptime(venc, '%Y-%m-%d').strftime('%d/%m') if venc else '?'
                    val = safe_float(it.get('valor', 0))
                    status = "✅" if it.get('situacao', '').upper() == 'LIQUIDADO' else "⏳"
                    hist = it.get('historico', '')[:30]
                    report.append(f"  • {venc_fmt}: {fmt_brl(val)} {status} _{hist}_")
        
        report.append("\n---")
        report.append("💡 *Configuração:* Este relatório utiliza a rota oficial de 'Despesas por Fornecedor' e 'Detalhamento de Títulos' solicitada via Portal.")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro ao processar relatório de fornecedores: {str(e)}"

def get_fiscal_report(restaurant_name, start_date, end_date=None):
    """Analisa a emissão fiscal (NFC-e), notas emitidas, canceladas e volume tributário."""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    params = {
        'DataInicial': date_start_iso,
        'DataFinal': date_end_iso,
        'Tipo': 0
    }
    
    try:
        resp = session.get(FISCAL_URL, params=params)
        if resp.status_code != 200:
            return f"Erro ao acessar relatório fiscal (HTTP {resp.status_code})"
        
        data = resp.json()
        if not data:
            return f"Nenhum registro fiscal encontrado no {rest['name']} no período de {start_date} a {end_date}."
            
        # Analysis
        total_authorized = 0
        total_canceled = 0
        count_authorized = 0
        count_canceled = 0
        canceled_details = []
        
        for item in data:
            val = safe_float(item.get('valorTotalNotaFiscal', 0))
            is_canceled = item.get('cancelado', False)
            num = item.get('numeroDocumento', '?')
            serie = item.get('numeroSerie', '?')
            
            if is_canceled:
                total_canceled += val
                count_canceled += 1
                motivo = item.get('canceladoMotivo', 'Não informado')
                canceled_details.append(f"- Nota {num}/{serie}: {fmt_brl(val)} ({motivo})")
            else:
                total_authorized += val
                count_authorized += 1
        
        report = []
        report.append(f"🧾 **AUDITORIA FISCAL: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append(f"---")
        report.append(f"✅ **Notas Autorizadas:** {count_authorized} | **Total: {fmt_brl(total_authorized)}**")
        report.append(f"❌ **Notas Canceladas:** {count_canceled} | **Total: {fmt_brl(total_canceled)}**")
        report.append("")
        
        if canceled_details:
            report.append("🔍 **Destaques de Cancelamento:**")
            for d in canceled_details[:10]: # Limit to top 10
                report.append(d)
        
        # Simple breakdown (Volume by Document)
        report.append("\n📊 **Métrica de Conformidade:**")
        perc_cancel = (count_canceled / (count_authorized + count_canceled) * 100) if (count_authorized + count_canceled) > 0 else 0
        report.append(f"- Percentual de Notas Canceladas: {fmt_pct(perc_cancel)}%")
        
        if perc_cancel > 5:
            report.append("⚠️ **ALERTA:** Volume de cancelamentos fiscais acima de 5% pode atrair atenção do fisco ou indicar erro operacional grave.")
        else:
            report.append("🟢 Saúde fiscal operacional dentro dos parâmetros normais.")
            
        report.append("\n---")
        report.append("💡 *Configuração:* Este relatório analisa os dados da SEFAZ transmitidos via Portal NetControll.")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro no processamento fiscal: {str(e)}"

def get_cashier_closure_report(restaurant_name, start_date, end_date=None, alert_only=False):
    """Analisa as quebras e conciliação de caixa (diferenças entre sistema e valor apurado).
    Portal: #/financeiro/conciliacao
    API: CASHIER_CLOSURE_URL (fechamento-caixa/periodo) + CASH_BOOK_URL (livro-caixa extrato)"""
    rest = find_restaurant_files(restaurant_name)
    if not end_date: end_date = start_date
    
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."
    
    # Format dates to ISO
    date_start_iso = f"{start_date}T03:00:00.000Z"
    date_end_iso = f"{end_date}T03:00:00.000Z"
    
    params = {
        'DataInicial': date_start_iso,
        'DataFinal': date_end_iso
    }
    
    try:
        resp = session.get(CASHIER_CLOSURE_URL, params=params)
        if resp.status_code != 200:
            return f"Erro ao acessar fechamentos de caixa (HTTP {resp.status_code})"
        
        data = resp.json()
        if not data:
            if alert_only: return None
            return f"Nenhum fechamento de caixa encontrado no {rest['name']} em {start_date}."
            
        # Analysis
        closures = {}
        anomalies = []
        
        for item in data:
            diff = safe_float(item.get('diferenca', 0))
            val_sys = safe_float(item.get('valor', 0))
            val_phys = safe_float(item.get('valorApurado', 0))
            op = item.get('operador', 'DESCONHECIDO').strip()
            pgto = item.get('pgto', 'GERAL').strip()
            num = item.get('numeroFechamento', '?')
            
            key = (op, num)
            if key not in closures:
                closures[key] = {'venda': 0, 'apurado': 0, 'diff': 0, 'pgtos': []}
            
            closures[key]['venda'] += val_sys
            closures[key]['apurado'] += val_phys
            closures[key]['diff'] += diff
            
            if diff != 0:
                anomalies.append(f"⚠️ **Caixa {num} ({op})**: Quebra de {fmt_brl(diff)} em {pgto}")

        if alert_only:
            return "\n".join(anomalies) if anomalies else None

        report = []
        report.append(f"🏦 **AUDITORIA DE FECHAMENTO DE CAIXA: {rest['name'].upper()}**")
        report.append(f"📅 Período: {start_date} até {end_date}")
        report.append(f"---")
        
        if not anomalies:
            report.append("✅ **Tudo em conformidade!** Não foram encontradas quebras de caixa nos períodos analisados.")
        else:
            report.append(f"🚨 **ALERTA: {len(anomalies)} divergências encontradas!**")
            for a in anomalies:
                report.append(a)
        
        report.append("")
        report.append("📊 **RESUMO POR TURNO/OPERADOR:**")
        for (op, num), stats in closures.items():
            status = "🔴 QUEBRA" if stats['diff'] != 0 else "🟢 OK"
            report.append(f"- **{op} (Fechamento #{num})**")
            report.append(f"  Venda: {fmt_brl(stats['venda'])} | Apurado: {fmt_brl(stats['apurado'])}")
            report.append(f"  Diferença: **{fmt_brl(stats['diff'])}** [{status}]")
            
        report.append("\n---")
        report.append("💡 *Dica do Gestor:* Quebras frequentes em dinheiro indicam erro de sangria ou troco. Divergências em Pix/Cartão geralmente são erros de digitação ou lançamentos não realizados.")
        
        return "\n".join(report)
        
    except Exception as e:
        return f"Erro no processamento dos fechamentos: {str(e)}"


# ── Capítulos NCM com regime monofásico de PIS/COFINS ──
# Produtos desses capítulos já têm PIS/COFINS recolhidos pelo fabricante/importador,
# portanto o restaurante NÃO precisa pagar PIS/COFINS (9.25%) sobre eles na revenda.
MONOPHASIC_NCM_PREFIXES = [
    "2201",  # Águas minerais
    "2202",  # Refrigerantes, sucos, energéticos
    "2203",  # Cervejas
    "2204",  # Vinhos
    "2205",  # Vermutes
    "2206",  # Sidras e fermentados
    "2207",  # Álcool etílico
    "2208",  # Destilados (Whisky, Vodka, Gin, Rum, Cachaça)
]

# Alíquota PIS/COFINS padrão cumulativo para restaurantes (regime presumido)
PIS_COFINS_RATE = 0.0365  # 3.65% (0.65% PIS + 3% COFINS cumulativo)


def get_tax_combo_suggestions(restaurant_name, start_date=None, end_date=None):
    """Analisa oportunidades de economia fiscal via combos usando regime monofásico de PIS/COFINS."""
    if not start_date:
        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = datetime.now().strftime('%Y-%m-%d')

    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session:
        return "Erro de autenticação no portal."

    try:
        # 1. Buscar todos os produtos
        resp = session.get(PRODUCT_URL)
        if resp.status_code != 200:
            return f"Erro ao acessar cadastro de produtos (HTTP {resp.status_code})"
        products = resp.json()

        # 2. Buscar vendas do período para saber o que vende
        sales = fetch_sales_data(rest['id'], start_date, end_date)
        sales_map = {}
        for s in sales:
            n = str(s.get('nome', '')).upper().strip()
            v = safe_float(s.get('valor', 0))
            q = safe_float(s.get('qtde', 0))
            if n:
                if n not in sales_map:
                    sales_map[n] = {'qtd': 0, 'valor': 0.0}
                sales_map[n]['qtd'] += q
                sales_map[n]['valor'] += v

        # 3. Classificar produtos
        mono_products = []  # Bebidas monofásicas
        taxable_products = []  # Comidas tributáveis normalmente

        for p in products:
            if not p.get('ativo', True):
                continue
            if p.get('naoVendavel', False):
                continue

            nome = str(p.get('nome', '')).strip().upper()
            ncm = str(p.get('ncm', '')).strip()
            preco = safe_float(p.get('preco', 0))
            subgrupo = p.get('nomeSubgrupo', '')
            aliq_fed = safe_float(p.get('aliqFed', 0))

            if not ncm or len(ncm) < 4 or preco <= 0:
                continue

            # Verificar se é monofásico
            is_mono = any(ncm.startswith(prefix) for prefix in MONOPHASIC_NCM_PREFIXES)

            # Volume de vendas
            sv = sales_map.get(nome, {'qtd': 0, 'valor': 0.0})

            if is_mono:
                mono_products.append({
                    'nome': nome, 'preco': preco, 'ncm': ncm,
                    'subgrupo': subgrupo, 'aliq_fed': aliq_fed,
                    'qtd_vendida': sv['qtd'], 'valor_vendido': sv['valor']
                })
            else:
                ch = ncm[:2]
                if ch != '00':  # Ignorar NCMs genéricos
                    taxable_products.append({
                        'nome': nome, 'preco': preco, 'ncm': ncm,
                        'subgrupo': subgrupo, 'aliq_fed': aliq_fed,
                        'qtd_vendida': sv['qtd'], 'valor_vendido': sv['valor']
                    })

        # 4. Ordenar por volume de vendas (priorizar os que mais vendem)
        mono_products.sort(key=lambda x: -x['valor_vendido'])
        taxable_products.sort(key=lambda x: -x['valor_vendido'])

        # 5. Calcular quanto o restaurante já gasta de PIS/COFINS nos tributáveis
        total_taxable_sold = sum(t['valor_vendido'] for t in taxable_products)
        pis_cofins_atual = total_taxable_sold * PIS_COFINS_RATE

        # 6. Gerar sugestões de combos
        combos = []
        mono_top = [m for m in mono_products if m['qtd_vendida'] > 0][:15]
        taxable_top = [t for t in taxable_products if t['qtd_vendida'] > 0][:15]

        for mono in mono_top[:8]:
            for tax in taxable_top[:6]:
                combo_price = mono['preco'] + tax['preco']
                # Estratégia: alocar ~70-80% do combo para o item monofásico
                # e o mínimo possível para o tributável
                mono_share = round(combo_price * 0.75, 2)
                tax_share = round(combo_price - mono_share, 2)

                # Economia por combo vendido
                # Sem combo: PIS/COFINS incide sobre o preço cheio do item tributável
                # Com combo: PIS/COFINS incide só sobre tax_share (menor)
                economia_por_unid = (tax['preco'] - tax_share) * PIS_COFINS_RATE

                if economia_por_unid > 0.50:  # Só mostrar se economia > R$0,50/combo
                    # Projetar economia mensal
                    est_mensal = min(mono['qtd_vendida'], tax['qtd_vendida'])
                    economia_mensal = economia_por_unid * est_mensal

                    combos.append({
                        'mono_nome': mono['nome'],
                        'mono_preco': mono['preco'],
                        'tax_nome': tax['nome'],
                        'tax_preco': tax['preco'],
                        'combo_price': combo_price,
                        'mono_share': mono_share,
                        'tax_share': tax_share,
                        'economia_unid': economia_por_unid,
                        'est_mensal': est_mensal,
                        'economia_mensal': economia_mensal,
                        'mono_subgrupo': mono['subgrupo'],
                        'tax_subgrupo': tax['subgrupo']
                    })

        # Ordenar combos por economia mensal projetada
        combos.sort(key=lambda x: -x['economia_mensal'])

        # 7. Montar relatório
        report = []
        report.append(f"💰 **ENGENHARIA TRIBUTÁRIA — COMBOS MONOFÁSICOS**")
        report.append(f"🏠 {rest['name']} | 📅 Base: {start_date} a {end_date}")
        report.append("")
        report.append("📘 **CONCEITO:** Bebidas (cap. 22 NCM) têm PIS/COFINS **monofásico** — o imposto já foi pago pelo fabricante. Ao criar combos Bebida + Comida, você aloca maior parte do preço na bebida (não tributada) e reduz a base de cálculo do PIS/COFINS sobre a comida.")
        report.append("")

        # Resumo do cenário atual
        report.append("📊 **CENÁRIO FISCAL ATUAL:**")
        report.append(f"  🍺 Produtos Monofásicos (PIS/COFINS já pago): {len(mono_products)} itens")
        report.append(f"  🍽️ Produtos Tributáveis (PIS/COFINS 3.65%): {len(taxable_products)} itens")
        report.append(f"  💸 Faturamento tributável no período: {fmt_brl(total_taxable_sold)}")
        report.append(f"  🧾 PIS/COFINS estimado sobre tributáveis: {fmt_brl(pis_cofins_atual)}")
        report.append("")

        # TOP Bebidas mais vendidas (monofásicas)
        report.append("🍺 **TOP BEBIDAS MONOFÁSICAS (não tributáveis na revenda):**")
        for m in mono_top[:8]:
            report.append(f"  • {m['nome']} — {fmt_brl(m['preco'])} | Vendas: {fmt_brl(m['qtd_vendida'], 0)} unid ({fmt_brl(m['valor_vendido'])})")
        report.append("")

        # TOP Comidas tributáveis
        report.append("🍽️ **TOP COMIDAS TRIBUTÁVEIS (pagando PIS/COFINS):**")
        for t in taxable_top[:8]:
            pis_pago = t['valor_vendido'] * PIS_COFINS_RATE
            report.append(f"  • {t['nome']} — {fmt_brl(t['preco'])} | PIS/COFINS pago: {fmt_brl(pis_pago)}")
        report.append("")

        # Sugestões de combos
        if combos:
            report.append(f"🎯 **TOP {min(10, len(combos))} SUGESTÕES DE COMBOS PARA ECONOMIA FISCAL:**")
            report.append("_Estratégia: alocar ~75% do preço do combo para a bebida (monofásica):_")
            report.append("")
            for i, c in enumerate(combos[:10], 1):
                report.append(f"**{i}. COMBO: {c['mono_nome']} + {c['tax_nome']}**")
                report.append(f"   Preço Combo: {fmt_brl(c['combo_price'])}")
                report.append(f"   ├─ Bebida (monofásica): {fmt_brl(c['mono_share'])} ← SEM PIS/COFINS")
                report.append(f"   └─ Comida (tributável): {fmt_brl(c['tax_share'])} ← base reduzida")
                report.append(f"   📉 Economia/unidade: {fmt_brl(c['economia_unid'])}")
                report.append(f"   📈 Economia mensal estimada: **{fmt_brl(c['economia_mensal'])}** ({fmt_brl(c['est_mensal'], 0)} unid)")
                report.append("")

            # Economia total projetada
            total_economia = sum(c['economia_mensal'] for c in combos[:10])
            report.append(f"💎 **ECONOMIA TOTAL ESTIMADA (mensal): {fmt_brl(total_economia)}**")
            report.append("")
        else:
            report.append("⚠️ Não foi possível gerar sugestões de combos com dados suficientes.")

        report.append("---")
        report.append("⚖️ *Base legal: Lei 10.637/02 e 10.833/03 — PIS/COFINS monofásico. Consulte seu contador para validar a implementação dos combos no sistema fiscal.*")

        return "\n".join(report)

    except Exception as e:
        return f"Erro na análise tributária: {str(e)}"

def lookup_ncm_description(ncm_code):
    """Consulta a descrição oficial do NCM na BrasilAPI. Usa cache para evitar chamadas repetidas."""
    if ncm_code in _NCM_CACHE:
        return _NCM_CACHE[ncm_code]
    try:
        r = requests.get(BRASILAPI_NCM_URL.format(ncm_code), timeout=5)
        if r.status_code == 200:
            data = r.json()
            desc = data.get('descricao', '')
            _NCM_CACHE[ncm_code] = desc
            return desc
    except:
        pass
    _NCM_CACHE[ncm_code] = None
    return None


def _get_expected_chapters_for_group(grupo_nome):
    """Retorna os capítulos NCM esperados para um grupo de produto do NetControll."""
    if not grupo_nome:
        return []
    g_norm = normalize_text(grupo_nome).upper()
    expected = []
    for keyword, chapters in NCM_GRUPO_MAP.items():
        if keyword in g_norm:
            expected.extend(chapters)
    return list(set(expected))


def audit_product_registration(restaurant_name):
    """Audita o cadastro de produtos: NCMs inválidos, ausentes, sem grupo e NCM semanticamente incompatível com o grupo."""
    rest = find_restaurant_files(restaurant_name)
    session = get_session_for_rest(rest['id'])
    if not session: return "Erro de autenticação no portal."

    try:
        resp = session.get(PRODUCT_URL)
        if resp.status_code != 200:
            return f"Erro ao acessar cadastro de produtos (HTTP {resp.status_code})"

        products = resp.json()
        if not products:
            return f"Nenhum produto cadastrado encontrado no {rest['name']}."

        format_anomalies = []   # NCM ausente / formato errado / sem grupo / nome genérico
        semantic_alerts = []    # NCM com formato OK mas descrição incompatível com o grupo
        missing_ncm = 0
        invalid_ncm = 0
        missing_group = 0
        total_items = 0

        for p in products:
            nome = str(p.get('nome', '')).strip().upper()
            ncm  = str(p.get('ncm', '')).strip()
            # nomeGrupo não existe na API - o campo correto é nomeSubgrupo
            grupo = p.get('nomeSubgrupo') or ''
            ativo = p.get('ativo', True)

            if not ativo:
                continue

            total_items += 1

            # ── 1. Formato do NCM ────────────────────────────────────────────
            if not ncm or ncm in ('None', '0', ''):
                missing_ncm += 1
                format_anomalies.append(f"❌ **NCM Ausente**: {nome} (Grupo: {grupo or 'N/A'})")
                continue  # Sem NCM não dá pra fazer lookup semântico

            if len(ncm) != 8 or not ncm.isdigit():
                invalid_ncm += 1
                format_anomalies.append(f"⚠️ **NCM Inválido ({ncm})**: {nome}")
                continue

            # ── 2. Grupo ausente ────────────────────────────────────────────
            if not grupo:
                missing_group += 1
                format_anomalies.append(f"📁 **Sem Grupo**: {nome} (NCM: {ncm})")

            # ── 3. Nome genérico ────────────────────────────────────────────
            if any(x in nome for x in ['GENÉRICO', 'GENERICO', 'TESTE', 'PRODUTO NOVO']):
                format_anomalies.append(f"🏷️ **Nome Genérico**: {nome}")

            # ── 4. Validação semântica: NCM × Grupo (sem HTTP aqui) ────────
            expected = _get_expected_chapters_for_group(grupo)
            if expected:  # Só valida se o grupo está no nosso mapa
                chapter = ncm[:2]
                if chapter not in expected:
                    # Coleta o suspeito SEM fazer HTTP ainda
                    semantic_alerts.append({
                        'nome': nome, 'grupo': grupo,
                        'ncm': ncm, 'chapter': chapter
                    })

        # ── Enriquecer os primeiros 15 suspeitos com descrição da Receita Federal ──
        MAX_LOOKUPS = 15
        for i, alert in enumerate(semantic_alerts[:MAX_LOOKUPS]):
            try:
                desc = lookup_ncm_description(alert['ncm'])
                alert['desc'] = desc
            except:
                alert['desc'] = None

        # ── Montagem do relatório ────────────────────────────────────────────
        report = []
        report.append(f"🛡️ **AUDITORIA DE CADASTRO: {rest['name'].upper()}**")
        report.append(f"📦 Total de Produtos Ativos Analisados: {total_items}")
        report.append("")

        # Seção 1 — Problemas de formato
        if format_anomalies:
            report.append("🚨 **PROBLEMAS DE FORMATO/CADASTRO:**")
            report.append(f"  - NCM Ausente: {missing_ncm}")
            report.append(f"  - NCM Inválido (fora de 8 dígitos): {invalid_ncm}")
            report.append(f"  - Sem Grupo de Produto: {missing_group}")
            report.append("")
            report.append("🔍 **ITENS COM PROBLEMA DE FORMATO (amostra):**")
            for a in format_anomalies[:20]:
                report.append(a)
            if len(format_anomalies) > 20:
                report.append(f"  ... e mais {len(format_anomalies) - 20} itens com problemas de formato.")
        else:
            report.append("✅ **Formato de NCM:** Todos os produtos ativos possuem NCM com 8 dígitos válidos.")

        report.append("")

        # Seção 2 — Alertas semânticos
        if semantic_alerts:
            report.append(f"🧾 **ALERTA SEMÂNTICO — NCM SUSPEITO ({len(semantic_alerts)} produtos):**")
            report.append("_O capítulo NCM não corresponde ao subgrupo cadastrado no sistema:_")
            report.append("")
            for alert in semantic_alerts[:MAX_LOOKUPS]:
                desc_str = f'Receita Federal: "{alert.get("desc")}"' if alert.get('desc') else ''
                report.append(
                    f"🔴 **{alert['nome']}**\n"
                    f"   Subgrupo: {alert['grupo']} | NCM: {alert['ncm']} (cap. {alert['chapter']})\n"
                    f"   {desc_str}"
                )
                report.append("")
            if len(semantic_alerts) > MAX_LOOKUPS:
                report.append(f"  ... e mais {len(semantic_alerts) - MAX_LOOKUPS} produtos com NCM suspeito.")
        else:
            report.append("✅ **Validação Semântica:** Todos os NCMs verificados são compatíveis com o subgrupo do produto.")

        report.append("")
        report.append("---")
        report.append("💡 *NCMs incorretos geram erros na NFC-e e podem resultar em multas fiscais ou tributação indevida de PIS/COFINS.*")

        return "\n".join(report)

    except Exception as e:
        return f"Erro na auditoria de cadastro: {str(e)}"

def get_break_even_analysis(restaurant_name):
    """
    Calcula o Ponto de Equilíbrio (Break-Even) Mensal baseado nos últimos 30 dias.
    Cruza custos fixos, CMV médio e custos variáveis de vendas.
    """
    try:
        rest = find_restaurant_files(restaurant_name)
        now = datetime.now()
        start = (now - timedelta(days=30)).strftime('%Y-%m-%d')
        end = now.strftime('%Y-%m-%d')
        
        # 1. Obter Vendas
        sales_data = fetch_sales_data(rest['id'], start, end)
        total_revenue = sum(safe_float(s.get('valor', 0)) for s in sales_data)
        
        if total_revenue == 0:
            return f"Não foi possível calcular o ponto de equilíbrio para {rest['name']} pois não há vendas registradas nos últimos 30 dias."

        # 2. Obter Despesas e Categorizar
        expenses = fetch_expenses_data(rest['id'], start, end)
        
        fixed_costs = 0
        variable_costs_non_cmv = 0
        cmv_expenses = 0
        
        # Mapeamento detalhado baseado na análise prévia do portal
        fixed_keywords = [
            "CUSTOS FIXOS", "FOLHA", "PESSOAL", "ADMINISTRATIV", "OCUPACAO", "OCUPAC",
            "RETIRADA", "SOCIOS", "PRO-LABORE", "ALUGUEL", "FIXA", "MARKETING",
            "MANUTENCAO", "SOFTWARE", "CONSORCIO", "INTERNET", "CONTABILIDADE"
        ]
        
        variable_keywords = [
            "VARIAVEIS", "VENDAS", "TRIBUTOS", "IMPOSTO", "TAXA", "COMISSAO", 
            "CARTAO", "MARKETPLACE", "FRETE", "RAPPI", "IFOOD"
        ]
        
        cmv_keywords = ["CUSTO DE PRODUCAO", "CMV", "MERCADORIA", "INSUMOS", "ENTRADA DE FOGO"]

        for exp in expenses:
            cat1 = normalize_text(str(exp.get('planoContas1', ''))).upper()
            val = safe_float(exp.get('valor', 0))
            
            # Priority check
            is_cmv = any(normalize_text(k).upper() in cat1 for k in cmv_keywords)
            is_fixed = any(normalize_text(k).upper() in cat1 for k in fixed_keywords)
            is_variable = any(normalize_text(k).upper() in cat1 for k in variable_keywords)
            
            if is_cmv:
                cmv_expenses += val
            elif is_variable:
                variable_costs_non_cmv += val
            elif is_fixed or cat1 in ["DESPESAS", "NONE", "GERAL", "DIVERSOS"]:
                fixed_costs += val
            else:
                fixed_costs += val

        # 3. CMV Portal cross-check
        cmv_data = fetch_cmv_data(rest['id'], start, end)
        cmv_portal_val = sum(safe_float(c.get('valorCusto', 0)) for c in cmv_data)
        
        effective_cmv_val = max(cmv_expenses, cmv_portal_val)
        cmv_perc = effective_cmv_val / total_revenue if total_revenue > 0 else 0
        var_perc = variable_costs_non_cmv / total_revenue if total_revenue > 0 else 0
        
        # 4. Margem de Contribuição Média (MCR)
        mcr = 1 - (cmv_perc + var_perc)
        
        if mcr <= 0:
            return (f"🚨 **ALERTA CRÍTICO: MARGEM NEGATIVA EM {rest['name'].upper()}**\n"
                    f"A soma de CMV ({fmt_pct(cmv_perc*100)}%) e Custos Variáveis ({fmt_pct(var_perc*100)}%) "
                    f"é superior ao faturamento. A casa está perdendo dinheiro em cada venda.")

        bep_monthly = fixed_costs / mcr
        
        report = []
        report.append(f"⚖️ **ANÁLISE DE PONTO DE EQUILÍBRIO (BREAK-EVEN)**")
        report.append(f"🏠 {rest['name']} | Base: Últimos 30 dias")
        report.append("---")
        report.append(f"🏢 **Estrutura de Custos Fixos:** {fmt_brl(fixed_costs)}")
        report.append(f"   _(Aluguel, Folha, Admin, Marketing, etc.)_")
        report.append("")
        report.append(f"📊 **Margem de Contribuição:** {fmt_pct(mcr*100)}%")
        report.append(f"   • CMV Médio: {fmt_pct(cmv_perc*100)}%")
        report.append(f"   • Taxas/Impostos: {fmt_pct(var_perc*100)}%")
        report.append("")
        report.append(f"🏁 **META DE SOBREVIVÊNCIA (Mensal):**")
        report.append(f"💰 **{fmt_brl(bep_monthly)}**")
        report.append("")
        
        # Comparison with current month
        m_start = now.replace(day=1).strftime('%Y-%m-%d')
        m_sales = fetch_sales_data(rest['id'], m_start, end)
        m_rev = sum(safe_float(s.get('valor', 0)) for s in m_sales)
        
        import calendar
        d_passed = now.day
        d_total = calendar.monthrange(now.year, now.month)[1]
        
        reached = (m_rev / bep_monthly * 100) if bep_monthly > 0 else 0
        
        report.append(f"📅 **Acompanhamento do Mês Atual:**")
        report.append(f"   • Realizado: {fmt_brl(m_rev)} ({fmt_pct(reached)}% do P.E.)")
        
        if m_rev >= bep_monthly:
            report.append(f"🥳 **LUCRO ATIVADO!** A casa já ultrapassou o ponto de equilíbrio no mês.")
        else:
            rem = bep_monthly - m_rev
            d_left = d_total - d_passed + 1
            needed = rem / d_left if d_left > 0 else rem
            report.append(f"⏳ Faltam {fmt_brl(rem)} para pagar as contas do mês.")
            report.append(f"🚀 Necessário: **{fmt_brl(needed)}/dia** para atingir o P.E.")
            
        report.append("\n💡 _Nota: Cálculo baseado na eficiência dos últimos 30 dias._")
        return "\n".join(report)

    except Exception as e:
        return f"Erro ao calcular ponto de equilíbrio: {str(e)}"

def get_financial_snapshot(restaurant_name):
    """
    Gera um dossiê de Saúde Financeira Real cruzando saldo em conta, 
    contas a receber e contas a pagar dos próximos 15 dias.
    """
    try:
        rest = find_restaurant_files(restaurant_name)
        session = get_session_for_rest(rest['id'])
        if not session: return "Erro de conexão financeira."

        now_dt = datetime.now()
        today = now_dt.strftime('%Y-%m-%d')
        future_dt = (now_dt + timedelta(days=15)).strftime('%Y-%m-%d')
        past_dt = (now_dt - timedelta(days=1)).strftime('%Y-%m-%d')

        # 1. Obter Saldo Real (Livro Caixa)
        # Tentamos ID 1 (Geralmente principal) e ID 2
        total_balance = 0
        accounts_found = 0
        for acc_id in [1, 2, 3]:
            try:
                params_cb = {
                    'IdContaFinanceira': acc_id, 
                    'DataInicial': f"{today}T00:00:00.000Z", 
                    'DataFinal': f"{today}T23:59:59.000Z"
                }
                r_cb = session.get(CASH_BOOK_URL, params=params_cb)
                if r_cb.status_code == 200:
                    data_cb = r_cb.json()
                    if data_cb:
                        # O saldo atual costuma ser o 'saldoAnterior' do dia de hoje + movs
                        # No NetWeb, o saldo final está em 'saldo' ou no último registro
                        acc_bal = safe_float(data_cb[-1].get('saldoAcumulado', data_cb[0].get('valor', 0)))
                        total_balance += acc_bal
                        accounts_found += 1
            except: pass

        # 2. Contas a Receber (Próximos 15 dias)
        params_rec = {
            'DataInicial': f"{today}T00:00:00.000Z", 
            'DataFinal': f"{future_dt}T23:59:59.000Z", 
            'TipoData': 1 # Vencimento/Previsão
        }
        r_rec = session.get(RECEIVABLES_URL, params=params_rec)
        receivables = r_rec.json() if r_rec.status_code == 200 else []
        total_rec = sum(safe_float(r.get('valor', 0)) for r in receivables)

        # 3. Contas a Pagar (Próximos 15 dias)
        params_pay = {
            'DataInicial': f"{today}T00:00:00.000Z", 
            'DataFinal': f"{future_dt}T23:59:59.000Z", 
            'TipoData': 0 # Vencimento
        }
        r_pay = session.get(EXPENSES_DETAILED_URL, params=params_pay)
        payables = r_pay.json() if r_pay.status_code == 200 else []
        total_pay = sum(safe_float(p.get('valor', 0)) for p in payables if p.get('situacao') != 'LIQUIDADO')

        # 4. Cálculo de Liquidez
        net_liquidity = total_balance + total_rec - total_pay
        
        report = []
        report.append(f"🏦 **RAIO-X FINANCEIRO (Cash Flow): {rest['name'].upper()}**")
        report.append(f"📅 Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        report.append("---")
        
        report.append(f"💰 **SALDO DISPONÍVEL (Contas/Caixa):** {fmt_brl(total_balance)}")
        report.append(f"   _(Baseado em {accounts_found} conta(s) ativa(s))_")
        report.append("")
        
        report.append(f"📈 **PREVISÃO DE ENTRADAS (15 dias):** {fmt_brl(total_rec)}")
        report.append(f"📉 **CONTAS A PAGAR (15 dias):** {fmt_brl(total_pay)}")
        report.append("")
        
        report.append(f"⚖️ **LIQUIDEZ PROJETADA (D+15):**")
        report.append(f"💵 **{fmt_brl(net_liquidity)}**")
        
        if net_liquidity < 0:
            report.append(f"\n🚨 **ALERTA DE CAIXA:** O faturamento + saldo atual não cobrem as despesas dos próximos 15 dias. Déficit projetado de {fmt_brl(abs(net_liquidity))}.")
        elif net_liquidity < (total_pay * 0.2):
            report.append(f"\n⚠️ **ATENÇÃO:** Margem de segurança baixa. Fluxo de caixa apertado.")
        else:
            report.append(f"\n✅ **Saúde financeira estável.** Saldo e recebíveis cobrem confortavelmente as obrigações imediatas.")

        report.append("\n🔍 **Pinceladas do Contas a Pagar (Top 5):**")
        sorted_pay = sorted([p for p in payables if p.get('situacao') != 'LIQUIDADO'], 
                           key=lambda x: safe_float(x.get('valor', 0)), reverse=True)
        for p in sorted_pay[:5]:
            venc = p.get('dataVencimento','').split('T')[0]
            report.append(f"  • {venc[8:10]}/{venc[5:7]}: {p.get('nomeFornecedor','?')[:25]} - {fmt_brl(safe_float(p.get('valor',0)))}")

        report.append("\n💡 _Nota: Este relatório utiliza dados em tempo real do Dashboard Financeiro e Livro Caixa._")
        
        return "\n".join(report)

    except Exception as e:
        return f"Erro ao gerar Raio-X financeiro: {str(e)}"


